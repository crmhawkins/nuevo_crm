"""
Servidor para procesamiento de justificaciones SEO
Recibe peticiones del CRM, ejecuta análisis SEO completo y devuelve archivos generados
"""

from flask import Flask, request, jsonify
import threading
import os
import sys
import time
import shutil
import requests
import json
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import pyautogui
from openpyxl import load_workbook
from urllib.parse import urlparse, urljoin
from app import generar_pdf_con_capturas
import imaplib
import email
import re
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from PIL import Image
from queue import Queue
from collections import namedtuple

app = Flask(__name__)

# Configuración
CALLBACK_TIMEOUT = 30  # Segundos para timeout al enviar archivos al CRM

# Cola de trabajos
cola_trabajos = Queue()
procesando = False
trabajo_actual = None

# Estructura de trabajo
Trabajo = namedtuple('Trabajo', ['url', 'justificacion_id', 'callback_url', 'user_name', 'tipo_analisis', 'timestamp_recibido'])


def obtener_directorio_base():
    """Obtiene el directorio base de la aplicación."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


def log_mensaje(mensaje):
    """Log de mensajes con timestamp"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"[{timestamp}] {mensaje}")


def consultar_ia_local(prompt):
    """Consulta la IA local para análisis de keywords"""
    try:
        url = "https://aiapi.hawkins.es/chat/chat"
        headers = {
            "x-api-key": "OllamaAPI_2024_K8mN9pQ2rS5tU7vW3xY6zA1bC4eF8hJ0lM",
            "Content-Type": "application/json"
        }
        
        payload = {
            "modelo": "gpt-oss:120b-cloud",
            "prompt": prompt
        }
        
        log_mensaje(f"🔗 Enviando petición a: {url}")
        log_mensaje(f"⚠️  Deshabilitando verificación SSL para API externa")
        log_mensaje(f"📏 Tamaño del prompt: {len(prompt)} caracteres")
        
        response = requests.post(url, headers=headers, json=payload, timeout=30, verify=False)
        
        log_mensaje(f"📊 Status Code: {response.status_code}")
        
        if response.status_code == 200:
            try:
                data = response.json()
                log_mensaje(f"✅ Respuesta JSON recibida")
                
                if data.get('success') and data.get('respuesta'):
                    log_mensaje(f"🎯 Keywords extraídas por IA: {data['respuesta']}")
                    return data['respuesta']
                else:
                    log_mensaje(f"❌ Respuesta sin éxito")
                    return None
            except json.JSONDecodeError as e:
                log_mensaje(f"❌ Error parseando JSON: {str(e)}")
                return None
        else:
            log_mensaje(f"❌ Error HTTP {response.status_code}")
            return None
            
    except requests.exceptions.Timeout:
        log_mensaje("⏰ Timeout en la petición a la IA local")
        return None
    except Exception as e:
        log_mensaje(f"💥 Error inesperado consultando IA local: {str(e)}")
        return None


def extraer_datos_web(driver, url):
    """Extrae datos estructurados de la página web para análisis"""
    try:
        log_mensaje(f"📊 Extrayendo datos de la página: {url}")
        
        title = ""
        meta_description = ""
        h1_tags = []
        h2_tags = []
        h3_tags = []
        content_sample = ""
        
        try:
            title = driver.find_element(By.TAG_NAME, 'title').text
            log_mensaje(f"✅ Título extraído: {title}")
        except:
            pass
            
        try:
            meta_desc_element = driver.find_element(By.CSS_SELECTOR, 'meta[name="description"]')
            meta_description = meta_desc_element.get_attribute('content') or ""
            log_mensaje(f"✅ Meta descripción extraída: {meta_description[:100]}...")
        except:
            pass
            
        try:
            h1_elements = driver.find_elements(By.TAG_NAME, 'h1')
            h1_tags = [h.text.strip() for h in h1_elements if h.text.strip()]
            log_mensaje(f"✅ H1 tags extraídos: {len(h1_tags)} elementos")
        except:
            pass
            
        try:
            h2_elements = driver.find_elements(By.TAG_NAME, 'h2')
            h2_tags = [h.text.strip() for h in h2_elements if h.text.strip()]
            log_mensaje(f"✅ H2 tags extraídos: {len(h2_tags)} elementos")
        except:
            pass
            
        try:
            h3_elements = driver.find_elements(By.TAG_NAME, 'h3')
            h3_tags = [h.text.strip() for h in h3_elements if h.text.strip()]
            log_mensaje(f"✅ H3 tags extraídos: {len(h3_tags)} elementos")
        except:
            pass
            
        try:
            body_element = driver.find_element(By.TAG_NAME, 'body')
            content_sample = body_element.text[:2000]
            log_mensaje(f"✅ Contenido extraído: {len(content_sample)} caracteres")
        except:
            pass
        
        domain = urlparse(url).netloc.replace('www.', '')
        url_words = []
        try:
            url_path = urlparse(url).path
            url_words = [word for word in url_path.split('/') if word and len(word) > 2]
            log_mensaje(f"✅ Palabras URL extraídas: {url_words}")
        except:
            pass
        
        datos = {
            'url': url,
            'title': title,
            'meta_description': meta_description,
            'h1_tags': h1_tags,
            'h2_tags': h2_tags,
            'h3_tags': h3_tags,
            'domain': domain,
            'url_words': url_words,
            'content_sample': content_sample
        }
        
        log_mensaje(f"✅ Datos completos extraídos exitosamente")
        return datos
        
    except Exception as e:
        log_mensaje(f"💥 Error extrayendo datos web: {str(e)}")
        return None


def crear_prompt_keywords(datos_web):
    """Crea el prompt optimizado para la IA local"""
    prompt = f"""Analiza estos datos de una página web y extrae exactamente las 5 keywords más relevantes para SEO, separadas por comas.

DATOS DE LA PÁGINA WEB:
- URL: {datos_web['url']}
- Título: {datos_web['title']}
- Descripción Meta: {datos_web['meta_description']}
- Encabezados H1: {', '.join(datos_web['h1_tags'])}
- Encabezados H2: {', '.join(datos_web['h2_tags'])}
- Encabezados H3: {', '.join(datos_web['h3_tags'])}
- Dominio: {datos_web['domain']}
- Palabras de URL: {', '.join(datos_web['url_words'])}
- Muestra de contenido: {datos_web['content_sample'][:500]}...

INSTRUCCIONES:
1. Identifica las palabras clave más importantes para SEO
2. Prioriza términos del título y encabezados
3. Incluye palabras relevantes del dominio/URL
4. Evita palabras muy genéricas como "página", "web", "sitio"
5. Máximo 5 keywords
6. Respuesta SOLO con las keywords separadas por comas, sin explicaciones

FORMATO DE RESPUESTA: keyword1, keyword2, keyword3, keyword4, keyword5"""
    
    return prompt


def extraer_keywords_automaticas(driver, url):
    """Sistema híbrido para extraer keywords automáticamente"""
    try:
        log_mensaje(f"🔍 Iniciando extracción automática de keywords de: {url}")
        
        datos_web = extraer_datos_web(driver, url)
        if not datos_web:
            log_mensaje("❌ No se pudieron extraer datos de la página web")
            return None
        
        log_mensaje(f"✅ Datos extraídos - Título: {datos_web['title']}")
        
        prompt = crear_prompt_keywords(datos_web)
        log_mensaje(f"✅ Prompt creado: {len(prompt)} caracteres")
        
        respuesta_ia = consultar_ia_local(prompt)
        
        if respuesta_ia:
            log_mensaje(f"✅ Respuesta recibida de IA: {respuesta_ia}")
            keywords = procesar_respuesta_ia(respuesta_ia)
            if keywords and len(keywords) > 0:
                log_mensaje(f"✅ Keywords procesadas exitosamente: {keywords}")
                return keywords
            else:
                log_mensaje("❌ No se pudieron procesar las keywords de la IA")
        else:
            log_mensaje("❌ No se recibió respuesta de la IA")
        
        log_mensaje("⚠️ Ejecutando fallback a análisis básico...")
        keywords_fallback = analisis_basico_keywords(datos_web)
        if keywords_fallback:
            log_mensaje(f"✅ Keywords fallback extraídas: {keywords_fallback}")
            return keywords_fallback
        else:
            log_mensaje("❌ Fallback también falló")
            
        log_mensaje("💥 Todos los métodos de extracción fallaron")
        return None
        
    except Exception as e:
        log_mensaje(f"💥 Error inesperado en extracción automática: {str(e)}")
        return None


def procesar_respuesta_ia(respuesta):
    """Procesa la respuesta de la IA local"""
    try:
        log_mensaje(f"🔄 Procesando respuesta de IA: '{respuesta}'")
        
        respuesta_limpia = respuesta.strip()
        keywords = [kw.strip() for kw in respuesta_limpia.split(',')]
        
        keywords_validas = []
        stop_words = ['página', 'web', 'sitio', 'home', 'inicio']
        
        for kw in keywords:
            kw = kw.strip()
            
            if len(kw) <= 2 or len(kw) >= 50:
                continue
            if kw.lower() in stop_words:
                continue
                
            keywords_validas.append(kw)
        
        resultado = keywords_validas[:5]
        log_mensaje(f"🎯 Resultado final (máximo 5): {resultado}")
        
        return resultado
        
    except Exception as e:
        log_mensaje(f"💥 Error procesando respuesta IA: {str(e)}")
        return None


def analisis_basico_keywords(datos_web):
    """Análisis básico de keywords como fallback"""
    try:
        keywords = []
        
        if datos_web['title']:
            title_words = re.findall(r'\b\w+\b', datos_web['title'].lower())
            keywords.extend([w for w in title_words if len(w) > 3])
        
        for h1 in datos_web['h1_tags']:
            h1_words = re.findall(r'\b\w+\b', h1.lower())
            keywords.extend([w for w in h1_words if len(w) > 3])
        
        for h2 in datos_web['h2_tags']:
            h2_words = re.findall(r'\b\w+\b', h2.lower())
            keywords.extend([w for w in h2_words if len(w) > 3])
        
        domain_words = re.findall(r'\b\w+\b', datos_web['domain'].lower())
        keywords.extend([w for w in domain_words if len(w) > 3])
        
        from collections import Counter
        word_count = Counter(keywords)
        
        stop_words = {'página', 'web', 'sitio', 'home', 'inicio', 'contenido', 'información', 'servicios', 'productos'}
        filtered_words = {word: count for word, count in word_count.items() if word not in stop_words}
        
        top_keywords = [word for word, count in Counter(filtered_words).most_common(5)]
        
        return top_keywords
        
    except Exception as e:
        log_mensaje(f"Error análisis básico: {str(e)}")
        return None


def buscar_urls_en_pagina(driver, url_base):
    """Busca las URLs de 'Sobre Nosotros' y 'Contacto' en la página web"""
    sobre_nosotros_url = None
    contacto_url = None
    
    try:
        enlaces = driver.find_elements(By.TAG_NAME, "a")
        
        for enlace in enlaces:
            try:
                href = enlace.get_attribute("href")
                texto = enlace.text.lower().strip()
                
                if not href:
                    continue
                
                if not sobre_nosotros_url and any(p in texto for p in ["sobre nosotros", "sobre", "quiénes somos", "quienes somos", "about", "acerca de"]):
                    sobre_nosotros_url = href
                
                if not contacto_url and any(p in texto for p in ["contacto", "contáctanos", "contactanos", "contact"]):
                    contacto_url = href
                
                if sobre_nosotros_url and contacto_url:
                    break
            except:
                continue
        
        if not sobre_nosotros_url or not contacto_url:
            for enlace in enlaces:
                try:
                    href = enlace.get_attribute("href")
                    if not href:
                        continue
                    
                    href_lower = href.lower()
                    
                    if not sobre_nosotros_url and any(p in href_lower for p in ["sobre", "about", "acerca", "quienes-somos"]):
                        sobre_nosotros_url = href
                    
                    if not contacto_url and any(p in href_lower for p in ["contacto", "contact"]):
                        contacto_url = href
                    
                    if sobre_nosotros_url and contacto_url:
                        break
                except:
                    continue
    
    except Exception as e:
        log_mensaje(f"Error buscando URLs: {str(e)}")
    
    if not sobre_nosotros_url:
        sobre_nosotros_url = urljoin(url_base, "/sobre-nosotros")
    if not contacto_url:
        contacto_url = urljoin(url_base, "/contacto")
    
    return sobre_nosotros_url, contacto_url


def modificar_excel_informe(url_web, nombre_web=None):
    """Crea una copia del archivo Excel y la modifica con los datos de la URL."""
    try:
        base_dir = obtener_directorio_base()
        archivo_excel_original = os.path.join(base_dir, 'Informe_de_revision.xlsx')
        
        if not os.path.exists(archivo_excel_original):
            log_mensaje(f"Archivo Excel no encontrado: {archivo_excel_original}")
            return None
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_copia = f'Informe_de_revision_{timestamp}.xlsx'
        archivo_excel_copia = os.path.join(base_dir, nombre_copia)
        
        shutil.copy2(archivo_excel_original, archivo_excel_copia)
        wb = load_workbook(archivo_excel_copia)
        hojas = wb.worksheets
        
        if len(hojas) < 4:
            log_mensaje("El archivo Excel no tiene suficientes hojas")
            return None
        
        pagina_2 = hojas[1]
        
        if not nombre_web:
            nombre_web = "Sitio Web"
        
        fecha_hoy = datetime.now().strftime('%d/%m/%Y')
        
        pagina_2['C15'] = nombre_web
        pagina_2['C17'] = url_web
        pagina_2['C21'] = fecha_hoy
        wb.save(archivo_excel_copia)
        return archivo_excel_copia
        
    except Exception as e:
        log_mensaje(f"Error modificando Excel: {str(e)}")
        return None


def modificar_excel_urls_adicionales(archivo_excel_copia, url_base, sobre_nosotros_url, contacto_url):
    """Modifica la página 4 del Excel copiado con las URLs adicionales y slugs personalizados."""
    try:
        if not archivo_excel_copia or not os.path.exists(archivo_excel_copia):
            return False
        
        wb = load_workbook(archivo_excel_copia)
        hojas = wb.worksheets
        
        if len(hojas) < 4:
            return False
        
        pagina_4 = hojas[3]
        
        pagina_4['F8'] = url_base
        pagina_4['F9'] = url_base
        pagina_4['F10'] = sobre_nosotros_url
        pagina_4['F11'] = contacto_url
        
        def generar_slug_personalizado(url):
            if not url:
                return ""
            
            parsed_url = urlparse(url)
            path = parsed_url.path.lower().strip('/')
            
            slug_mapping = {
                'sobre-nosotros': 'Sobre Nosotros',
                'sobre': 'Sobre Nosotros',
                'nosotros': 'Nosotros',
                'sobre-mi': 'Sobre Mí',
                'quienes-somos': 'Quiénes Somos',
                'acerca-de': 'Acerca de',
                'about': 'Sobre Nosotros',
                'contacto': 'Contacto',
                'contactanos': 'Contáctanos',
                'contact': 'Contacto',
                'contacto.html': 'Contacto',
                'contactanos.html': 'Contáctanos'
            }
            
            if path in slug_mapping:
                return slug_mapping[path]
            
            for key, value in slug_mapping.items():
                if key in path:
                    return value
            
            if path:
                return path.replace('-', ' ').replace('_', ' ').title()
            
            return "Página"
        
        slug_sobre_nosotros = generar_slug_personalizado(sobre_nosotros_url)
        slug_contacto = generar_slug_personalizado(contacto_url)
        
        def determinar_tipo_pagina(url):
            if not url:
                return "Otras páginas"
            
            parsed_url = urlparse(url)
            path = parsed_url.path.lower().strip('/')
            
            if any(keyword in path for keyword in ['contacto', 'contact', 'contactanos']):
                return "Contacto"
            elif any(keyword in path for keyword in ['sobre', 'nosotros', 'about', 'acerca', 'quienes']):
                return "Acerca de"
            else:
                return "Otras páginas"
        
        tipo_sobre_nosotros = determinar_tipo_pagina(sobre_nosotros_url)
        tipo_contacto = determinar_tipo_pagina(contacto_url)
        
        def generar_breadcrumbs(slug):
            return f"Inicio > {slug}"
        
        pagina_4['C10'] = "Página Web"
        pagina_4['E10'] = tipo_sobre_nosotros
        pagina_4['G10'] = generar_breadcrumbs(slug_sobre_nosotros)
        
        pagina_4['C11'] = "Página Web"
        pagina_4['E11'] = tipo_contacto
        pagina_4['G11'] = generar_breadcrumbs(slug_contacto)
        
        wb.save(archivo_excel_copia)
        return True
    
    except Exception as e:
        log_mensaje(f"Error modificando Excel (URLs adicionales): {str(e)}")
        return False


def normalizar_url(url):
    """Normaliza la URL agregando https:// si no lo tiene y quitando slash final"""
    if not url:
        return ""
    
    url = url.strip()
    
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    
    if url.endswith('/'):
        url = url[:-1]
    
    return url


def obtener_codigo_verificacion_email():
    """Obtiene el código de verificación del último email de Dondominio"""
    try:
        log_mensaje("  📧 Conectando a servidor de email...")
        
        mail = imaplib.IMAP4_SSL('imap.ionos.es', 993)
        mail.login('dondominio@hawkins.es', 'R4t4-2025')
        mail.select('inbox')
        
        log_mensaje("  📧 Buscando emails de Dondominio...")
        
        status, messages = mail.search(None, 'FROM', 'info@dondominio.com')
        email_ids = messages[0].split()
        
        if not email_ids:
            log_mensaje("  ❌ No se encontraron emails de Dondominio")
            return None
        
        latest_email_id = email_ids[-1]
        status, msg_data = mail.fetch(latest_email_id, '(RFC822)')
        email_body = msg_data[0][1]
        email_message = email.message_from_bytes(email_body)
        
        email_content = ""
        if email_message.is_multipart():
            for part in email_message.walk():
                if part.get_content_type() == "text/plain":
                    email_content = part.get_payload(decode=True).decode()
                    break
        else:
            email_content = email_message.get_payload(decode=True).decode()
        
        log_mensaje("  📧 Email obtenido, extrayendo código...")
        
        codigo_pattern = r'\b(\d{6})\b'
        matches = re.findall(codigo_pattern, email_content)
        
        if matches:
            codigo = matches[-1]
            log_mensaje(f"  ✅ Código de verificación encontrado: {codigo}")
            return codigo
        else:
            log_mensaje("  ❌ No se encontró código de 6 dígitos en el email")
            return None
            
    except Exception as e:
        log_mensaje(f"  ❌ Error obteniendo código de email: {str(e)[:50]}")
        return None
    finally:
        try:
            mail.close()
            mail.logout()
        except:
            pass


def obtener_nombre_web_con_ia(url, driver_temp=None):
    """Obtiene el nombre de la web usando IA local"""
    cerrar_driver = False
    try:
        log_mensaje("  🤖 Obteniendo nombre con IA...")
        
        if not driver_temp:
            chrome_options = Options()
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            driver_temp = webdriver.Chrome(options=chrome_options)
            cerrar_driver = True
        
        try:
            driver_temp.get(url)
            time.sleep(3)
            
            title = ""
            h1_text = ""
            meta_description = ""
            domain = ""
            
            try:
                title = driver_temp.find_element(By.TAG_NAME, 'title').text
            except:
                pass
                
            try:
                h1_element = driver_temp.find_element(By.TAG_NAME, 'h1')
                h1_text = h1_element.text
            except:
                pass
            
            try:
                meta_desc_element = driver_temp.find_element(By.CSS_SELECTOR, 'meta[name="description"]')
                meta_description = meta_desc_element.get_attribute('content') or ""
            except:
                pass
            
            try:
                domain = urlparse(url).netloc.replace('www.', '')
            except:
                pass
            
            prompt = f"""Analiza estos datos de una página web y proporciona SOLO el nombre de la empresa o sitio web, sin explicaciones adicionales.

URL: {url}
Dominio: {domain}
Título: {title}
H1: {h1_text}
Meta descripción: {meta_description[:200]}

Responde únicamente con el nombre de la empresa/sitio web, máximo 3 palabras."""
            
            respuesta = consultar_ia_local(prompt)
            
            if respuesta and isinstance(respuesta, str):
                nombre_web = respuesta.strip()
                nombre_web = nombre_web.replace('"', '').replace("'", '').strip()
                if len(nombre_web) > 50:
                    nombre_web = domain.split('.')[0].title()
                log_mensaje(f"  ✓ Nombre: {nombre_web}")
                return nombre_web
            else:
                log_mensaje("  ❌ Error IA, usando dominio")
                return domain.split('.')[0].title() if domain else "Sitio Web"
                
        finally:
            if cerrar_driver and driver_temp:
                driver_temp.quit()
            
    except Exception as e:
        log_mensaje(f"  ❌ Error obteniendo nombre: {str(e)[:30]}")
        return None


def generar_pdf_titularidad_dominio(screenshots, url_web, nombre_web, timestamp, pdfs_dir):
    """Genera PDF de titularidad del dominio"""
    try:
        if not screenshots:
            log_mensaje("  ❌ No hay screenshots")
            return None
        
        pdf_filename = f'Titularidad_Dominio_{timestamp}.pdf'
        pdf_path = os.path.join(pdfs_dir, pdf_filename)
        
        c = canvas.Canvas(pdf_path, pagesize=A4)
        page_width, page_height = A4
        
        # Portada
        c.setFillColorRGB(1, 1, 1)
        c.rect(0, 0, page_width, page_height, fill=1, stroke=0)
        
        c.setFillColorRGB(0.2, 0.45, 0.75)
        c.rect(0, page_height - 3, page_width, 3, fill=1, stroke=0)
        
        c.setFillColorRGB(0.1, 0.1, 0.1)
        c.setFont("Helvetica-Bold", 20)
        c.drawString(50, page_height - 50, "SITIO WEB Y PRESENCIA BÁSICA EN INTERNET")
        
        c.setFillColorRGB(0.2, 0.2, 0.2)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, page_height - 75, "1. Mantenimiento del Dominio")
        
        c.setFillColorRGB(0.3, 0.3, 0.3)
        c.setFont("Helvetica", 12)
        c.drawString(50, page_height - 105, "Verificación de titularidad del dominio mediante capturas de pantalla")
        c.drawString(50, page_height - 125, "de los paneles de control de los registradores correspondientes.")
        
        c.setFillColorRGB(0.1, 0.1, 0.1)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, page_height - 160, f"Sitio Web: {nombre_web if nombre_web else 'N/A'}")
        c.drawString(50, page_height - 180, f"URL: {url_web}")
        c.drawString(50, page_height - 200, f"Fecha de análisis: {datetime.now().strftime('%d/%m/%Y a las %H:%M')}")
        
        c.setFillColorRGB(0.5, 0.5, 0.5)
        c.setFont("Helvetica", 7)
        c.drawCentredString(page_width/2, 30, f"Justificado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
        c.showPage()
        
        # Páginas de screenshots
        for i, screenshot_path in enumerate(screenshots):
            if os.path.exists(screenshot_path):
                try:
                    c.setFillColorRGB(1, 1, 1)
                    c.rect(0, 0, page_width, page_height, fill=1, stroke=0)
                    
                    c.setFillColorRGB(0.2, 0.45, 0.75)
                    c.rect(0, page_height - 3, page_width, 3, fill=1, stroke=0)
                    
                    c.setFillColorRGB(0.1, 0.1, 0.1)
                    c.setFont("Helvetica-Bold", 16)
                    titulo_captura = f"EVIDENCIA {i+1}: Panel de Control del Dominio"
                    c.drawString(50, page_height - 50, titulo_captura)
                    
                    c.setStrokeColorRGB(0.2, 0.4, 0.8)
                    c.setLineWidth(1)
                    c.line(50, page_height - 70, page_width - 50, page_height - 70)
                    
                    img = Image.open(screenshot_path)
                    img_width, img_height = img.size
                    
                    max_width = page_width - 100
                    max_height = page_height - 200
                    
                    scale_x = max_width / img_width
                    scale_y = max_height / img_height
                    scale = min(scale_x, scale_y)
                    
                    new_width = img_width * scale
                    new_height = img_height * scale
                    
                    x = (page_width - new_width) / 2
                    y = (page_height - new_height) / 2 - 50
                    
                    c.setStrokeColorRGB(0.8, 0.8, 0.8)
                    c.setLineWidth(2)
                    c.rect(x - 10, y - 10, new_width + 20, new_height + 20)
                    
                    c.drawImage(screenshot_path, x, y, width=new_width, height=new_height)
                    
                    c.setFillColorRGB(0.1, 0.1, 0.1)
                    c.setFont("Helvetica", 10)
                    descripcion = f"Archivo: {os.path.basename(screenshot_path)}"
                    c.drawString(50, y - 30, descripcion)
                    
                    c.setFillColorRGB(0.5, 0.5, 0.5)
                    c.setFont("Helvetica", 7)
                    c.drawCentredString(page_width/2, 30, f"Justificado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                    
                    c.showPage()
                    
                except Exception as e:
                    log_mensaje(f"  ✗ Error screenshot {i+1}: {str(e)[:30]}")
        
        c.save()
        log_mensaje(f"  ✓ PDF de titularidad generado: {pdf_filename}")
        return pdf_path
        
    except Exception as e:
        log_mensaje(f"  ❌ Error generando PDF de titularidad: {str(e)[:30]}")
        return None


def generar_pdf_publicidad(fullpage_screenshot, url_web, timestamp, pdfs_dir):
    """Genera PDF de publicidad que solo contiene la captura de la web del cliente"""
    try:
        if not fullpage_screenshot or not os.path.exists(fullpage_screenshot):
            log_mensaje("  ❌ No hay captura de página completa para generar PDF de publicidad")
            return None
        
        pdf_filename = f'Publicidad_{timestamp}.pdf'
        pdf_path = os.path.join(pdfs_dir, pdf_filename)
        
        c = canvas.Canvas(pdf_path, pagesize=A4)
        page_width, page_height = A4
        
        try:
            img = Image.open(fullpage_screenshot)
            img_width, img_height = img.size
            
            max_width = page_width - 20
            max_height = page_height - 20
            
            scale_x = max_width / img_width
            scale_y = max_height / img_height
            scale = min(scale_x, scale_y)
            
            new_width = img_width * scale
            new_height = img_height * scale
            
            x = (page_width - new_width) / 2
            y = (page_height - new_height) / 2
            
            c.drawImage(fullpage_screenshot, x, y, width=new_width, height=new_height)
            
        except Exception as e:
            log_mensaje(f"  ❌ Error cargando imagen: {str(e)}")
            return None
        
        c.save()
        log_mensaje(f"  ✓ PDF de publicidad generado: {pdf_filename}")
        return pdf_path
        
    except Exception as e:
        log_mensaje(f"  ❌ Error generando PDF de publicidad: {str(e)[:30]}")
        return None


def actualizar_estado_crm(callback_url, estado, mensaje=''):
    """Actualiza el estado de la justificación en el CRM"""
    try:
        # Extraer base_url del callback
        base_url = callback_url.rsplit('/receive/', 1)[0]
        justificacion_id = callback_url.rsplit('/', 1)[1]
        update_url = f"{base_url}/update-estado/{justificacion_id}"
        
        response = requests.post(update_url, json={
            'estado': estado,
            'mensaje': mensaje
        }, timeout=10)
        
        if response.status_code == 200:
            log_mensaje(f"  ✅ Estado actualizado a: {estado}")
            return True
        else:
            log_mensaje(f"  ⚠️ Error actualizando estado: {response.status_code}")
            return False
    except Exception as e:
        log_mensaje(f"  ⚠️ Error actualizando estado: {str(e)[:50]}")
        return False


def procesar_cola_trabajos():
    """Worker que procesa la cola de trabajos uno a uno"""
    global procesando, trabajo_actual
    
    while True:
        try:
            # Esperar a que haya un trabajo en la cola
            trabajo = cola_trabajos.get()
            
            if trabajo is None:  # Señal de parada
                break
            
            procesando = True
            trabajo_actual = trabajo
            
            log_mensaje("=" * 80)
            log_mensaje(f"📥 INICIANDO TRABAJO DESDE LA COLA")
            log_mensaje(f"  Cola actual: {cola_trabajos.qsize()} trabajos pendientes")
            log_mensaje(f"  Justificación ID: {trabajo.justificacion_id}")
            log_mensaje("=" * 80)
            
            # Actualizar estado a "procesando" en el CRM
            actualizar_estado_crm(trabajo.callback_url, 'procesando', 'Análisis SEO en curso...')
            
            # Ejecutar el análisis
            resultado = ejecutar_analisis_seo_completo(
                trabajo.url,
                trabajo.justificacion_id,
                trabajo.callback_url,
                trabajo.tipo_analisis
            )
            
            if resultado:
                log_mensaje(f"✅ Trabajo completado exitosamente - ID #{trabajo.justificacion_id}")
            else:
                log_mensaje(f"❌ Trabajo falló - ID #{trabajo.justificacion_id}")
                actualizar_estado_crm(trabajo.callback_url, 'error', 'Error en el análisis SEO')
            
        except Exception as e:
            log_mensaje(f"❌ Error procesando trabajo: {str(e)}")
            if trabajo_actual:
                actualizar_estado_crm(trabajo_actual.callback_url, 'error', f'Error: {str(e)[:100]}')
        
        finally:
            procesando = False
            trabajo_actual = None
            cola_trabajos.task_done()


def ejecutar_analisis_seo_completo(url, justificacion_id, callback_url, tipo_analisis='web'):
    """Ejecuta el análisis SEO completo (igual que la GUI) y envía archivos al CRM"""
    driver = None
    
    try:
        log_mensaje("=" * 80)
        log_mensaje(f"🚀 INICIANDO ANÁLISIS SEO PARA JUSTIFICACIÓN #{justificacion_id}")
        log_mensaje(f"📍 URL: {url}")
        log_mensaje(f"🔗 Callback: {callback_url}")
        log_mensaje(f"🏷️ Tipo: {tipo_analisis.upper()}")
        log_mensaje("=" * 80)
        
        base_dir = obtener_directorio_base()
        screenshots_dir = os.path.join(base_dir, 'screenshots')
        pdfs_dir = os.path.join(base_dir, 'pdfs_generados')
        os.makedirs(screenshots_dir, exist_ok=True)
        os.makedirs(pdfs_dir, exist_ok=True)
        
        url = normalizar_url(url)
        log_mensaje(f"URL normalizada: {url}")
        
        # Configurar Chrome (SIN headless - ventana visible)
        log_mensaje("⚙️ Configurando Chrome...")
        chrome_options = Options()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        prefs = {"credentials_enable_service": False, "profile.password_manager_enabled": False}
        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.add_argument('--start-maximized')
        
        log_mensaje("✅ Chrome configurado (ventana visible)")
        
        driver = webdriver.Chrome(options=chrome_options)
        driver.maximize_window()
        wait = WebDriverWait(driver, 20)
        
        # EXTRACCIÓN AUTOMÁTICA DE KEYWORDS
        log_mensaje("🔍 Extrayendo keywords automáticamente...")
        chrome_options_temp = Options()
        chrome_options_temp.add_argument('--headless')
        chrome_options_temp.add_argument('--no-sandbox')
        chrome_options_temp.add_argument('--disable-dev-shm-usage')
        
        driver_temp = webdriver.Chrome(options=chrome_options_temp)
        driver_temp.get(url)
        time.sleep(2)
        
        keywords = extraer_keywords_automaticas(driver_temp, url)
        driver_temp.quit()
        
        if not keywords or len(keywords) == 0:
            log_mensaje("❌ No se pudieron extraer keywords")
            return None
        
        log_mensaje(f"✅ Keywords extraídas: {', '.join(keywords)}")
        
        # Determinar si es ecommerce y generar competidores
        es_ecommerce = (tipo_analisis.lower() == 'ecommerce')
        competidores = []
        
        if es_ecommerce:
            log_mensaje("🛒 ECOMMERCE detectado - Generando competidores...")
            actualizar_estado_crm(callback_url, 'procesando', 'Buscando competidores...')
            from app_gui import SEOAnalyzerGUI
            
            # Crear instancia temporal para usar el método de búsqueda
            class TempGUI:
                def log(self, msg):
                    log_mensaje(f"    {msg}")
            
            temp_gui = TempGUI()
            
            # Usar el método de búsqueda de competidores de la GUI
            chrome_options_comp = Options()
            chrome_options_comp.add_argument('--no-sandbox')
            chrome_options_comp.add_argument('--disable-dev-shm-usage')
            driver_comp = webdriver.Chrome(options=chrome_options_comp)
            
            try:
                # Importar y usar la función de búsqueda
                competidores = buscar_competidores_simple_server(driver_comp, keywords, url, temp_gui)
                if competidores and len(competidores) >= 2:
                    competidores = competidores[:2]
                    log_mensaje(f"✅ Competidores generados: {', '.join(competidores)}")
                else:
                    log_mensaje("⚠️ No se pudieron generar suficientes competidores")
            finally:
                driver_comp.quit()
        else:
            log_mensaje("🌐 WEB normal - Sin competidores")
            competidores = []
        
        # ===== RANKALYZE - ANÁLISIS COMPLETO (IGUAL QUE LA GUI) =====
        
        log_mensaje("=" * 80)
        log_mensaje("📧 LOGIN EN HERRAMIENTA SEO")
        log_mensaje("=" * 80)
        driver.get('https://rankalyze.net/login')
        time.sleep(2)
        
        username_input = wait.until(EC.element_to_be_clickable((By.NAME, 'email')))
        username_input.clear()
        time.sleep(0.3)
        username_input.send_keys('admin@admin.com')
        time.sleep(0.3)
        
        password_input = wait.until(EC.element_to_be_clickable((By.NAME, 'password')))
        password_input.clear()
        time.sleep(0.3)
        password_input.send_keys('R4t4-2020')
        time.sleep(0.3)
        
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="submitButton"]')))
        submit_button.click()
        time.sleep(3)
        
        current_url = driver.current_url
        if 'login' in current_url.lower():
            raise Exception("Login falló - Verificar credenciales")
        
        log_mensaje("✅ Login OK")
        
        log_mensaje("=" * 80)
        log_mensaje("📝 LLENANDO FORMULARIO")
        log_mensaje("=" * 80)
        se_form_link = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/aside/nav/ul/li[2]/a')))
        se_form_link.click()
        time.sleep(2)
        
        random_btn = wait.until(EC.element_to_be_clickable((By.ID, 'randomBtn')))
        random_btn.click()
        alert = wait.until(EC.alert_is_present())
        alert.accept()
        log_mensaje("✅ Alert OK")
        
        log_mensaje(f"🔑 Ingresando {len(keywords)} keywords...")
        for i, keyword in enumerate(keywords):
            keyword_input = wait.until(EC.element_to_be_clickable((By.ID, f'palabra_{i}')))
            keyword_input.clear()
            time.sleep(0.2)
            keyword_input.send_keys(Keys.CONTROL + "a")
            keyword_input.send_keys(Keys.DELETE)
            time.sleep(0.2)
            keyword_input.send_keys(keyword)
            time.sleep(0.2)
            log_mensaje(f"  ✓ Keyword {i + 1}: {keyword}")
        
        # Llenar competidores si es ECOMMERCE
        if es_ecommerce and competidores:
            log_mensaje(f"🏢 Ingresando {len(competidores)} competidores...")
            
            if len(competidores) > 0:
                log_mensaje(f"  📝 Competidor 1: {competidores[0]}")
                c1 = wait.until(EC.element_to_be_clickable((By.ID, 'competitor1_url')))
                c1.clear()
                time.sleep(0.2)
                c1.send_keys(Keys.CONTROL + "a")
                c1.send_keys(Keys.DELETE)
                time.sleep(0.2)
                c1.send_keys(competidores[0])
                log_mensaje("  ✅ Competidor 1 ingresado")
            
            if len(competidores) > 1:
                log_mensaje(f"  📝 Competidor 2: {competidores[1]}")
                c2 = wait.until(EC.element_to_be_clickable((By.ID, 'competitor2_url')))
                c2.clear()
                time.sleep(0.2)
                c2.send_keys(Keys.CONTROL + "a")
                c2.send_keys(Keys.DELETE)
                time.sleep(0.2)
                c2.send_keys(competidores[1])
                log_mensaje("  ✅ Competidor 2 ingresado")
        
        fecha_hoy = datetime.now().strftime('%d/%m/%Y')
        fecha_input = wait.until(EC.element_to_be_clickable((By.ID, 'fecha_inicio')))
        fecha_input.clear()
        time.sleep(0.2)
        fecha_input.send_keys(Keys.CONTROL + "a")
        fecha_input.send_keys(Keys.DELETE)
        time.sleep(0.2)
        fecha_input.send_keys(fecha_hoy)
        
        url_input = wait.until(EC.element_to_be_clickable((By.ID, 'url')))
        url_input.clear()
        time.sleep(0.2)
        url_input.send_keys(Keys.CONTROL + "a")
        url_input.send_keys(Keys.DELETE)
        time.sleep(0.2)
        url_input.send_keys(url)
        
        submit_final_btn = wait.until(EC.element_to_be_clickable((By.ID, 'submitBtn')))
        submit_final_btn.click()
        log_mensaje("✅ Formulario enviado")
        time.sleep(10)
        
        # ===== CAPTURAS DE RANKALYZE =====
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Vista General
        log_mensaje("📸 Captura Vista General...")
        vista_general_screenshot = None
        try:
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(0.5)
            vista_general_filename = f'{screenshots_dir}/vista_general_{timestamp}.png'
            driver.save_screenshot(vista_general_filename)
            vista_general_screenshot = vista_general_filename
            log_mensaje("  ✅ Vista General capturada")
        except Exception as e:
            log_mensaje(f"  ❌ Vista General: {str(e)[:30]}")
        
        # Indexación
        log_mensaje("📸 Captura Indexación...")
        indexacion_screenshots = []
        try:
            index1_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/main/div[4]/div[1]')))
            driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", index1_element)
            driver.execute_script("window.scrollBy(0, -100);")
            time.sleep(0.3)
            index1_filename = f'{screenshots_dir}/indexacion_1_{timestamp}.png'
            index1_element.screenshot(index1_filename)
            indexacion_screenshots.append(index1_filename)
            log_mensaje("  ✅ Indexación 1")
        except Exception as e:
            log_mensaje(f"  ❌ Indexación 1: {str(e)[:30]}")
        
        try:
            index2_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/main/div[4]/div[2]')))
            driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", index2_element)
            driver.execute_script("window.scrollBy(0, -100);")
            time.sleep(0.3)
            index2_filename = f'{screenshots_dir}/indexacion_2_{timestamp}.png'
            index2_element.screenshot(index2_filename)
            indexacion_screenshots.append(index2_filename)
            log_mensaje("  ✅ Indexación 2")
        except Exception as e:
            log_mensaje(f"  ❌ Indexación 2: {str(e)[:30]}")
        
        # Keywords
        log_mensaje("📸 Captura Keywords...")
        keyword_tabs = driver.find_elements(By.CLASS_NAME, 'keyword-tab')
        log_mensaje(f"Encontradas {len(keyword_tabs)} keyword tabs")
        
        screenshots_guardadas = []
        for index, tab in enumerate(keyword_tabs):
            try:
                log_mensaje(f"  Capturando keyword {index + 1}/{len(keyword_tabs)}...")
                keyword_text = tab.text.strip().replace(' ', '_').replace('/', '_')
                data_keyword = tab.get_attribute('data-keyword')
                
                tab.click()
                
                wait_fast = WebDriverWait(driver, 5)
                try:
                    keywords_section = wait_fast.until(EC.visibility_of_element_located((By.ID, 'keywords-section')))
                except:
                    keywords_section = wait_fast.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.keywords-section')))
                
                driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", keywords_section)
                driver.execute_script("window.scrollBy(0, -100);")
                
                filename = f'{screenshots_dir}/keyword_{data_keyword}_{keyword_text}_{timestamp}.png'
                keywords_section.screenshot(filename)
                screenshots_guardadas.append(filename)
                log_mensaje(f"    ✅ Keyword {index + 1} capturada")
            except Exception as e:
                log_mensaje(f"    ❌ Error en keyword {index + 1}: {str(e)[:30]}")
                continue
        
        log_mensaje(f"✅ Total keywords capturadas: {len(screenshots_guardadas)}")
        
        # Capturas de competencia si es ECOMMERCE
        competencia_screenshots = []
        analisis_competencia_ia = None
        metrics_screenshot = None
        
        if es_ecommerce and competidores:
            log_mensaje("🏆 Capturando análisis de competencia...")
            actualizar_estado_crm(callback_url, 'procesando', 'Analizando competencia...')
            try:
                # Capturas de competencia
                comp1_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/main/div[5]/div[1]')))
                driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", comp1_element)
                driver.execute_script("window.scrollBy(0, -100);")
                time.sleep(0.3)
                comp1_filename = f'{screenshots_dir}/competencia_1_{timestamp}.png'
                comp1_element.screenshot(comp1_filename)
                competencia_screenshots.append(comp1_filename)
                log_mensaje(f"  ✅ Competencia 1 capturada")
            except:
                log_mensaje(f"  ❌ Error captura competencia 1")
            
            try:
                comp2_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/main/div[5]/div[2]')))
                driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", comp2_element)
                driver.execute_script("window.scrollBy(0, -100);")
                time.sleep(0.3)
                comp2_filename = f'{screenshots_dir}/competencia_2_{timestamp}.png'
                comp2_element.screenshot(comp2_filename)
                competencia_screenshots.append(comp2_filename)
                log_mensaje(f"  ✅ Competencia 2 capturada")
            except:
                log_mensaje(f"  ❌ Error captura competencia 2")
            
            log_mensaje(f"✅ Total capturas competencia: {len(competencia_screenshots)}")
        
        # WHOIS
        whois_screenshot = None
        log_mensaje("🌐 Captura WHOIS...")
        
        url_sin_protocolo = url.replace('https://', '').replace('http://', '').replace('www.', '').split('/')[0].rstrip('/')
        is_dominio_es = url_sin_protocolo.endswith('.es')
        
        if is_dominio_es:
            log_mensaje("  Dominio .es detectado")
            nombre_dominio = url_sin_protocolo.replace('.es', '')
            try:
                search_url = f'https://nic.es/sgnd/dominio/publicBuscarDominios.action?tDominio.nombreDominio={nombre_dominio}&flag=activado'
                driver.get(search_url)
                time.sleep(3)
                ver_datos_btn = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'verDatosNRGTM')))
                ver_datos_btn.click()
                time.sleep(2)
                buscar_button = wait.until(EC.element_to_be_clickable((By.NAME, 'Buscar')))
                buscar_button.click()
                time.sleep(3)
                target_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.center-on-page-plantillaBase.pageBase')))
                driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", target_element)
                whois_filename = f'{screenshots_dir}/whois_domain_es_{timestamp}.png'
                target_element.screenshot(whois_filename)
                whois_screenshot = whois_filename
                log_mensaje("  ✅ WHOIS capturado")
            except Exception as e:
                log_mensaje(f"  ❌ WHOIS: {str(e)[:30]}")
        else:
            log_mensaje("  Dominio internacional detectado")
            try:
                whois_url = f'https://www.whois.com/whois/{url_sin_protocolo}'
                driver.get(whois_url)
                time.sleep(4)
                whois_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/main')))
                driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", whois_element)
                driver.execute_script("window.scrollBy(0, -50);")
                whois_filename = f'{screenshots_dir}/whois_domain_{timestamp}.png'
                whois_element.screenshot(whois_filename)
                whois_screenshot = whois_filename
                log_mensaje("  ✅ WHOIS capturado")
            except Exception as e:
                log_mensaje(f"  ❌ WHOIS: {str(e)[:30]}")
        
        # H1/H2 y Excel
        h1_screenshot = None
        h2_screenshot = None
        log_mensaje("🏷️ Captura H1/H2 y modificación de Excel...")
        
        try:
            driver.get(url)
            time.sleep(3)
            
            log_mensaje("📊 Modificando Excel (Página 2)...")
            nombre_web = obtener_nombre_web_con_ia(url)
            excel_copia = modificar_excel_informe(url, nombre_web)
            log_mensaje("  ✅ Excel Pág 2 modificado")
            
            log_mensaje("🔍 Buscando URLs de Sobre Nosotros y Contacto...")
            sobre_nosotros_url, contacto_url = buscar_urls_en_pagina(driver, url)
            log_mensaje(f"  ✅ Sobre Nosotros: {sobre_nosotros_url}")
            log_mensaje(f"  ✅ Contacto: {contacto_url}")
            
            if excel_copia:
                log_mensaje("📊 Modificando Excel (Página 4)...")
                modificar_excel_urls_adicionales(excel_copia, url, sobre_nosotros_url, contacto_url)
                log_mensaje("  ✅ Excel Pág 4 modificado")
            
            # Capturar H1
            try:
                h1_element = driver.find_element(By.TAG_NAME, 'h1')
                h1_html = h1_element.get_attribute('outerHTML')
                h1_text = h1_element.text
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", h1_element)
                driver.execute_script("""
                    arguments[0].style.outline = '3px solid red';
                    arguments[0].style.backgroundColor = 'rgba(255, 255, 0, 0.3)';
                    var p = document.createElement('div');
                    p.id = 'custom-inspector-h1';
                    p.style.cssText = 'position: fixed; right: 20px; top: 80px; width: 400px; max-height: 600px; background: #282828; color: #fff; padding: 20px; border-radius: 8px; box-shadow: 0 8px 32px rgba(0,0,0,0.5); font-family: Courier New, monospace; font-size: 13px; z-index: 999999; overflow-y: auto;';
                    var t = document.createElement('div');
                    t.style.cssText = 'color: #61dafb; font-weight: bold; margin-bottom: 15px; font-size: 14px; border-bottom: 2px solid #61dafb; padding-bottom: 8px;';
                    t.textContent = 'Inspector - H1';
                    p.appendChild(t);
                    var c = document.createElement('pre');
                    c.style.cssText = 'margin: 0; white-space: pre-wrap; word-wrap: break-word; color: #e06c75;';
                    c.textContent = arguments[1];
                    p.appendChild(c);
                    var i = document.createElement('div');
                    i.style.cssText = 'margin-top: 15px; padding-top: 15px; border-top: 1px solid #444; color: #98c379;';
                    i.innerHTML = '<strong>Texto:</strong> ' + arguments[2];
                    p.appendChild(i);
                    document.body.appendChild(p);
                """, h1_element, h1_html, h1_text)
                time.sleep(1.5)
                h1_filename = f'{screenshots_dir}/h1_inspector_{timestamp}.png'
                driver.save_screenshot(h1_filename)
                h1_screenshot = h1_filename
                log_mensaje("  ✅ H1 capturado")
                driver.execute_script("arguments[0].style.outline = ''; arguments[0].style.backgroundColor = ''; var p = document.getElementById('custom-inspector-h1'); if (p) p.remove();", h1_element)
            except:
                log_mensaje("  ❌ H1")
            
            # Capturar H2
            try:
                h2_element = driver.find_element(By.TAG_NAME, 'h2')
                h2_html = h2_element.get_attribute('outerHTML')
                h2_text = h2_element.text
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", h2_element)
                driver.execute_script("""
                    arguments[0].style.outline = '3px solid blue';
                    arguments[0].style.backgroundColor = 'rgba(0, 255, 255, 0.3)';
                    var p = document.createElement('div');
                    p.id = 'custom-inspector-h2';
                    p.style.cssText = 'position: fixed; right: 20px; top: 80px; width: 400px; max-height: 600px; background: #282828; color: #fff; padding: 20px; border-radius: 8px; box-shadow: 0 8px 32px rgba(0,0,0,0.5); font-family: Courier New, monospace; font-size: 13px; z-index: 999999; overflow-y: auto;';
                    var t = document.createElement('div');
                    t.style.cssText = 'color: #61dafb; font-weight: bold; margin-bottom: 15px; font-size: 14px; border-bottom: 2px solid #61dafb; padding-bottom: 8px;';
                    t.textContent = 'Inspector - H2';
                    p.appendChild(t);
                    var c = document.createElement('pre');
                    c.style.cssText = 'margin: 0; white-space: pre-wrap; word-wrap: break-word; color: #e06c75;';
                    c.textContent = arguments[1];
                    p.appendChild(c);
                    var i = document.createElement('div');
                    i.style.cssText = 'margin-top: 15px; padding-top: 15px; border-top: 1px solid #444; color: #98c379;';
                    i.innerHTML = '<strong>Texto:</strong> ' + arguments[2];
                    p.appendChild(i);
                    document.body.appendChild(p);
                """, h2_element, h2_html, h2_text)
                time.sleep(1.5)
                h2_filename = f'{screenshots_dir}/h2_inspector_{timestamp}.png'
                driver.save_screenshot(h2_filename)
                h2_screenshot = h2_filename
                log_mensaje("  ✅ H2 capturado")
                driver.execute_script("arguments[0].style.outline = ''; arguments[0].style.backgroundColor = ''; var p = document.getElementById('custom-inspector-h2'); if (p) p.remove();", h2_element)
            except:
                log_mensaje("  ❌ H2")
        except Exception as e:
            log_mensaje(f"  ❌ Error en H1/H2: {str(e)[:30]}")
        
        # Pantalla completa
        fullpage_screenshot = None
        log_mensaje("🖥️ Captura pantalla completa...")
        try:
            driver.get(url)
            time.sleep(3)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            fullpage_filename = f'{screenshots_dir}/fullpage_{timestamp}.png'
            fullpage_screenshot_img = pyautogui.screenshot()
            fullpage_screenshot_img.save(fullpage_filename)
            fullpage_screenshot = fullpage_filename
            log_mensaje("  ✅ Pantalla completa capturada")
        except Exception as e:
            log_mensaje(f"  ❌ Pantalla completa: {str(e)[:30]}")
        
        # Panel de control del dominio
        log_mensaje("🌐 Captura panel de control del dominio...")
        panel_control_screenshots = capturar_panel_control_dominio(url, driver, timestamp, screenshots_dir)
        
        driver.quit()
        driver = None
        
        # ===== GENERACIÓN DE PDFS =====
        
        log_mensaje("=" * 80)
        log_mensaje("📄 GENERANDO PDFS")
        log_mensaje("=" * 80)
        
        # PDF Principal (Justificación)
        log_mensaje("📄 Generando PDF de justificación...")
        actualizar_estado_crm(callback_url, 'procesando', 'Generando PDF de justificación...')
        
        pdf_justificacion = generar_pdf_con_capturas(
            screenshots_guardadas, keywords, url,
            competidores=competidores,
            whois_screenshot=whois_screenshot,
            h1_screenshot=h1_screenshot,
            h2_screenshot=h2_screenshot,
            fullpage_screenshot=fullpage_screenshot,
            competencia_screenshots=competencia_screenshots if es_ecommerce else [],
            metrics_screenshot=metrics_screenshot if es_ecommerce else None,
            vista_general_screenshot=vista_general_screenshot,
            indexacion_screenshots=indexacion_screenshots,
            analisis_competencia_ia=analisis_competencia_ia if es_ecommerce else None
        )
        
        if not pdf_justificacion:
            log_mensaje("❌ Error generando PDF de justificación")
            return None
        
        log_mensaje(f"✅ PDF Justificación: {os.path.basename(pdf_justificacion)}")
        
        # PDF Titularidad
        pdf_titularidad = None
        if panel_control_screenshots:
            log_mensaje("📄 Generando PDF de titularidad...")
            actualizar_estado_crm(callback_url, 'procesando', 'Generando PDF de titularidad...')
            pdf_titularidad = generar_pdf_titularidad_dominio(
                panel_control_screenshots, url, nombre_web, timestamp, pdfs_dir
            )
            if pdf_titularidad:
                log_mensaje(f"✅ PDF Titularidad: {os.path.basename(pdf_titularidad)}")
            else:
                log_mensaje("❌ Error PDF Titularidad")
        else:
            log_mensaje("⚠️ No hay capturas para PDF de titularidad")
        
        # PDF Publicidad
        pdf_publicidad = None
        if fullpage_screenshot and os.path.exists(fullpage_screenshot):
            log_mensaje("📄 Generando PDF de publicidad...")
            actualizar_estado_crm(callback_url, 'procesando', 'Generando PDF de publicidad...')
            pdf_publicidad = generar_pdf_publicidad(
                fullpage_screenshot, url, timestamp, pdfs_dir
            )
            if pdf_publicidad:
                log_mensaje(f"✅ PDF Publicidad: {os.path.basename(pdf_publicidad)}")
            else:
                log_mensaje("❌ Error PDF Publicidad")
        else:
            log_mensaje("⚠️ No hay captura para PDF de publicidad")
        
        # ===== ENVIAR ARCHIVOS AL CRM =====
        
        log_mensaje("=" * 80)
        log_mensaje("📤 ENVIANDO ARCHIVOS AL CRM")
        log_mensaje("=" * 80)
        
        if not pdf_justificacion or not pdf_titularidad or not pdf_publicidad:
            log_mensaje("❌ No se generaron todos los PDFs requeridos")
            actualizar_estado_crm(callback_url, 'error', 'No se generaron todos los PDFs')
            return None
        
        actualizar_estado_crm(callback_url, 'procesando', 'Enviando archivos al CRM...')
        enviar_archivos_al_crm(callback_url, pdf_justificacion, pdf_titularidad, pdf_publicidad, justificacion_id)
        
        log_mensaje("=" * 80)
        log_mensaje(f"✅ ANÁLISIS COMPLETADO PARA JUSTIFICACIÓN #{justificacion_id}")
        log_mensaje("=" * 80)
        
        return True
        
    except Exception as e:
        log_mensaje(f"💥 ERROR GENERAL: {str(e)}")
        import traceback
        log_mensaje(f"Traceback: {traceback.format_exc()}")
        return None
        
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass


def capturar_panel_control_dominio(url_web, driver, timestamp, screenshots_dir):
    """Captura screenshots del panel de control del dominio (IGUAL QUE GUI)"""
    screenshots = []
    
    try:
        domain = urlparse(url_web).netloc.replace('www.', '')
        domain_name_without_tld = domain.split('.')[0]
        domain_name_with_tld = domain
        
        log_mensaje(f"  🔍 Dominio completo: {domain_name_with_tld}")
        log_mensaje(f"  🔍 Dominio sin TLD: {domain_name_without_tld}")
        
        log_mensaje("  🔍 Intentando con Ionos...")
        ionos_url = f"https://my.ionos.es/domain-privacy/{domain_name_with_tld}?linkId=ct.tab.domainlist.privacy"
        
        try:
            log_mensaje("  📍 Accediendo a Ionos...")
            driver.get(ionos_url)
            time.sleep(3)
            
            current_url = driver.current_url
            if "login" in current_url or "signin" in current_url:
                log_mensaje("  🔐 Login en Ionos...")
                wait = WebDriverWait(driver, 15)
                
                try:
                    username_input = wait.until(EC.element_to_be_clickable((By.NAME, "username")))
                    username_input.clear()
                    username_input.send_keys("i-pointsite.com")
                    time.sleep(0.5)
                    
                    next_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit']")))
                    next_button.click()
                    time.sleep(2)
                    log_mensaje("  ✅ Username enviado")
                    
                    password_input = wait.until(EC.element_to_be_clickable((By.NAME, "password")))
                    password_input.clear()
                    password_input.send_keys("R4t420223!")
                    time.sleep(0.5)
                    
                    login_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit']")))
                    login_button.click()
                    time.sleep(3)
                    log_mensaje("  ✅ Login completado")
                except Exception as login_error:
                    log_mensaje(f"  ❌ Error en login: {str(login_error)[:50]}")
            else:
                log_mensaje("  ✅ Ya autenticado en Ionos")
            
            current_url = driver.current_url
            if "my.ionos.es/domains" in current_url or "my.ionos.es" not in current_url:
                log_mensaje("  ❌ No está en Ionos - probando Dondominio")
                raise Exception("Dominio no en Ionos")
            else:
                log_mensaje("  ✅ En panel de Ionos")
            
            # Capturas de Ionos
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight/2);")
            time.sleep(1)
            
            wait = WebDriverWait(driver, 10)
            elemento = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/main/div[2]/section[3]")))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elemento)
            driver.execute_script("window.scrollBy(0, -50);")
            time.sleep(0.5)
            
            driver.execute_script("document.body.style.zoom='1.2'")
            time.sleep(0.3)
            
            screenshot_filename = f'{screenshots_dir}/ionos_elemento_{timestamp}.png'
            elemento.screenshot(screenshot_filename)
            screenshots.append(screenshot_filename)
            log_mensaje("  ✅ Ionos elemento capturado")
            
            driver.execute_script("document.body.style.zoom='1.0'")
            
            driver.execute_script("document.body.style.zoom='0.85'")
            time.sleep(0.5)
            
            screenshot_filename = f'{screenshots_dir}/ionos_completa_{timestamp}.png'
            pyautogui.screenshot(screenshot_filename)
            screenshots.append(screenshot_filename)
            log_mensaje("  ✅ Ionos completa capturada")
            
            driver.execute_script("document.body.style.zoom='1.0'")
            
            return screenshots
            
        except Exception as e:
            log_mensaje(f"  ❌ Error con Ionos: {str(e)[:30]}")
            log_mensaje("  🔄 Probando con Dondominio...")
            
            # DONDOMINIO - Lógica completa
            try:
                log_mensaje("  🔐 Accediendo a login de Dondominio...")
                driver.get("https://www.dondominio.com/admin/")
                time.sleep(3)
                
                wait = WebDriverWait(driver, 15)
                
                # Login en Dondominio con Shadow DOM
                log_mensaje("  🔐 Iniciando sesión en Dondominio...")
                
                # Usuario con Shadow DOM
                try:
                    driver.execute_script("""
                        var text = 'THWORK';
                        var ddInput = document.querySelector('dd-input[name="user"]');
                        if (ddInput && ddInput.shadowRoot) {
                            var realInput = ddInput.shadowRoot.querySelector('input');
                            if (realInput) {
                                realInput.focus();
                                realInput.value = '';
                                realInput.value = text;
                                realInput.dispatchEvent(new Event('input', { bubbles: true }));
                                realInput.dispatchEvent(new Event('change', { bubbles: true }));
                                realInput.dispatchEvent(new Event('blur', { bubbles: true }));
                                return 'SUCCESS';
                            }
                        }
                        return 'NO_SHADOW_ROOT';
                    """)
                    time.sleep(1)
                    log_mensaje("  ✅ Usuario ingresado")
                except Exception as e:
                    log_mensaje(f"  ❌ Error usuario: {str(e)[:30]}")
                
                # Contraseña con Shadow DOM
                try:
                    driver.execute_script("""
                        var text = 'R4t4-2025';
                        var ddPassword = document.querySelector('dd-password[name="password"]');
                        if (ddPassword && ddPassword.shadowRoot) {
                            var realInput = ddPassword.shadowRoot.querySelector('input');
                            if (realInput) {
                                realInput.focus();
                                realInput.value = '';
                                realInput.value = text;
                                realInput.dispatchEvent(new Event('input', { bubbles: true }));
                                realInput.dispatchEvent(new Event('change', { bubbles: true }));
                                realInput.dispatchEvent(new Event('blur', { bubbles: true }));
                                return 'SUCCESS';
                            }
                        }
                        return 'NO_SHADOW_ROOT';
                    """)
                    time.sleep(1)
                    log_mensaje("  ✅ Contraseña ingresada")
                except Exception as e:
                    log_mensaje(f"  ❌ Error contraseña: {str(e)[:30]}")
                
                # Click login
                try:
                    login_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@type='submit']")))
                    login_button.click()
                    time.sleep(5)
                    log_mensaje("  ✅ Login realizado")
                    
                    # Verificar 2FA
                    try:
                        time.sleep(2)
                        twofa_element = driver.find_element(By.CLASS_NAME, "input-2fa-code")
                        if twofa_element:
                            log_mensaje("  🔐 Verificación 2FA detectada")
                            codigo = obtener_codigo_verificacion_email()
                            if codigo and len(codigo) == 6:
                                log_mensaje(f"  📧 Código obtenido: {codigo}")
                                for i, digito in enumerate(codigo, 1):
                                    try:
                                        input_field = driver.find_element(By.NAME, f"code-sub-{i}")
                                        input_field.clear()
                                        input_field.send_keys(digito)
                                        time.sleep(0.2)
                                    except:
                                        pass
                                
                                verify_button = driver.find_element(By.CSS_SELECTOR, "button.btn.btn-dark.btn-lg.btn-block[type='submit']")
                                verify_button.click()
                                time.sleep(3)
                                log_mensaje("  ✅ Verificación 2FA completada")
                    except:
                        log_mensaje("  ✓ No se requiere 2FA")
                        
                except Exception as e:
                    log_mensaje(f"  ❌ Error en login: {str(e)[:30]}")
                
                # Ir a lista de dominios
                log_mensaje("  📍 Navegando a lista de dominios...")
                driver.get("https://www.dondominio.com/admin/domains/list/")
                time.sleep(3)
                
                # Buscar dominio
                log_mensaje(f"  🔍 Buscando dominio: {domain_name_without_tld}")
                try:
                    filter_input = wait.until(EC.element_to_be_clickable((By.NAME, "filterName")))
                    filter_input.click()
                    time.sleep(0.5)
                    filter_input.clear()
                    time.sleep(0.5)
                    filter_input.send_keys(domain_name_without_tld)
                    time.sleep(1)
                    filter_input.send_keys(Keys.RETURN)
                    time.sleep(3)
                    log_mensaje("  ✅ Dominio buscado")
                except Exception as e:
                    log_mensaje(f"  ❌ Error buscando: {str(e)[:30]}")
                
                # Zoom y capturas
                driver.execute_script("document.body.style.zoom='0.9'")
                time.sleep(1)
                
                screenshot_filename = f'{screenshots_dir}/dondominio_completa_{timestamp}.png'
                pyautogui.screenshot(screenshot_filename)
                screenshots.append(screenshot_filename)
                log_mensaje("  ✅ Dondominio completa")
                
                # Elemento específico
                try:
                    elemento = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[1]/div[4]/div/div[3]/div/div/div/div[2]/div[2]")))
                    driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", elemento)
                    driver.execute_script("window.scrollBy(0, -100);")
                    time.sleep(0.5)
                    
                    screenshot_filename = f'{screenshots_dir}/dondominio_elemento_{timestamp}.png'
                    elemento.screenshot(screenshot_filename)
                    screenshots.append(screenshot_filename)
                    log_mensaje("  ✅ Dondominio elemento")
                except:
                    pass
                
                # Click en enlace
                try:
                    enlace = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[1]/div[4]/div/div[3]/div/div/div/div[1]/div[2]/div[1]/div[3]/div[2]/div[1]/div[1]/a")))
                    enlace.click()
                    time.sleep(3)
                    
                    screenshot_filename = f'{screenshots_dir}/dondominio_final_{timestamp}.png'
                    pyautogui.screenshot(screenshot_filename)
                    screenshots.append(screenshot_filename)
                    log_mensaje("  ✅ Dondominio final")
                except:
                    pass
                
                return screenshots
                
            except Exception as e:
                log_mensaje(f"  ❌ Error con Dondominio: {str(e)[:30]}")
                return []
    
    except Exception as e:
        log_mensaje(f"  💥 Error general capturando panel: {str(e)[:30]}")
        return []


def buscar_competidores_simple_server(driver, keywords, url, temp_gui):
    """Búsqueda simple de competidores para servidor"""
    try:
        competidores = []
        
        if keywords:
            keyword_principal = keywords[0]
            temp_gui.log(f"🔍 Buscando competidores para: {keyword_principal}")
            
            # Usar SerpAPI
            api_key = "d6627f12f17390dd5a229ef805199b662065d1fc8e1276c686fb270381a242a2"
            search_terms = f'"{keyword_principal}" empresa servicios'
            
            temp_gui.log(f"🔗 Buscando con SerpAPI: {search_terms}")
            
            params = {
                "q": search_terms,
                "api_key": api_key,
                "engine": "google",
                "num": 10,
                "gl": "es",
                "hl": "es"
            }
            
            try:
                response = requests.get("https://serpapi.com/search", params=params, timeout=10)
                
                if response.status_code == 200:
                    data = response.json()
                    organic_results = data.get("organic_results", [])
                    temp_gui.log(f"✓ Encontrados {len(organic_results)} resultados")
                    
                    dominios_excluidos = [
                        'google.com', 'youtube.com', 'facebook.com', 'instagram.com', 'twitter.com',
                        'linkedin.com', 'wikipedia.org', 'amazon.com', 'reddit.com', 'github.com'
                    ]
                    
                    for result in organic_results:
                        try:
                            href = result.get("link", "")
                            if not href or not href.startswith('http'):
                                continue
                            
                            es_valido = True
                            for dominio in dominios_excluidos:
                                if dominio in href.lower():
                                    es_valido = False
                                    break
                            
                            parsed_url = urlparse(href)
                            domain = parsed_url.netloc.lower()
                            clean_url = f"https://{domain}"
                            
                            if es_valido and clean_url not in competidores and url not in clean_url:
                                if ('.com' in domain or '.es' in domain or '.org' in domain):
                                    competidores.append(clean_url)
                                    temp_gui.log(f"✅ Competidor: {clean_url}")
                                    if len(competidores) >= 2:
                                        break
                        except:
                            continue
            except Exception as e:
                temp_gui.log(f"❌ Error SerpAPI: {str(e)[:30]}")
        
        return competidores[:2]
        
    except Exception as e:
        temp_gui.log(f"❌ Error buscando competidores: {str(e)[:30]}")
        return []


def enviar_archivos_al_crm(callback_url, pdf_justificacion, pdf_titularidad, pdf_publicidad, justificacion_id):
    """Envía los 3 archivos generados de vuelta al CRM"""
    try:
        log_mensaje(f"📤 Enviando archivos al CRM...")
        log_mensaje(f"  Callback URL: {callback_url}")
        log_mensaje(f"  Justificación: {os.path.basename(pdf_justificacion)}")
        log_mensaje(f"  Titularidad: {os.path.basename(pdf_titularidad)}")
        log_mensaje(f"  Publicidad: {os.path.basename(pdf_publicidad)}")
        
        files = {
            'archivo_just': ('justificacion.pdf', open(pdf_justificacion, 'rb'), 'application/pdf'),
            'archivo_titularidad': ('titularidad.pdf', open(pdf_titularidad, 'rb'), 'application/pdf'),
            'archivo_publicidad': ('publicidad.pdf', open(pdf_publicidad, 'rb'), 'application/pdf')
        }
        
        response = requests.post(callback_url, files=files, timeout=CALLBACK_TIMEOUT)
        
        # Cerrar archivos
        for name, (_, file_obj, _) in files.items():
            file_obj.close()
        
        if response.status_code == 200:
            log_mensaje(f"✅ Archivos enviados correctamente al CRM")
            log_mensaje(f"  Respuesta: {response.json()}")
            return True
        else:
            log_mensaje(f"❌ Error enviando archivos: HTTP {response.status_code}")
            log_mensaje(f"  Respuesta: {response.text}")
            return False
            
    except Exception as e:
        log_mensaje(f"❌ Error enviando archivos al CRM: {str(e)}")
        return False


# ===== ENDPOINTS DEL SERVIDOR =====

@app.route('/sgbasc', methods=['POST'])
def recibir_justificacion():
    """
    Endpoint principal que recibe peticiones del CRM y las encola
    """
    try:
        data = request.get_json()
        
        if not data:
            log_mensaje("❌ Request sin datos JSON")
            return jsonify({'error': 'No se recibieron datos'}), 400
        
        url = data.get('url')
        justificacion_id = data.get('justificacion_id')
        callback_url = data.get('callback_url')
        user_name = data.get('user_name', 'Usuario')
        tipo_analisis = data.get('tipo_analisis', 'web')  # 'web' o 'ecommerce'
        
        if not url:
            log_mensaje("❌ Request sin URL")
            return jsonify({'error': 'URL no proporcionada'}), 400
        
        if not callback_url:
            log_mensaje("❌ Request sin callback_url")
            return jsonify({'error': 'callback_url no proporcionado'}), 400
        
        log_mensaje("=" * 80)
        log_mensaje("🔔 NUEVA JUSTIFICACIÓN RECIBIDA")
        log_mensaje("=" * 80)
        log_mensaje(f"  📋 ID: {justificacion_id}")
        log_mensaje(f"  👤 Usuario: {user_name}")
        log_mensaje(f"  🔗 URL: {url}")
        log_mensaje(f"  📞 Callback: {callback_url}")
        log_mensaje(f"  🏷️ Tipo: {tipo_analisis.upper()}")
        log_mensaje(f"  ⏰ Timestamp: {data.get('timestamp')}")
        log_mensaje("=" * 80)
        
        # Crear trabajo y añadir a la cola
        trabajo = Trabajo(
            url=url,
            justificacion_id=justificacion_id,
            callback_url=callback_url,
            user_name=user_name,
            tipo_analisis=tipo_analisis,
            timestamp_recibido=datetime.now().isoformat()
        )
        
        cola_trabajos.put(trabajo)
        posicion_cola = cola_trabajos.qsize()
        
        # Actualizar estado en el CRM
        if posicion_cola > 1:
            actualizar_estado_crm(callback_url, 'en_cola', f'En cola - Posición: {posicion_cola}')
            log_mensaje(f"📥 Trabajo añadido a la cola - Posición: {posicion_cola}")
        else:
            actualizar_estado_crm(callback_url, 'en_cola', 'Procesando pronto...')
            log_mensaje(f"📥 Trabajo añadido a la cola - Se procesará inmediatamente")
        
        return jsonify({
            'success': True,
            'message': 'Justificación recibida y encolada',
            'justificacion_id': justificacion_id,
            'posicion_cola': posicion_cola,
            'estado': 'en_cola'
        }), 200
        
    except Exception as e:
        log_mensaje(f"❌ Error en endpoint /sgbasc: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/health', methods=['GET'])
def health_check():
    """Endpoint para verificar que el servidor está funcionando"""
    return jsonify({
        'status': 'ok',
        'service': 'Procesador de Justificaciones SEO - Hawkins CRM',
        'timestamp': datetime.now().isoformat()
    }), 200


@app.route('/stats', methods=['GET'])
def get_stats():
    """Endpoint para obtener estadísticas"""
    global procesando, trabajo_actual
    
    return jsonify({
        'status': 'activo',
        'servicio': 'Procesador SEO',
        'version': '1.0',
        'procesando': procesando,
        'trabajos_en_cola': cola_trabajos.qsize(),
        'trabajo_actual': {
            'id': trabajo_actual.justificacion_id,
            'url': trabajo_actual.url,
            'tipo': trabajo_actual.tipo_analisis
        } if trabajo_actual else None
    }), 200


@app.route('/cola', methods=['GET'])
def ver_cola():
    """Endpoint para ver el estado de la cola"""
    global procesando, trabajo_actual
    
    return jsonify({
        'procesando': procesando,
        'trabajos_en_cola': cola_trabajos.qsize(),
        'trabajo_actual': {
            'justificacion_id': trabajo_actual.justificacion_id,
            'url': trabajo_actual.url,
            'user_name': trabajo_actual.user_name,
            'tipo_analisis': trabajo_actual.tipo_analisis,
            'timestamp_recibido': trabajo_actual.timestamp_recibido
        } if trabajo_actual else None
    }), 200


if __name__ == '__main__':
    log_mensaje("=" * 80)
    log_mensaje("🚀 SERVIDOR DE JUSTIFICACIONES SEO INICIADO")
    log_mensaje("=" * 80)
    log_mensaje(f"📍 Escuchando en: http://0.0.0.0:5000")
    log_mensaje(f"📍 Endpoint principal: POST /sgbasc")
    log_mensaje(f"📍 Health check: GET /health")
    log_mensaje(f"📍 Cola de trabajos: GET /cola")
    log_mensaje(f"📍 Stats: GET /stats")
    log_mensaje("=" * 80)
    
    # Iniciar worker de la cola en un thread separado
    worker_thread = threading.Thread(target=procesar_cola_trabajos)
    worker_thread.daemon = True
    worker_thread.start()
    log_mensaje("✅ Worker de cola iniciado")
    
    # Iniciar servidor Flask
    app.run(
        host='0.0.0.0',
        port=5000,
        debug=False,
        threaded=True
    )

