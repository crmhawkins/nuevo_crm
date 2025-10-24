import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import threading
from datetime import datetime
import time
import os
import sys
import subprocess
import platform
import shutil
import requests
import json
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import pyautogui
from openpyxl import load_workbook
from urllib.parse import urlparse, urljoin
from app import generar_pdf_con_capturas
import imaplib
import email
import re
import time
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from PIL import Image


def obtener_directorio_base():
    """Obtiene el directorio base de la aplicación."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


def consultar_ia_local(prompt):
    """Consulta la IA local para análisis de keywords"""
    try:
        url = "https://aiapi.hawkins.es/chat/chat"
        headers = {
            "x-api-key": "OllamaAPI_2024_K8mN9pQ2rS5tU7vW3xY6zA1bC4eF8hJ0lM",
            "Content-Type": "application/json"
        }
        
        payload = {
            "modelo": "gpt-oss:120b-cloud",
            "prompt": prompt
        }
        
        print(f"🔗 Enviando petición a: {url}")
        print(f"⚠️  Deshabilitando verificación SSL para API externa")
        print(f"📋 Headers: {headers}")
        print(f"📦 Payload: {json.dumps(payload, indent=2)}")
        print(f"📏 Tamaño del prompt: {len(prompt)} caracteres")
        
        response = requests.post(url, headers=headers, json=payload, timeout=30, verify=False)
        
        print(f"📊 Status Code: {response.status_code}")
        print(f"📄 Response Headers: {dict(response.headers)}")
        
        if response.status_code == 200:
            try:
                data = response.json()
                print(f"✅ Respuesta JSON recibida: {json.dumps(data, indent=2)}")
                
                if data.get('success') and data.get('respuesta'):
                    print(f"🎯 Keywords extraídas por IA: {data['respuesta']}")
                    return data['respuesta']
                else:
                    print(f"❌ Respuesta sin éxito: success={data.get('success')}, respuesta={data.get('respuesta')}")
                    return None
            except json.JSONDecodeError as e:
                print(f"❌ Error parseando JSON: {str(e)}")
                print(f"📄 Respuesta raw: {response.text}")
                return None
        else:
            print(f"❌ Error HTTP {response.status_code}")
            print(f"📄 Respuesta: {response.text}")
            return None
            
    except requests.exceptions.Timeout:
        print("⏰ Timeout en la petición a la IA local")
        return None
    except requests.exceptions.ConnectionError as e:
        print(f"🔌 Error de conexión: {str(e)}")
        return None
    except requests.exceptions.RequestException as e:
        print(f"🌐 Error en la petición: {str(e)}")
        return None
    except Exception as e:
        print(f"💥 Error inesperado consultando IA local: {str(e)}")
        return None


def extraer_datos_web(driver, url):
    """Extrae datos estructurados de la página web para análisis"""
    try:
        print(f"📊 Extrayendo datos de la página: {url}")
        
        # Obtener elementos básicos
        title = ""
        meta_description = ""
        h1_tags = []
        h2_tags = []
        h3_tags = []
        content_sample = ""
        
        try:
            title = driver.find_element(By.TAG_NAME, 'title').text
            print(f"✅ Título extraído: {title}")
        except Exception as e:
            print(f"❌ Error extrayendo título: {str(e)}")
            pass
            
        try:
            meta_desc_element = driver.find_element(By.CSS_SELECTOR, 'meta[name="description"]')
            meta_description = meta_desc_element.get_attribute('content') or ""
            print(f"✅ Meta descripción extraída: {meta_description[:100]}...")
        except Exception as e:
            print(f"❌ Error extrayendo meta descripción: {str(e)}")
            pass
            
        try:
            h1_elements = driver.find_elements(By.TAG_NAME, 'h1')
            h1_tags = [h.text.strip() for h in h1_elements if h.text.strip()]
            print(f"✅ H1 tags extraídos: {len(h1_tags)} elementos")
        except Exception as e:
            print(f"❌ Error extrayendo H1: {str(e)}")
            pass
            
        try:
            h2_elements = driver.find_elements(By.TAG_NAME, 'h2')
            h2_tags = [h.text.strip() for h in h2_elements if h.text.strip()]
            print(f"✅ H2 tags extraídos: {len(h2_tags)} elementos")
        except Exception as e:
            print(f"❌ Error extrayendo H2: {str(e)}")
            pass
            
        try:
            h3_elements = driver.find_elements(By.TAG_NAME, 'h3')
            h3_tags = [h.text.strip() for h in h3_elements if h.text.strip()]
            print(f"✅ H3 tags extraídos: {len(h3_tags)} elementos")
        except Exception as e:
            print(f"❌ Error extrayendo H3: {str(e)}")
            pass
            
        try:
            body_element = driver.find_element(By.TAG_NAME, 'body')
            content_sample = body_element.text[:2000]  # Primeros 2000 caracteres
            print(f"✅ Contenido extraído: {len(content_sample)} caracteres")
        except Exception as e:
            print(f"❌ Error extrayendo contenido: {str(e)}")
            pass
        
        # Extraer palabras del dominio y URL
        domain = urlparse(url).netloc.replace('www.', '')
        url_words = []
        try:
            url_path = urlparse(url).path
            url_words = [word for word in url_path.split('/') if word and len(word) > 2]
            print(f"✅ Palabras URL extraídas: {url_words}")
        except Exception as e:
            print(f"❌ Error extrayendo palabras URL: {str(e)}")
            pass
        
        datos = {
            'url': url,
            'title': title,
            'meta_description': meta_description,
            'h1_tags': h1_tags,
            'h2_tags': h2_tags,
            'h3_tags': h3_tags,
            'domain': domain,
            'url_words': url_words,
            'content_sample': content_sample
        }
        
        print(f"✅ Datos completos extraídos exitosamente")
        return datos
        
    except Exception as e:
        print(f"💥 Error extrayendo datos web: {str(e)}")
        import traceback
        print(f"📋 Traceback: {traceback.format_exc()}")
        return None


def crear_prompt_keywords(datos_web):
    """Crea el prompt optimizado para la IA local"""
    prompt = f"""Analiza estos datos de una página web y extrae exactamente las 5 keywords más relevantes para SEO, separadas por comas.

DATOS DE LA PÁGINA WEB:
- URL: {datos_web['url']}
- Título: {datos_web['title']}
- Descripción Meta: {datos_web['meta_description']}
- Encabezados H1: {', '.join(datos_web['h1_tags'])}
- Encabezados H2: {', '.join(datos_web['h2_tags'])}
- Encabezados H3: {', '.join(datos_web['h3_tags'])}
- Dominio: {datos_web['domain']}
- Palabras de URL: {', '.join(datos_web['url_words'])}
- Muestra de contenido: {datos_web['content_sample'][:500]}...

INSTRUCCIONES:
1. Identifica las palabras clave más importantes para SEO
2. Prioriza términos del título y encabezados
3. Incluye palabras relevantes del dominio/URL
4. Evita palabras muy genéricas como "página", "web", "sitio"
5. Máximo 5 keywords
6. Respuesta SOLO con las keywords separadas por comas, sin explicaciones

FORMATO DE RESPUESTA: keyword1, keyword2, keyword3, keyword4, keyword5"""
    
    return prompt


def extraer_keywords_automaticas(driver, url):
    """Sistema híbrido para extraer keywords automáticamente"""
    try:
        print(f"🔍 Iniciando extracción automática de keywords de: {url}")
        
        # Paso 1: Extraer datos de la página web
        print("📊 Paso 1: Extrayendo datos de la página web...")
        datos_web = extraer_datos_web(driver, url)
        if not datos_web:
            print("❌ No se pudieron extraer datos de la página web")
            return None
        
        print(f"✅ Datos extraídos exitosamente:")
        print(f"   - Título: {datos_web['title']}")
        print(f"   - Meta descripción: {datos_web['meta_description'][:100]}...")
        print(f"   - H1 tags: {len(datos_web['h1_tags'])} elementos")
        print(f"   - H2 tags: {len(datos_web['h2_tags'])} elementos")
        print(f"   - H3 tags: {len(datos_web['h3_tags'])} elementos")
        print(f"   - Dominio: {datos_web['domain']}")
        print(f"   - Palabras URL: {datos_web['url_words']}")
        print(f"   - Contenido muestra: {len(datos_web['content_sample'])} caracteres")
            
        # Paso 2: Crear prompt para IA local
        print("📝 Paso 2: Creando prompt para IA local...")
        prompt = crear_prompt_keywords(datos_web)
        print(f"✅ Prompt creado: {len(prompt)} caracteres")
        print(f"📋 Preview del prompt: {prompt[:200]}...")
        
        # Paso 3: Consultar IA local
        print("🤖 Paso 3: Consultando IA local...")
        respuesta_ia = consultar_ia_local(prompt)
        
        if respuesta_ia:
            print(f"✅ Respuesta recibida de IA: {respuesta_ia}")
            # Paso 4: Procesar respuesta de la IA
            print("🔄 Paso 4: Procesando respuesta de la IA...")
            keywords = procesar_respuesta_ia(respuesta_ia)
            if keywords and len(keywords) > 0:
                print(f"✅ Keywords procesadas exitosamente: {keywords}")
                return keywords
            else:
                print("❌ No se pudieron procesar las keywords de la IA")
        else:
            print("❌ No se recibió respuesta de la IA")
        
        # Paso 5: Fallback a análisis básico si falla la IA
        print("⚠️ Paso 5: Ejecutando fallback a análisis básico...")
        keywords_fallback = analisis_basico_keywords(datos_web)
        if keywords_fallback:
            print(f"✅ Keywords fallback extraídas: {keywords_fallback}")
            return keywords_fallback
        else:
            print("❌ Fallback también falló")
            
        print("💥 Todos los métodos de extracción fallaron")
        return None
        
    except Exception as e:
        print(f"💥 Error inesperado en extracción automática: {str(e)}")
        import traceback
        print(f"📋 Traceback completo: {traceback.format_exc()}")
        return None


def procesar_respuesta_ia(respuesta):
    """Procesa la respuesta de la IA local"""
    try:
        print(f"🔄 Procesando respuesta de IA: '{respuesta}'")
        
        # Limpiar la respuesta
        respuesta_limpia = respuesta.strip()
        print(f"🧹 Respuesta limpia: '{respuesta_limpia}'")
        
        # Buscar keywords separadas por comas
        keywords = [kw.strip() for kw in respuesta_limpia.split(',')]
        print(f"📝 Keywords separadas: {keywords}")
        
        # Filtrar keywords válidas
        keywords_validas = []
        stop_words = ['página', 'web', 'sitio', 'home', 'inicio']
        
        for kw in keywords:
            kw = kw.strip()
            print(f"🔍 Evaluando keyword: '{kw}'")
            
            if len(kw) <= 2:
                print(f"   ❌ Muy corta: {len(kw)} caracteres")
                continue
            if len(kw) >= 50:
                print(f"   ❌ Muy larga: {len(kw)} caracteres")
                continue
            if kw.lower() in stop_words:
                print(f"   ❌ Palabra común: {kw}")
                continue
                
            print(f"   ✅ Keyword válida: {kw}")
            keywords_validas.append(kw)
        
        print(f"✅ Keywords válidas encontradas: {keywords_validas}")
        
        # Limitar a 5 keywords
        resultado = keywords_validas[:5]
        print(f"🎯 Resultado final (máximo 5): {resultado}")
        
        return resultado
        
    except Exception as e:
        print(f"💥 Error procesando respuesta IA: {str(e)}")
        import traceback
        print(f"📋 Traceback: {traceback.format_exc()}")
        return None


def analisis_basico_keywords(datos_web):
    """Análisis básico de keywords como fallback"""
    try:
        keywords = []
        
        # Extraer del título
        if datos_web['title']:
            title_words = re.findall(r'\b\w+\b', datos_web['title'].lower())
            keywords.extend([w for w in title_words if len(w) > 3])
        
        # Extraer de H1
        for h1 in datos_web['h1_tags']:
            h1_words = re.findall(r'\b\w+\b', h1.lower())
            keywords.extend([w for w in h1_words if len(w) > 3])
        
        # Extraer de H2
        for h2 in datos_web['h2_tags']:
            h2_words = re.findall(r'\b\w+\b', h2.lower())
            keywords.extend([w for w in h2_words if len(w) > 3])
        
        # Extraer del dominio
        domain_words = re.findall(r'\b\w+\b', datos_web['domain'].lower())
        keywords.extend([w for w in domain_words if len(w) > 3])
        
        # Contar frecuencia y seleccionar las más comunes
        from collections import Counter
        word_count = Counter(keywords)
        
        # Filtrar palabras muy comunes
        stop_words = {'página', 'web', 'sitio', 'home', 'inicio', 'contenido', 'información', 'servicios', 'productos'}
        filtered_words = {word: count for word, count in word_count.items() if word not in stop_words}
        
        # Obtener las 5 más frecuentes
        top_keywords = [word for word, count in Counter(filtered_words).most_common(5)]
        
        return top_keywords
        
    except Exception as e:
        print(f"Error análisis básico: {str(e)}")
        return None


def buscar_urls_en_pagina(driver, url_base):
    """Busca las URLs de 'Sobre Nosotros' y 'Contacto' en la página web"""
    sobre_nosotros_url = None
    contacto_url = None
    
    try:
        enlaces = driver.find_elements(By.TAG_NAME, "a")
        
        for enlace in enlaces:
            try:
                href = enlace.get_attribute("href")
                texto = enlace.text.lower().strip()
                
                if not href:
                    continue
                
                if not sobre_nosotros_url and any(p in texto for p in ["sobre nosotros", "sobre", "quiénes somos", "quienes somos", "about", "acerca de"]):
                    sobre_nosotros_url = href
                
                if not contacto_url and any(p in texto for p in ["contacto", "contáctanos", "contactanos", "contact"]):
                    contacto_url = href
                
                if sobre_nosotros_url and contacto_url:
                    break
            except:
                continue
        
        if not sobre_nosotros_url or not contacto_url:
            for enlace in enlaces:
                try:
                    href = enlace.get_attribute("href")
                    if not href:
                        continue
                    
                    href_lower = href.lower()
                    
                    if not sobre_nosotros_url and any(p in href_lower for p in ["sobre", "about", "acerca", "quienes-somos"]):
                        sobre_nosotros_url = href
                    
                    if not contacto_url and any(p in href_lower for p in ["contacto", "contact"]):
                        contacto_url = href
                    
                    if sobre_nosotros_url and contacto_url:
                        break
                except:
                    continue
    
    except Exception as e:
        print(f"Error buscando URLs: {str(e)}")
    
    if not sobre_nosotros_url:
        sobre_nosotros_url = urljoin(url_base, "/sobre-nosotros")
    if not contacto_url:
        contacto_url = urljoin(url_base, "/contacto")
    
    return sobre_nosotros_url, contacto_url


def modificar_excel_informe(url_web, nombre_web=None):
    """Crea una copia del archivo Excel y la modifica con los datos de la URL."""
    try:
        base_dir = obtener_directorio_base()
        archivo_excel_original = os.path.join(base_dir, 'Informe_de_revision.xlsx')
        
        if not os.path.exists(archivo_excel_original):
            print(f"Archivo Excel no encontrado: {archivo_excel_original}")
            return None
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_copia = f'Informe_de_revision_{timestamp}.xlsx'
        archivo_excel_copia = os.path.join(base_dir, nombre_copia)
        
        shutil.copy2(archivo_excel_original, archivo_excel_copia)
        wb = load_workbook(archivo_excel_copia)
        hojas = wb.worksheets
        
        if len(hojas) < 4:
            print("El archivo Excel no tiene suficientes hojas")
            return None
        
        pagina_2 = hojas[1]
        
        if not nombre_web:
            nombre_web = "Sitio Web"
        
        # Obtener fecha de hoy en formato DD/MM/AAAA
        fecha_hoy = datetime.now().strftime('%d/%m/%Y')
        
        pagina_2['C15'] = nombre_web
        pagina_2['C17'] = url_web
        pagina_2['C21'] = fecha_hoy
        wb.save(archivo_excel_copia)
        return archivo_excel_copia
        
    except Exception as e:
        print(f"Error modificando Excel: {str(e)}")
        return None


def modificar_excel_urls_adicionales(archivo_excel_copia, url_base, sobre_nosotros_url, contacto_url):
    """Modifica la página 4 del Excel copiado con las URLs adicionales y slugs personalizados."""
    try:
        if not archivo_excel_copia or not os.path.exists(archivo_excel_copia):
            return False
        
        wb = load_workbook(archivo_excel_copia)
        hojas = wb.worksheets
        
        if len(hojas) < 4:
            return False
        
        pagina_4 = hojas[3]
        
        # URLs principales
        pagina_4['F8'] = url_base
        pagina_4['F9'] = url_base
        pagina_4['F10'] = sobre_nosotros_url
        pagina_4['F11'] = contacto_url
        
        # Generar slugs personalizados basados en las URLs
        def generar_slug_personalizado(url):
            """Genera un slug personalizado basado en la URL"""
            if not url:
                return ""
            
            # Extraer la ruta de la URL
            parsed_url = urlparse(url)
            path = parsed_url.path.lower().strip('/')
            
            # Mapeo de rutas comunes a slugs personalizados
            slug_mapping = {
                'sobre-nosotros': 'Sobre Nosotros',
                'sobre': 'Sobre Nosotros', 
                'nosotros': 'Nosotros',
                'sobre-mi': 'Sobre Mí',
                'quienes-somos': 'Quiénes Somos',
                'acerca-de': 'Acerca de',
                'about': 'Sobre Nosotros',
                'contacto': 'Contacto',
                'contactanos': 'Contáctanos',
                'contact': 'Contacto',
                'contacto.html': 'Contacto',
                'contactanos.html': 'Contáctanos'
            }
            
            # Buscar coincidencia exacta
            if path in slug_mapping:
                return slug_mapping[path]
            
            # Buscar coincidencia parcial
            for key, value in slug_mapping.items():
                if key in path:
                    return value
            
            # Si no hay coincidencia, usar la ruta capitalizada
            if path:
                return path.replace('-', ' ').replace('_', ' ').title()
            
            return "Página"
        
        # Generar slugs para las páginas
        slug_sobre_nosotros = generar_slug_personalizado(sobre_nosotros_url)
        slug_contacto = generar_slug_personalizado(contacto_url)
        
        # Determinar el tipo de página
        def determinar_tipo_pagina(url):
            """Determina el tipo de página basado en la URL"""
            if not url:
                return "Otras páginas"
            
            parsed_url = urlparse(url)
            path = parsed_url.path.lower().strip('/')
            
            if any(keyword in path for keyword in ['contacto', 'contact', 'contactanos']):
                return "Contacto"
            elif any(keyword in path for keyword in ['sobre', 'nosotros', 'about', 'acerca', 'quienes']):
                return "Acerca de"
            else:
                return "Otras páginas"
        
        tipo_sobre_nosotros = determinar_tipo_pagina(sobre_nosotros_url)
        tipo_contacto = determinar_tipo_pagina(contacto_url)
        
        # Generar breadcrumbs
        def generar_breadcrumbs(slug):
            """Genera breadcrumbs basados en el slug"""
            return f"Inicio > {slug}"
        
        # Modificar las celdas específicas
        # Fila 10 - Sobre Nosotros
        pagina_4['C10'] = "Página Web"
        pagina_4['E10'] = tipo_sobre_nosotros
        pagina_4['G10'] = generar_breadcrumbs(slug_sobre_nosotros)
        
        # Fila 11 - Contacto  
        pagina_4['C11'] = "Página Web"
        pagina_4['E11'] = tipo_contacto
        pagina_4['G11'] = generar_breadcrumbs(slug_contacto)
        
        wb.save(archivo_excel_copia)
        return True
    
    except Exception as e:
        print(f"Error modificando Excel (URLs adicionales): {str(e)}")
        return False


class SEOAnalyzerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Analizador SEO - Rankalyze")
        self.root.geometry("900x750")
        self.root.resizable(True, True)
        
        self.base_dir = obtener_directorio_base()
        self.screenshots_dir = os.path.join(self.base_dir, 'screenshots')
        self.pdfs_dir = os.path.join(self.base_dir, 'pdfs_generados')
        os.makedirs(self.screenshots_dir, exist_ok=True)
        os.makedirs(self.pdfs_dir, exist_ok=True)
        self.ultimo_pdf = None
        
        style = ttk.Style()
        style.theme_use('clam')
        
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)
        
        main_frame = ttk.Frame(root, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        main_frame.columnconfigure(0, weight=1)
        
        title_label = tk.Label(main_frame, text="🔍 Analizador SEO Automático", font=("Helvetica", 16, "bold"), fg="#1a5490")
        title_label.grid(row=0, column=0, pady=(0, 15), sticky=tk.W + tk.E)
        
        input_frame = ttk.LabelFrame(main_frame, text="Datos del Análisis", padding="15")
        input_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        input_frame.columnconfigure(1, weight=1)
        
        tk.Label(input_frame, text="URL:", font=("Helvetica", 9, "bold")).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.url_entry = ttk.Entry(input_frame, width=70, font=("Helvetica", 9))
        self.url_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        self.url_entry.insert(0, "")
        
        tk.Label(input_frame, text="Keywords (máx 5):", font=("Helvetica", 9, "bold")).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.keywords_entry = ttk.Entry(input_frame, width=70, font=("Helvetica", 9))
        self.keywords_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        tk.Label(input_frame, text="Competidores (máx 2):", font=("Helvetica", 9, "bold")).grid(row=2, column=0, sticky=tk.W, pady=5)
        self.competidores_entry = ttk.Entry(input_frame, width=70, font=("Helvetica", 9))
        self.competidores_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        self.ecommerce_var = tk.BooleanVar()
        ecommerce_check = ttk.Checkbutton(input_frame, text="¿Es un ecommerce? (incluye análisis de competencia)", variable=self.ecommerce_var)
        ecommerce_check.grid(row=3, column=1, sticky=tk.W, pady=10, padx=(10, 0))
        
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, pady=15)
        
        self.start_button = tk.Button(button_frame, text="▶ Iniciar Análisis", command=self.iniciar_analisis, bg="#1a5490", fg="white", font=("Helvetica", 11, "bold"), padx=30, pady=12, cursor="hand2", relief=tk.RAISED, bd=2)
        self.start_button.pack(side=tk.LEFT, padx=5)
        
        self.open_pdf_button = tk.Button(button_frame, text="📄 Abrir Último PDF", command=self.abrir_ultimo_pdf, bg="#28a745", fg="white", font=("Helvetica", 11, "bold"), padx=20, pady=12, cursor="hand2", relief=tk.RAISED, bd=2, state=tk.DISABLED)
        self.open_pdf_button.pack(side=tk.LEFT, padx=5)
        
        progress_frame = ttk.LabelFrame(main_frame, text="Progreso", padding="10")
        progress_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)
        
        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        
        self.status_label = tk.Label(progress_frame, text="Listo para iniciar", font=("Helvetica", 9), fg="#666")
        self.status_label.grid(row=1, column=0)
        
        log_frame = ttk.LabelFrame(main_frame, text="Log del Proceso", padding="10")
        log_frame.grid(row=4, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(4, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, width=80, height=18, bg="#f5f5f5", fg="#333", font=("Consolas", 8))
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.buscar_ultimo_pdf()
    
    def buscar_ultimo_pdf(self):
        try:
            if os.path.exists(self.pdfs_dir):
                pdfs = [f for f in os.listdir(self.pdfs_dir) if f.endswith('.pdf')]
                if pdfs:
                    pdfs_completos = [os.path.join(self.pdfs_dir, f) for f in pdfs]
                    ultimo = max(pdfs_completos, key=os.path.getmtime)
                    if os.path.exists(ultimo):
                        self.ultimo_pdf = os.path.abspath(ultimo)
                        self.open_pdf_button.config(state=tk.NORMAL)
        except Exception as e:
            print(f"Error buscando PDF: {e}")
    
    def log(self, mensaje):
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.log_text.insert(tk.END, f"[{timestamp}] {mensaje}\n")
        self.log_text.see(tk.END)
        self.root.update()
    
    def set_status(self, status):
        self.status_label.config(text=status)
        self.root.update()
    
    def abrir_ultimo_pdf(self):
        self.buscar_ultimo_pdf()
        
        if not self.ultimo_pdf:
            messagebox.showerror("Error", "No hay ningún PDF generado aún")
            return
        
        pdf_path = os.path.abspath(self.ultimo_pdf)
        if not os.path.exists(pdf_path):
            messagebox.showerror("Error", f"El PDF no existe:\n{pdf_path}")
            return
        
        try:
            self.log(f"\n📂 Abriendo PDF: {os.path.basename(pdf_path)}")
            if platform.system() == 'Windows':
                os.startfile(pdf_path)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', pdf_path])
            else:
                subprocess.run(['xdg-open', pdf_path])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el PDF:\n{str(e)}")
    
    def normalizar_url(self, url):
        """Normaliza la URL agregando https:// si no lo tiene y quitando slash final"""
        if not url:
            return ""
        
        url = url.strip()
        
        # Si no tiene protocolo, agregar https://
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        
        # Quitar slash final si existe
        if url.endswith('/'):
            url = url[:-1]
        
        return url
    
    def obtener_codigo_verificacion_email(self):
        """Obtiene el código de verificación del último email de Dondominio"""
        try:
            self.log("  📧 Conectando a servidor de email...")
            
            # Configuración del servidor IMAP de Ionos
            mail = imaplib.IMAP4_SSL('imap.ionos.es', 993)
            mail.login('dondominio@hawkins.es', 'R4t4-2025')
            mail.select('inbox')
            
            self.log("  📧 Buscando emails de Dondominio...")
            
            # Buscar emails de info@dondominio.com
            status, messages = mail.search(None, 'FROM', 'info@dondominio.com')
            email_ids = messages[0].split()
            
            if not email_ids:
                self.log("  ❌ No se encontraron emails de Dondominio")
                return None
            
            # Obtener el email más reciente
            latest_email_id = email_ids[-1]
            status, msg_data = mail.fetch(latest_email_id, '(RFC822)')
            email_body = msg_data[0][1]
            email_message = email.message_from_bytes(email_body)
            
            # Extraer el contenido del email
            email_content = ""
            if email_message.is_multipart():
                for part in email_message.walk():
                    if part.get_content_type() == "text/plain":
                        email_content = part.get_payload(decode=True).decode()
                        break
            else:
                email_content = email_message.get_payload(decode=True).decode()
            
            self.log("  📧 Email obtenido, extrayendo código...")
            
            # Buscar código de 6 dígitos en el email
            codigo_pattern = r'\b(\d{6})\b'
            matches = re.findall(codigo_pattern, email_content)
            
            if matches:
                codigo = matches[-1]  # Tomar el último código encontrado
                self.log(f"  ✅ Código de verificación encontrado: {codigo}")
                return codigo
            else:
                self.log("  ❌ No se encontró código de 6 dígitos en el email")
                self.log(f"  📧 Contenido del email: {email_content[:200]}...")
                return None
                
        except Exception as e:
            self.log(f"  ❌ Error obteniendo código de email: {str(e)[:50]}")
            return None
        finally:
            try:
                mail.close()
                mail.logout()
            except:
                pass
    
    def iniciar_analisis(self):
        url_raw = self.url_entry.get().strip()
        if not url_raw:
            messagebox.showerror("Error", "Por favor ingresa una URL válida")
            return
        
        # Normalizar la URL
        url = self.normalizar_url(url_raw)
        self.log(f"URL normalizada: {url}")
        self.log("=" * 60)
        self.log("🚀 INICIANDO NUEVO ANÁLISIS")
        self.log("=" * 60)
        
        keywords_raw = self.keywords_entry.get().strip()
        
        # Si no hay keywords o son valores por defecto, usar extracción automática
        if not keywords_raw or keywords_raw in ['keyword1, keyword2, keyword3', 'keyword1,keyword2,keyword3']:
            self.log("🔍 Extrayendo keywords automáticamente...")
            self.log("⚠️ Esto puede tomar unos segundos...")
            self.root.update()
            
            # Crear un driver temporal para extraer keywords
            try:
                chrome_options = Options()
                chrome_options.add_argument('--headless')  # Modo headless para extracción
                chrome_options.add_argument('--no-sandbox')
                chrome_options.add_argument('--disable-dev-shm-usage')
                
                driver_temp = webdriver.Chrome(options=chrome_options)
                driver_temp.get(url)
                time.sleep(2)
                
                keywords = extraer_keywords_automaticas(driver_temp, url)
                
                if keywords and len(keywords) > 0:
                    self.log(f"✅ Keywords extraídas: {', '.join(keywords)}")
                    # Actualizar el campo de keywords en la GUI
                    self.keywords_entry.delete(0, tk.END)
                    self.keywords_entry.insert(0, ', '.join(keywords))
                    
                    # Generar competidores automáticamente solo si el checkbox de ecommerce está activo
                    checkbox_activo = self.ecommerce_var.get()
                    self.log(f"🔍 Estado del checkbox ecommerce: {checkbox_activo}")
                    
                    if checkbox_activo:
                        self.log("🤖 Generando competidores automáticamente (ecommerce activo)...")
                        competidores_generados = self.buscar_competidores_simple(driver_temp, keywords, url)
                        
                        if competidores_generados and len(competidores_generados) >= 2:
                            competidores_para_gui = competidores_generados[:2]
                            self.log(f"✅ Competidores generados: {', '.join(competidores_para_gui)}")
                            self.log(f"📝 Actualizando campo de competidores en la GUI...")
                            # Actualizar el campo de competidores en la GUI
                            self.competidores_entry.delete(0, tk.END)
                            self.competidores_entry.insert(0, ', '.join(competidores_para_gui))
                            self.log(f"✅ Campo de competidores actualizado con: {', '.join(competidores_para_gui)}")
                        else:
                            self.log("⚠️ No se pudieron generar competidores automáticamente")
                    else:
                        self.log("ℹ️ Checkbox de ecommerce no activo - no se generan competidores automáticamente")
                    
                else:
                    messagebox.showerror("Error", "No se pudieron extraer keywords automáticamente. Por favor ingrésalas manualmente.")
                    return
                
                driver_temp.quit()
                    
            except Exception as e:
                self.log(f"❌ Error extrayendo keywords: {str(e)}")
                messagebox.showerror("Error", "Error extrayendo keywords automáticamente. Por favor ingrésalas manualmente.")
                return
        else:
            keywords = [k.strip() for k in keywords_raw.split(',') if k.strip()][:5]
        
        competidores_raw = self.competidores_entry.get().strip()
        competidores = [c.strip() for c in competidores_raw.split(',') if c.strip()][:2] if competidores_raw else []
        
        # Log de competidores ingresados
        self.log(f"📋 Competidores ingresados: {len(competidores)}")
        if competidores:
            for i, comp in enumerate(competidores, 1):
                self.log(f"  {i}. {comp}")
        else:
            self.log("  ⚠️ No hay competidores ingresados")
        
        if competidores:
            competidores = [c for c in competidores if 'competidor' not in c.lower() and c != 'https://']
            self.log(f"📋 Competidores válidos después del filtrado: {len(competidores)}")
            if competidores:
                for i, comp in enumerate(competidores, 1):
                    self.log(f"  {i}. {comp}")
        
        es_ecommerce = self.ecommerce_var.get()
        
        self.start_button.config(state=tk.DISABLED)
        self.progress.start()
        
        self.log("=" * 60)
        self.log("📊 INICIANDO PROCESO PRINCIPAL")
        self.log("=" * 60)
        
        thread = threading.Thread(target=self.ejecutar_proceso, args=(url, keywords, competidores, es_ecommerce))
        thread.daemon = True
        thread.start()
    
    def ejecutar_proceso(self, url, keywords, competidores, es_ecommerce=False):
        driver = None
        try:
            self.log("=" * 50)
            self.log("🚀 INICIANDO ANÁLISIS SEO")
            if es_ecommerce:
                self.log("🛒 ECOMMERCE - Incluye análisis de competencia")
            self.log("=" * 50)
            self.set_status("Configurando Chrome...")
            
            chrome_options = Options()
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            prefs = {"credentials_enable_service": False, "profile.password_manager_enabled": False}
            chrome_options.add_experimental_option("prefs", prefs)
            chrome_options.add_argument('--start-maximized')
            
            self.log("✓ Chrome configurado")
            
            driver = webdriver.Chrome(options=chrome_options)
            driver.maximize_window()
            wait = WebDriverWait(driver, 20)
            
            # ===== RANKALYZE - ANÁLISIS COMPLETO =====
            
            self.set_status("Login...")
            self.log("\n" + "=" * 50)
            self.log("📧 LOGIN EN HERRAMIENTA SEO")
            self.log("=" * 50)
            driver.get('https://rankalyze.net/login')
            time.sleep(2)
            
            username_input = wait.until(EC.element_to_be_clickable((By.NAME, 'email')))
            username_input.clear()
            time.sleep(0.3)
            username_input.send_keys('admin@admin.com')
            time.sleep(0.3)
            
            password_input = wait.until(EC.element_to_be_clickable((By.NAME, 'password')))
            password_input.clear()
            time.sleep(0.3)
            password_input.send_keys('R4t4-2020')
            time.sleep(0.3)
            
            submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="submitButton"]')))
            submit_button.click()
            time.sleep(3)
            
            current_url = driver.current_url
            if 'login' in current_url.lower():
                raise Exception("Login falló - Verificar credenciales")
            
            self.log("✓ Login OK")
            
            self.set_status("Formulario...")
            self.log("\n" + "=" * 50)
            self.log("📝 LLENANDO FORMULARIO")
            self.log("=" * 50)
            se_form_link = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/aside/nav/ul/li[2]/a')))
            se_form_link.click()
            time.sleep(2)
            
            random_btn = wait.until(EC.element_to_be_clickable((By.ID, 'randomBtn')))
            random_btn.click()
            alert = wait.until(EC.alert_is_present())
            alert.accept()
            self.log("✓ Alert OK")
            
            self.set_status("Keywords...")
            self.log(f"\n🔑 {len(keywords)} keywords...")
            for i, keyword in enumerate(keywords):
                keyword_input = wait.until(EC.element_to_be_clickable((By.ID, f'palabra_{i}')))
                keyword_input.clear()
                time.sleep(0.2)
                keyword_input.send_keys(Keys.CONTROL + "a")
                keyword_input.send_keys(Keys.DELETE)
                time.sleep(0.2)
                keyword_input.send_keys(keyword)
                time.sleep(0.2)
                self.log(f"  ✓ {i + 1}")
            
            if competidores:
                self.log(f"🏢 Competidores...")
                self.log(f"  📋 Enviando {len(competidores)} competidores a la herramienta SEO:")
                for i, comp in enumerate(competidores, 1):
                    self.log(f"    {i}. {comp}")
            
            if len(competidores) > 0:
                self.log(f"  📝 Llenando competitor1_url con: {competidores[0]}")
                c1 = wait.until(EC.element_to_be_clickable((By.ID, 'competitor1_url')))
                c1.clear()
                time.sleep(0.2)
                c1.send_keys(Keys.CONTROL + "a")
                c1.send_keys(Keys.DELETE)
                time.sleep(0.2)
                c1.send_keys(competidores[0])
                self.log("  ✅ Competitor 1 enviado")
            
            if len(competidores) > 1:
                self.log(f"  📝 Llenando competitor2_url con: {competidores[1]}")
                c2 = wait.until(EC.element_to_be_clickable((By.ID, 'competitor2_url')))
                c2.clear()
                time.sleep(0.2)
                c2.send_keys(Keys.CONTROL + "a")
                c2.send_keys(Keys.DELETE)
                time.sleep(0.2)
                c2.send_keys(competidores[1])
                self.log("  ✅ Competitor 2 enviado")
                self.log("✓ OK")
            
            fecha_hoy = datetime.now().strftime('%d/%m/%Y')
            fecha_input = wait.until(EC.element_to_be_clickable((By.ID, 'fecha_inicio')))
            fecha_input.clear()
            time.sleep(0.2)
            fecha_input.send_keys(Keys.CONTROL + "a")
            fecha_input.send_keys(Keys.DELETE)
            time.sleep(0.2)
            fecha_input.send_keys(fecha_hoy)
            
            url_input = wait.until(EC.element_to_be_clickable((By.ID, 'url')))
            url_input.clear()
            time.sleep(0.2)
            url_input.send_keys(Keys.CONTROL + "a")
            url_input.send_keys(Keys.DELETE)
            time.sleep(0.2)
            url_input.send_keys(url)
            self.set_status("Enviando...")
            submit_final_btn = wait.until(EC.element_to_be_clickable((By.ID, 'submitBtn')))
            submit_final_btn.click()
            self.log("✓ Enviado")
            time.sleep(10)
            
            # ===== CAPTURAS DE RANKALYZE =====
            
            screenshots_dir = self.screenshots_dir
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # 1. Capturar Vista General
            self.set_status("Vista General...")
            self.log(f"\n📊 Vista General...")
            vista_general_screenshot = None
            try:
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(0.5)
                vista_general_filename = f'{screenshots_dir}/vista_general_{timestamp}.png'
                driver.save_screenshot(vista_general_filename)
                vista_general_screenshot = vista_general_filename
                self.log("  ✓ Vista General")
            except Exception as e:
                self.log(f"  ✗ Vista General: {str(e)[:30]}")
            
            # 2. Capturar Indexación
            self.set_status("Indexación...")
            self.log(f"\n📊 Indexación...")
            indexacion_screenshots = []
            try:
                index1_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/main/div[4]/div[1]')))
                driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", index1_element)
                driver.execute_script("window.scrollBy(0, -100);")
                time.sleep(0.3)
                index1_filename = f'{screenshots_dir}/indexacion_1_{timestamp}.png'
                index1_element.screenshot(index1_filename)
                indexacion_screenshots.append(index1_filename)
                self.log("  ✓ Indexación 1")
            except Exception as e:
                self.log(f"  ✗ Indexación 1: {str(e)[:30]}")
            
            try:
                index2_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/main/div[4]/div[2]')))
                driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", index2_element)
                driver.execute_script("window.scrollBy(0, -100);")
                time.sleep(0.3)
                index2_filename = f'{screenshots_dir}/indexacion_2_{timestamp}.png'
                index2_element.screenshot(index2_filename)
                indexacion_screenshots.append(index2_filename)
                self.log("  ✓ Indexación 2")
            except Exception as e:
                self.log(f"  ✗ Indexación 2: {str(e)[:30]}")
            
            # 3. Capturar Vista General (pantalla completa de Rankalyze)
            self.set_status("Vista General...")
            self.log(f"\n📊 Vista General...")
            vista_general_screenshot = None
            try:
                # Scroll al inicio de la página
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(0.5)
                
                # Captura de pantalla completa de la página
                vista_general_filename = f'{screenshots_dir}/vista_general_{timestamp}.png'
                driver.save_screenshot(vista_general_filename)
                vista_general_screenshot = vista_general_filename
                self.log("  ✓ Vista General capturada")
            except Exception as e:
                self.log(f"  ✗ Vista General: {str(e)[:30]}")
            
            # 4. Capturar Keywords
            self.set_status("Capturas...")
            self.log(f"\n📸 Keywords...")
            keyword_tabs = driver.find_elements(By.CLASS_NAME, 'keyword-tab')
            self.log(f"{len(keyword_tabs)} tabs")
            
            screenshots_guardadas = []
            for index, tab in enumerate(keyword_tabs):
                try:
                    self.log(f"  {index + 1}/{len(keyword_tabs)}...")
                    keyword_text = tab.text.strip().replace(' ', '_').replace('/', '_')
                    data_keyword = tab.get_attribute('data-keyword')
                    
                    tab.click()
                    
                    wait_fast = WebDriverWait(driver, 5)
                    try:
                        keywords_section = wait_fast.until(EC.visibility_of_element_located((By.ID, 'keywords-section')))
                    except:
                        keywords_section = wait_fast.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.keywords-section')))
                    
                    driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", keywords_section)
                    driver.execute_script("window.scrollBy(0, -100);")
                    
                    filename = f'{screenshots_dir}/keyword_{data_keyword}_{keyword_text}_{timestamp}.png'
                    keywords_section.screenshot(filename)
                    screenshots_guardadas.append(filename)
                    self.log("    ✓")
                except Exception as e:
                    self.log(f"    ✗ {str(e)[:30]}")
                    continue
            
            self.log(f"Total: {len(screenshots_guardadas)}")
            
            # # 5. Capturar Indexación (elementos /html/body/div/main/div[4]/div[1] y /html/body/div/main/div[4]/div[2])
            self.set_status("Indexación...")
            self.log(f"\n📊 Indexación...")
            indexacion_screenshots = []
            try:
                # Captura 1: /html/body/div/main/div[4]/div[1]
                index1_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/main/div[4]/div[1]')))
                driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", index1_element)
                driver.execute_script("window.scrollBy(0, -100);")
                time.sleep(0.3)
                index1_filename = f'{screenshots_dir}/indexacion_1_{timestamp}.png'
                index1_element.screenshot(index1_filename)
                indexacion_screenshots.append(index1_filename)
                self.log("  ✓ Indexación 1")
            except Exception as e:
                self.log(f"  ✗ Indexación 1: {str(e)[:30]}")
            
            try:
                # Captura 2: /html/body/div/main/div[4]/div[2]
                index2_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/main/div[4]/div[2]')))
                driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", index2_element)
                driver.execute_script("window.scrollBy(0, -100);")
                time.sleep(0.3)
                index2_filename = f'{screenshots_dir}/indexacion_2_{timestamp}.png'
                index2_element.screenshot(index2_filename)
                indexacion_screenshots.append(index2_filename)
                self.log("  ✓ Indexación 2")
            except Exception as e:
                self.log(f"  ✗ Indexación 2: {str(e)[:30]}")
            
            competencia_screenshots = []
            metrics_screenshot = None  # Se mantiene como None ya que no se captura
            analisis_competencia_ia = None
            
            if es_ecommerce:
                # ===== CAPTURAS DE COMPETENCIA (ANTES DE SALIR DE RANKALYZE) =====
                self.log(f"\n🏆 Capturas de Competencia...")
                try:
                    comp1_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/main/div[5]/div[1]')))
                    driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", comp1_element)
                    driver.execute_script("window.scrollBy(0, -100);")
                    time.sleep(0.3)
                    comp1_filename = f'{screenshots_dir}/competencia_1_{timestamp}.png'
                    comp1_element.screenshot(comp1_filename)
                    competencia_screenshots.append(comp1_filename)
                    self.log(f"  ✓ Comp 1: {comp1_filename}")
                except Exception as e:
                    self.log(f"  ✗ Comp 1: {str(e)[:30]}")
                
                try:
                    comp2_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/main/div[5]/div[2]')))
                    driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", comp2_element)
                    driver.execute_script("window.scrollBy(0, -100);")
                    time.sleep(0.3)
                    comp2_filename = f'{screenshots_dir}/competencia_2_{timestamp}.png'
                    comp2_element.screenshot(comp2_filename)
                    competencia_screenshots.append(comp2_filename)
                    self.log(f"  ✓ Comp 2: {comp2_filename}")
                except Exception as e:
                    self.log(f"  ✗ Comp 2: {str(e)[:30]}")
                
                self.log(f"  📊 Total capturas de competencia: {len(competencia_screenshots)}")
                
                self.log(f"\n📊 Métricas...")
                try:
                    metrics_element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'metrics-section')))
                    driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", metrics_element)
                    driver.execute_script("window.scrollBy(0, -100);")
                    time.sleep(0.3)
                    metrics_filename = f'{screenshots_dir}/metrics_section_{timestamp}.png'
                    metrics_element.screenshot(metrics_filename)
                    metrics_screenshot = metrics_filename
                    self.log("  ✓")
                except:
                    self.log("  ✗")
                
                # ===== ANÁLISIS DE COMPETENCIA CON IA (DESPUÉS DE LAS CAPTURAS) =====
                self.log(f"\n🤖 Análisis de Competencia con IA...")
                analisis_competencia_ia = self.analizar_competencia_con_ia(url, keywords, driver)
            
            time.sleep(1)
            
            whois_screenshot = None
            self.set_status("WHOIS...")
            self.log(f"\n🌐 WHOIS...")
            
            url_sin_protocolo = url.replace('https://', '').replace('http://', '').replace('www.', '').split('/')[0].rstrip('/')
            is_dominio_es = url_sin_protocolo.endswith('.es')
            
            if is_dominio_es:
                self.log("  Dominio .es")
                nombre_dominio = url_sin_protocolo.replace('.es', '')
                try:
                    search_url = f'https://nic.es/sgnd/dominio/publicBuscarDominios.action?tDominio.nombreDominio={nombre_dominio}&flag=activado'
                    driver.get(search_url)
                    time.sleep(3)
                    ver_datos_btn = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'verDatosNRGTM')))
                    ver_datos_btn.click()
                    time.sleep(2)
                    buscar_button = wait.until(EC.element_to_be_clickable((By.NAME, 'Buscar')))
                    buscar_button.click()
                    time.sleep(3)
                    target_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.center-on-page-plantillaBase.pageBase')))
                    driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", target_element)
                    whois_filename = f'{screenshots_dir}/whois_domain_es_{timestamp}.png'
                    target_element.screenshot(whois_filename)
                    whois_screenshot = whois_filename
                    self.log("  ✓")
                except Exception as e:
                    self.log(f"  ✗ {str(e)[:30]}")
                else:
                    self.log("  Internacional")
                    try:
                        whois_url = f'https://www.whois.com/whois/{url_sin_protocolo}'
                        driver.get(whois_url)
                        time.sleep(4)
                        whois_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/main')))
                        driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", whois_element)
                        driver.execute_script("window.scrollBy(0, -50);")
                        whois_filename = f'{screenshots_dir}/whois_domain_{timestamp}.png'
                        whois_element.screenshot(whois_filename)
                        whois_screenshot = whois_filename
                        self.log("  ✓")
                    except Exception as e:
                        self.log(f"  ✗ {str(e)[:30]}")
            
            h1_screenshot = None
            h2_screenshot = None
            self.set_status("H1/H2...")
            self.log(f"\n🏷️ H1/H2...")
            
            try:
                driver.get(url)
                time.sleep(3)
                
                self.log("\n📊 Modificando Excel (Página 2)...")
                nombre_web = self.obtener_nombre_web_con_ia_gui(url)
                excel_copia = modificar_excel_informe(url, nombre_web)
                self.log("  ✓ Excel Pág 2")
                
                self.log("🔍 Buscando URLs...")
                sobre_nosotros_url, contacto_url = buscar_urls_en_pagina(driver, url)
                self.log(f"  ✓ Sobre Nosotros: {sobre_nosotros_url}")
                self.log(f"  ✓ Contacto: {contacto_url}")
                
                if excel_copia:
                    self.log("📊 Modificando Excel (Página 4)...")
                    modificar_excel_urls_adicionales(excel_copia, url, sobre_nosotros_url, contacto_url)
                    self.log("  ✓ Excel Pág 4")
                
                try:
                    h1_element = driver.find_element(By.TAG_NAME, 'h1')
                    h1_html = h1_element.get_attribute('outerHTML')
                    h1_text = h1_element.text
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", h1_element)
                    driver.execute_script("""
                        arguments[0].style.outline = '3px solid red';
                        arguments[0].style.backgroundColor = 'rgba(255, 255, 0, 0.3)';
                        var p = document.createElement('div');
                        p.id = 'custom-inspector-h1';
                        p.style.cssText = 'position: fixed; right: 20px; top: 80px; width: 400px; max-height: 600px; background: #282828; color: #fff; padding: 20px; border-radius: 8px; box-shadow: 0 8px 32px rgba(0,0,0,0.5); font-family: Courier New, monospace; font-size: 13px; z-index: 999999; overflow-y: auto; display: block !important; visibility: visible !important;';
                        var t = document.createElement('div');
                        t.style.cssText = 'color: #61dafb; font-weight: bold; margin-bottom: 15px; font-size: 14px; border-bottom: 2px solid #61dafb; padding-bottom: 8px;';
                        t.textContent = 'Inspector - H1';
                        p.appendChild(t);
                        var c = document.createElement('pre');
                        c.style.cssText = 'margin: 0; white-space: pre-wrap; word-wrap: break-word; color: #e06c75;';
                        c.textContent = arguments[1];
                        p.appendChild(c);
                        var i = document.createElement('div');
                        i.style.cssText = 'margin-top: 15px; padding-top: 15px; border-top: 1px solid #444; color: #98c379;';
                        i.innerHTML = '<strong>Texto:</strong> ' + arguments[2];
                        p.appendChild(i);
                        document.body.appendChild(p);
                    """, h1_element, h1_html, h1_text)
                    time.sleep(1.5)
                    h1_filename = f'{screenshots_dir}/h1_inspector_{timestamp}.png'
                    driver.save_screenshot(h1_filename)
                    h1_screenshot = h1_filename
                    self.log("  ✓ H1")
                    driver.execute_script("arguments[0].style.outline = ''; arguments[0].style.backgroundColor = ''; var p = document.getElementById('custom-inspector-h1'); if (p) p.remove();", h1_element)
                except:
                    self.log("  ✗ H1")
                
                try:
                    h2_element = driver.find_element(By.TAG_NAME, 'h2')
                    h2_html = h2_element.get_attribute('outerHTML')
                    h2_text = h2_element.text
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", h2_element)
                    driver.execute_script("""
                        arguments[0].style.outline = '3px solid blue';
                        arguments[0].style.backgroundColor = 'rgba(0, 255, 255, 0.3)';
                        var p = document.createElement('div');
                        p.id = 'custom-inspector-h2';
                        p.style.cssText = 'position: fixed; right: 20px; top: 80px; width: 400px; max-height: 600px; background: #282828; color: #fff; padding: 20px; border-radius: 8px; box-shadow: 0 8px 32px rgba(0,0,0,0.5); font-family: Courier New, monospace; font-size: 13px; z-index: 999999; overflow-y: auto; display: block !important; visibility: visible !important;';
                        var t = document.createElement('div');
                        t.style.cssText = 'color: #61dafb; font-weight: bold; margin-bottom: 15px; font-size: 14px; border-bottom: 2px solid #61dafb; padding-bottom: 8px;';
                        t.textContent = 'Inspector - H2';
                        p.appendChild(t);
                        var c = document.createElement('pre');
                        c.style.cssText = 'margin: 0; white-space: pre-wrap; word-wrap: break-word; color: #e06c75;';
                        c.textContent = arguments[1];
                        p.appendChild(c);
                        var i = document.createElement('div');
                        i.style.cssText = 'margin-top: 15px; padding-top: 15px; border-top: 1px solid #444; color: #98c379;';
                        i.innerHTML = '<strong>Texto:</strong> ' + arguments[2];
                        p.appendChild(i);
                        document.body.appendChild(p);
                    """, h2_element, h2_html, h2_text)
                    time.sleep(1.5)
                    h2_filename = f'{screenshots_dir}/h2_inspector_{timestamp}.png'
                    driver.save_screenshot(h2_filename)
                    h2_screenshot = h2_filename
                    self.log("  ✓ H2")
                    driver.execute_script("arguments[0].style.outline = ''; arguments[0].style.backgroundColor = ''; var p = document.getElementById('custom-inspector-h2'); if (p) p.remove();", h2_element)
                except:
                    self.log("  ✗ H2")
            except Exception as e:
                self.log(f"  ✗ {str(e)[:30]}")
            
            fullpage_screenshot = None
            self.set_status("Pantalla completa...")
            self.log(f"\n🖥️ Pantalla completa...")
            try:
                driver.get(url)
                time.sleep(3)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
                fullpage_filename = f'{screenshots_dir}/fullpage_{timestamp}.png'
                fullpage_screenshot_img = pyautogui.screenshot()
                fullpage_screenshot_img.save(fullpage_filename)
                fullpage_screenshot = fullpage_filename
                self.log("  ✓")
            except Exception as e:
                self.log(f"  ✗ {str(e)[:30]}")
        
        
            # Las variables de capturas ya están definidas arriba con sus valores reales
            # No reinicializar para no perder las capturas generadas
            
            # ===== CAPTURAS DE DOMINIO (IONOS/DONDOMINIO) =====
            
            # Capturar panel de control del dominio
            self.set_status("Panel de control...")
            self.log(f"\n🌐 Panel de control del dominio...")
            panel_control_screenshots = self.capturar_panel_control_dominio_gui(url, driver, timestamp)
            
            driver.quit()
            driver = None
            
            # ===== GENERACIÓN DE PDF COMENTADA PARA PRUEBA RÁPIDA =====
            # Descomenta esta sección cuando quieras generar PDFs completos
            
            self.set_status("Generando PDF...")
            self.log(f"\n📄 Generando PDF...")
            
            # Debug: verificar qué capturas tenemos
            self.log(f"  📊 Capturas disponibles:")
            self.log(f"    - Keywords: {len(screenshots_guardadas) if screenshots_guardadas else 0}")
            self.log(f"    - WHOIS: {'Sí' if whois_screenshot else 'No'}")
            self.log(f"    - H1: {'Sí' if h1_screenshot else 'No'}")
            self.log(f"    - H2: {'Sí' if h2_screenshot else 'No'}")
            self.log(f"    - Fullpage: {'Sí' if fullpage_screenshot else 'No'}")
            self.log(f"    - Competencia: {len(competencia_screenshots) if competencia_screenshots else 0}")
            if competencia_screenshots:
                for i, comp in enumerate(competencia_screenshots):
                    self.log(f"      Competencia {i+1}: {comp}")
            self.log(f"    - Análisis IA: {'Sí' if analisis_competencia_ia and analisis_competencia_ia.get('analisis') else 'No'}")
            if analisis_competencia_ia and analisis_competencia_ia.get('analisis'):
                self.log(f"      Texto análisis: {analisis_competencia_ia.get('analisis', '')[:100]}...")
            self.log(f"    - Métricas: {'Sí' if metrics_screenshot else 'No'}")
            self.log(f"    - Vista General: {'Sí' if vista_general_screenshot else 'No'}")
            self.log(f"    - Indexación: {len(indexacion_screenshots) if indexacion_screenshots else 0}")
            
            pdf_generado = generar_pdf_con_capturas(
                screenshots_guardadas, keywords, url,
                competidores=competidores,
                whois_screenshot=whois_screenshot,
                h1_screenshot=h1_screenshot,
                h2_screenshot=h2_screenshot,
                fullpage_screenshot=fullpage_screenshot,
                competencia_screenshots=competencia_screenshots,
                metrics_screenshot=metrics_screenshot,
                vista_general_screenshot=vista_general_screenshot,
                indexacion_screenshots=indexacion_screenshots,
                analisis_competencia_ia=analisis_competencia_ia
            )
            if pdf_generado:
                self.ultimo_pdf = os.path.abspath(pdf_generado)
                self.open_pdf_button.config(state=tk.NORMAL)
                self.log(f"✓ PDF: {os.path.basename(pdf_generado)}")
                self.log(f"📂 Guardado en: {os.path.dirname(self.ultimo_pdf)}")
                
                # Generar PDF del panel de control si hay screenshots
                if panel_control_screenshots:
                    self.set_status("Generando PDF panel de control...")
                    self.log(f"\n📄 Generando PDF panel de control...")
                    nombre_web = self.obtener_nombre_web_con_ia_gui(url)
                    pdf_panel_control = self.generar_pdf_titularidad_dominio(panel_control_screenshots, url, nombre_web, timestamp)
                    if pdf_panel_control:
                        self.log(f"✓ PDF Panel Control: {os.path.basename(pdf_panel_control)}")
                    else:
                        self.log("✗ Error PDF Panel Control")
                
                # Generar PDF de publicidad si existe fullpage_screenshot
                if fullpage_screenshot and os.path.exists(fullpage_screenshot):
                    self.set_status("Generando PDF de publicidad...")
                    self.log(f"\n📄 Generando PDF de publicidad...")
                    pdf_publicidad = self.generar_pdf_publicidad(fullpage_screenshot, url, timestamp)
                    if pdf_publicidad:
                        self.log(f"✓ PDF Publicidad: {os.path.basename(pdf_publicidad)}")
                    else:
                        self.log("✗ Error PDF Publicidad")
                
                self.set_status("¡Completado!")
                messagebox.showinfo("Éxito", f"PDFs generados:\n{os.path.basename(pdf_generado)}")
            else:
                self.log("✗ Error PDF")
                self.set_status("Error en PDF")
            
            # ===== SOLO GENERAR PDF DEL PANEL DE CONTROL =====
            if panel_control_screenshots:
                self.set_status("Generando PDF panel de control...")
                self.log(f"\n📄 Generando PDF panel de control...")
                nombre_web = self.obtener_nombre_web_con_ia_gui(url)
                pdf_panel_control = self.generar_pdf_titularidad_dominio(panel_control_screenshots, url, nombre_web, timestamp)
                if pdf_panel_control:
                    self.log(f"✓ PDF Panel Control: {os.path.basename(pdf_panel_control)}")
                    
                    # Generar PDF de publicidad si existe fullpage_screenshot
                    if fullpage_screenshot and os.path.exists(fullpage_screenshot):
                        self.set_status("Generando PDF de publicidad...")
                        self.log(f"\n📄 Generando PDF de publicidad...")
                        pdf_publicidad = self.generar_pdf_publicidad(fullpage_screenshot, url, timestamp)
                        if pdf_publicidad:
                            self.log(f"✓ PDF Publicidad: {os.path.basename(pdf_publicidad)}")
                        else:
                            self.log("✗ Error PDF Publicidad")
                    
                    self.set_status("¡Completado!")
                    messagebox.showinfo("Éxito", f"PDF Panel de Control generado:\n{os.path.basename(pdf_panel_control)}")
                else:
                    self.log("✗ Error PDF Panel Control")
                    self.set_status("Error en PDF")
            else:
                self.log("✗ No se capturaron screenshots del panel de control")
                self.set_status("Sin capturas de panel de control")
            
        except Exception as e:
            self.log(f"\n✗ ERROR: {str(e)}")
            messagebox.showerror("Error", str(e))
            
        finally:
            if driver:
                try:
                    driver.quit()
                except:
                    pass
                    
            self.log("=" * 60)
            self.log("✅ ANÁLISIS COMPLETADO")
            self.log("=" * 60)
            
            self.start_button.config(state=tk.NORMAL)
            self.progress.stop()
            
            try:
                if not self.status_label.cget('text').startswith('¡Completado'):
                    self.set_status("Listo para nuevo análisis")
            except:
                self.set_status("Listo")

    def capturar_panel_control_dominio_gui(self, url_web, driver, timestamp):
        """Captura screenshots del panel de control del dominio para la GUI"""
        screenshots = []
        
        try:
            # Extraer dominio sin protocolo y www
            domain = urlparse(url_web).netloc.replace('www.', '')
            domain_name_without_tld = domain.split('.')[0]  # Solo el nombre sin TLD para Dondominio
            domain_name_with_tld = domain  # Dominio completo con TLD para Ionos
            
            self.log(f"  🔍 Dominio completo: {domain_name_with_tld}")
            self.log(f"  🔍 Dominio sin TLD: {domain_name_without_tld}")
            
            # Intentar primero con Ionos
            self.log("  🔍 Intentando con Ionos...")
            ionos_url = f"https://my.ionos.es/domain-privacy/{domain_name_with_tld}?linkId=ct.tab.domainlist.privacy"
            
            try:
                # 1. Acceder a la URL especificada
                self.log("  📍 Accediendo a URL especificada...")
                driver.get(ionos_url)
                time.sleep(3)
                
                # 2. Verificar si nos redirige al login
                current_url = driver.current_url
                if "login" in current_url or "signin" in current_url:
                    self.log("  🔐 Redirigido al login - iniciando sesión...")
                    wait = WebDriverWait(driver, 15)
                    
                    try:
                        # Paso 1: Rellenar username y click en siguiente
                        self.log("  👤 Rellenando username...")
                        username_input = None
                        for selector in [By.NAME, By.ID, By.CSS_SELECTOR]:
                            try:
                                if selector == By.NAME:
                                    username_input = wait.until(EC.element_to_be_clickable((By.NAME, "username")))
                                elif selector == By.ID:
                                    username_input = wait.until(EC.element_to_be_clickable((By.ID, "username")))
                                elif selector == By.CSS_SELECTOR:
                                    username_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[name='username']")))
                                break
                            except:
                                continue
                        
                        if username_input:
                            username_input.clear()
                            username_input.send_keys("i-pointsite.com")
                            time.sleep(0.5)
                        
                            # Click en siguiente
                            self.log("  ➡️ Click en siguiente...")
                            next_button = None
                            for selector in [By.CSS_SELECTOR, By.XPATH, By.CLASS_NAME]:
                                try:
                                    if selector == By.CSS_SELECTOR:
                                        next_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit']")))
                                    elif selector == By.XPATH:
                                        next_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Siguiente') or contains(text(), 'Next') or contains(text(), 'Continuar')]")))
                                    elif selector == By.CLASS_NAME:
                                        next_button = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "button--with-loader")))
                                    break
                                except:
                                    continue
                            
                            if next_button:
                                next_button.click()
                                time.sleep(2)
                                self.log("  ✓ Username enviado")
                        
                        # Paso 2: Rellenar password y click en siguiente
                        self.log("  🔑 Rellenando password...")
                        password_input = None
                        for selector in [By.NAME, By.ID, By.CSS_SELECTOR]:
                            try:
                                if selector == By.NAME:
                                    password_input = wait.until(EC.element_to_be_clickable((By.NAME, "password")))
                                elif selector == By.ID:
                                    password_input = wait.until(EC.element_to_be_clickable((By.ID, "password")))
                                elif selector == By.CSS_SELECTOR:
                                    password_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[name='password']")))
                                break
                            except:
                                continue
                        
                        if password_input:
                            password_input.clear()
                            password_input.send_keys("R4t420223!")
                            time.sleep(0.5)
                            
                            # Click en siguiente/login final
                            self.log("  ➡️ Click en siguiente...")
                            login_button = None
                            for selector in [By.CSS_SELECTOR, By.XPATH, By.CLASS_NAME]:
                                try:
                                    if selector == By.CSS_SELECTOR:
                                        login_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit']")))
                                    elif selector == By.XPATH:
                                        login_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Login') or contains(text(), 'Entrar') or contains(text(), 'Siguiente') or contains(text(), 'Next')]")))
                                    elif selector == By.CLASS_NAME:
                                        login_button = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "button--with-loader")))
                                    break
                                except:
                                    continue
                            
                            if login_button:
                                login_button.click()
                                time.sleep(3)
                                self.log("  ✓ Login completado")
                            else:
                                self.log("  ❌ No se encontró botón de login")
                        else:
                            self.log("  ❌ No se encontró campo de password")
                            
                    except Exception as login_error:
                        self.log(f"  ❌ Error en login: {str(login_error)[:50]}")
                else:
                    self.log("  ✓ Ya autenticado en Ionos")
                
                # 3. Verificar si nos redirige a la URL especificada (Ionos)
                self.log("  🔍 Verificando redirección...")
                current_url = driver.current_url
                if "my.ionos.es/domains" in current_url or "my.ionos.es" not in current_url:
                    self.log("  ❌ No se redirigió a la URL especificada - es Dondominio")
                    raise Exception("Dominio no en Ionos")
                else:
                    self.log("  ✓ Redirigido correctamente a Ionos")
                
                # Captura 1: Scroll a la mitad de la pantalla primero
                try:
                    self.log("  📍 Haciendo scroll a la mitad de la pantalla...")
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight/2);")
                    time.sleep(1)
                    self.log("  ✓ Scroll a la mitad completado")
                except Exception as e:
                    self.log(f"  ✗ Error en scroll: {str(e)[:30]}")
                
                # Captura 2: Elemento específico - probar múltiples selectores
                try:
                    elemento = None
                    wait = WebDriverWait(driver, 10)
                    
                    # Probar múltiples selectores para el elemento
                    selectors = [
                        (By.XPATH, "/html/body/main/div[2]/section[3]"),
                        (By.XPATH, "//section[3]"),
                        (By.XPATH, "//main//section[last()]"),
                        (By.CSS_SELECTOR, "main div:nth-child(2) section:nth-child(3)"),
                        (By.CSS_SELECTOR, "section:nth-child(3)"),
                        (By.CLASS_NAME, "domain-info"),
                        (By.CLASS_NAME, "privacy-info")
                    ]
                    
                    for selector_type, selector_value in selectors:
                        try:
                            elemento = wait.until(EC.presence_of_element_located((selector_type, selector_value)))
                            self.log(f"  ✓ Elemento encontrado con selector: {selector_type}")
                            break
                        except:
                            continue
                    
                    if elemento:
                        # Hacer scroll al elemento y centrarlo mejor
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elemento)
                        driver.execute_script("window.scrollBy(0, -50);")
                        time.sleep(0.5)
                        
                        # Hacer la captura más grande usando zoom
                        driver.execute_script("document.body.style.zoom='1.2'")
                        time.sleep(0.3)
                        
                        screenshot_filename = f'{self.screenshots_dir}/ionos_elemento_{timestamp}.png'
                        elemento.screenshot(screenshot_filename)
                        screenshots.append(screenshot_filename)
                        self.log("  ✓ Ionos elemento (ampliado)")
                        
                        # Restaurar zoom
                        driver.execute_script("document.body.style.zoom='1.0'")
                    else:
                        self.log("  ⚠️ No se encontró elemento específico, usando captura completa")
                        
                except Exception as e:
                    self.log(f"  ✗ Ionos elemento: {str(e)[:30]}")
                
                # Captura 3: Pantalla completa sin scroll adicional con zoom 85%
                try:
                    # Aplicar zoom al 85% (sin hacer scroll adicional)
                    driver.execute_script("document.body.style.zoom='0.85'")
                    time.sleep(0.5)
                    
                    screenshot_filename = f'{self.screenshots_dir}/ionos_completa_{timestamp}.png'
                    pyautogui.screenshot(screenshot_filename)
                    screenshots.append(screenshot_filename)
                    self.log("  ✓ Ionos completa (zoom 85%, sin scroll)")
                    
                    # Restaurar zoom
                    driver.execute_script("document.body.style.zoom='1.0'")
                except Exception as e:
                    self.log(f"  ✗ Ionos completa: {str(e)[:30]}")
                
                return screenshots
                
            except Exception as e:
                self.log(f"  ❌ Error con Ionos: {str(e)[:30]}")
                self.log("  🔄 Probando con Dondominio...")
                
                # Intentar con Dondominio
                try:
                    # Primero ir a la página de login de Dondominio
                    self.log("  🔐 Accediendo a login de Dondominio...")
                    driver.get("https://www.dondominio.com/admin/")
                    time.sleep(3)
                    
                    wait = WebDriverWait(driver, 15)
                    
                    # Verificar si ya estamos en la página de login
                    current_url = driver.current_url
                    self.log(f"  📍 URL actual: {current_url}")
                    
                    # Login en Dondominio
                    self.log("  🔐 Iniciando sesión en Dondominio...")
                    
                    try:
                        # Usuario - probar múltiples selectores más específicos
                        user_input = None
                        selectors_user = [
                            (By.NAME, "user"),
                            (By.NAME, "username"),
                            (By.NAME, "email"),
                            (By.ID, "user"),
                            (By.ID, "username"),
                            (By.ID, "email"),
                            (By.CSS_SELECTOR, "input[name='user']"),
                            (By.CSS_SELECTOR, "input[name='username']"),
                            (By.CSS_SELECTOR, "input[name='email']"),
                            (By.CSS_SELECTOR, "input[type='text']"),
                            (By.CSS_SELECTOR, "input[type='email']")
                        ]
                        
                        for selector_type, selector_value in selectors_user:
                            try:
                                user_input = wait.until(EC.element_to_be_clickable((selector_type, selector_value)))
                                self.log(f"  ✓ Campo usuario encontrado con: {selector_type} = {selector_value}")
                                break
                            except:
                                continue
                        
                        if user_input:
                            # Esperar a que el elemento esté completamente cargado
                            time.sleep(1)
                            
                            # Limpiar campo de forma más robusta
                            try:
                                # Hacer scroll al elemento
                                driver.execute_script("arguments[0].scrollIntoView(true);", user_input)
                                time.sleep(0.5)
                                
                                # Intentar escribir en Shadow DOM
                                usuario_escrito = False
                                
                                # Método 1: Acceder al Shadow DOM directamente
                                try:
                                    driver.execute_script("""
                                        var text = 'THWORK';
                                        
                                        // Buscar el elemento dd-input con name="user"
                                        var ddInput = document.querySelector('dd-input[name="user"]');
                                        if (ddInput && ddInput.shadowRoot) {
                                            // Acceder al input real dentro del shadow DOM
                                            var realInput = ddInput.shadowRoot.querySelector('input');
                                            if (realInput) {
                                                // Escribir en el input real
                                                realInput.focus();
                                                realInput.value = '';
                                                realInput.value = text;
                                                
                                                // Disparar eventos en el input real
                                                realInput.dispatchEvent(new Event('input', { bubbles: true }));
                                                realInput.dispatchEvent(new Event('change', { bubbles: true }));
                                                realInput.dispatchEvent(new Event('blur', { bubbles: true }));
                                                
                                                return 'SUCCESS';
                                            } else {
                                                return 'NO_INPUT_IN_SHADOW';
                                            }
                                        } else {
                                            return 'NO_SHADOW_ROOT';
                                        }
                                    """)
                                    time.sleep(1)
                                    
                                    # Verificar si se escribió correctamente
                                    valor_verificacion = driver.execute_script("""
                                        var ddInput = document.querySelector('dd-input[name="user"]');
                                        if (ddInput && ddInput.shadowRoot) {
                                            var realInput = ddInput.shadowRoot.querySelector('input');
                                            return realInput ? realInput.value : 'NO_INPUT';
                                        }
                                        return 'NO_SHADOW';
                                    """)
                                    
                                    if valor_verificacion == 'THWORK':
                                        self.log("  ✓ Usuario ingresado (Shadow DOM)")
                                        usuario_escrito = True
                                    else:
                                        self.log(f"  ⚠️ Shadow DOM falló: valor='{valor_verificacion}'")
                                except Exception as e1:
                                    self.log(f"  ⚠️ Error Shadow DOM: {str(e1)[:30]}")
                                
                                # Método 2: Intentar con diferentes selectores de Shadow DOM
                                if not usuario_escrito:
                                    try:
                                        driver.execute_script("""
                                            var text = 'THWORK';
                                            
                                            // Probar diferentes selectores
                                            var selectors = [
                                                'dd-input[name="user"]',
                                                'dd-input[data-name="user"]',
                                                'dd-input',
                                                'input[name="user"]'
                                            ];
                                            
                                            for (var i = 0; i < selectors.length; i++) {
                                                var element = document.querySelector(selectors[i]);
                                                if (element) {
                                                    // Si tiene shadowRoot, acceder al input interno
                                                    if (element.shadowRoot) {
                                                        var realInput = element.shadowRoot.querySelector('input');
                                                        if (realInput) {
                                                            realInput.focus();
                                                            realInput.value = '';
                                                            realInput.value = text;
                                                            realInput.dispatchEvent(new Event('input', { bubbles: true }));
                                                            realInput.dispatchEvent(new Event('change', { bubbles: true }));
                                                            return 'SUCCESS_' + selectors[i];
                                                        }
                                                    } else {
                                                        // Si no tiene shadowRoot, es un input normal
                                                        element.focus();
                                                        element.value = '';
                                                        element.value = text;
                                                        element.dispatchEvent(new Event('input', { bubbles: true }));
                                                        element.dispatchEvent(new Event('change', { bubbles: true }));
                                                        return 'SUCCESS_NORMAL_' + selectors[i];
                                                    }
                                                }
                                            }
                                            return 'NO_ELEMENT_FOUND';
                                        """)
                                        time.sleep(1)
                                        self.log("  ✓ Usuario ingresado (Múltiples selectores)")
                                        usuario_escrito = True
                                    except Exception as e2:
                                        self.log(f"  ⚠️ Error múltiples selectores: {str(e2)[:30]}")
                                
                                # Método 3: Selenium estándar como último recurso
                                if not usuario_escrito:
                                    try:
                                        user_input.click()
                                        time.sleep(0.5)
                                        user_input.clear()
                                        time.sleep(0.5)
                                        user_input.send_keys("THWORK")
                                        time.sleep(0.5)
                                        self.log("  ✓ Usuario ingresado (Selenium)")
                                        usuario_escrito = True
                                    except Exception as e3:
                                        self.log(f"  ⚠️ Error Selenium: {str(e3)[:30]}")
                            except Exception as e_general:
                                self.log(f"  ❌ Error general en usuario: {str(e_general)[:30]}")
                        
                        # Verificar que el usuario se haya escrito correctamente
                        if user_input:
                            try:
                                valor_actual = user_input.get_attribute('value')
                                if valor_actual == 'THWORK':
                                    self.log("  ✅ Usuario verificado correctamente")
                                else:
                                    self.log(f"  ⚠️ Usuario no coincide: '{valor_actual}' (esperado: 'THWORK')")
                            except:
                                self.log("  ⚠️ No se pudo verificar el valor del usuario")
                        else:
                            self.log("  ❌ No se encontró campo de usuario")
                        
                        # Forzar validación del formulario después de escribir usuario
                        try:
                            driver.execute_script("""
                                // Buscar el formulario y forzar validación
                                var forms = document.querySelectorAll('form');
                                for (var i = 0; i < forms.length; i++) {
                                    forms[i].dispatchEvent(new Event('input', { bubbles: true }));
                                    forms[i].dispatchEvent(new Event('change', { bubbles: true }));
                                }
                                
                                // También disparar eventos en los elementos dd-input
                                var ddInputs = document.querySelectorAll('dd-input, dd-password');
                                for (var i = 0; i < ddInputs.length; i++) {
                                    if (ddInputs[i].shadowRoot) {
                                        var inputs = ddInputs[i].shadowRoot.querySelectorAll('input');
                                        for (var j = 0; j < inputs.length; j++) {
                                            inputs[j].dispatchEvent(new Event('input', { bubbles: true }));
                                            inputs[j].dispatchEvent(new Event('change', { bubbles: true }));
                                        }
                                    }
                                }
                            """)
                            time.sleep(0.5)
                            self.log("  ✓ Validación de formulario forzada")
                        except Exception as e:
                            self.log(f"  ⚠️ Error forzando validación: {str(e)[:30]}")
                        
                        # Contraseña - probar múltiples selectores más específicos
                        password_input = None
                        selectors_password = [
                            (By.NAME, "password"),
                            (By.NAME, "pass"),
                            (By.ID, "password"),
                            (By.ID, "pass"),
                            (By.CSS_SELECTOR, "input[name='password']"),
                            (By.CSS_SELECTOR, "input[name='pass']"),
                            (By.CSS_SELECTOR, "input[type='password']")
                        ]
                        
                        for selector_type, selector_value in selectors_password:
                            try:
                                password_input = wait.until(EC.element_to_be_clickable((selector_type, selector_value)))
                                self.log(f"  ✓ Campo contraseña encontrado con: {selector_type} = {selector_value}")
                                break
                            except:
                                continue
                        
                        if password_input:
                            # Esperar a que el elemento esté completamente cargado
                            time.sleep(1)
                            
                            # Limpiar campo de forma más robusta
                            try:
                                # Hacer scroll al elemento
                                driver.execute_script("arguments[0].scrollIntoView(true);", password_input)
                                time.sleep(0.5)
                                
                                # Intentar escribir en Shadow DOM
                                contraseña_escrita = False
                                
                                # Método 1: Acceder al Shadow DOM directamente
                                try:
                                    driver.execute_script("""
                                        var text = 'R4t4-2025';
                                        
                                        // Buscar el elemento dd-password con name="password"
                                        var ddPassword = document.querySelector('dd-password[name="password"]');
                                        if (ddPassword && ddPassword.shadowRoot) {
                                            // Acceder al input real dentro del shadow DOM
                                            var realInput = ddPassword.shadowRoot.querySelector('input');
                                            if (realInput) {
                                                // Escribir en el input real
                                                realInput.focus();
                                                realInput.value = '';
                                                realInput.value = text;
                                                
                                                // Disparar eventos en el input real
                                                realInput.dispatchEvent(new Event('input', { bubbles: true }));
                                                realInput.dispatchEvent(new Event('change', { bubbles: true }));
                                                realInput.dispatchEvent(new Event('blur', { bubbles: true }));
                                                
                                                return 'SUCCESS';
                                            } else {
                                                return 'NO_INPUT_IN_SHADOW';
                                            }
                                        } else {
                                            return 'NO_SHADOW_ROOT';
                                        }
                                    """)
                                    time.sleep(1)
                                    
                                    # Verificar si se escribió correctamente
                                    valor_verificacion = driver.execute_script("""
                                        var ddPassword = document.querySelector('dd-password[name="password"]');
                                        if (ddPassword && ddPassword.shadowRoot) {
                                            var realInput = ddPassword.shadowRoot.querySelector('input');
                                            return realInput ? realInput.value : 'NO_INPUT';
                                        }
                                        return 'NO_SHADOW';
                                    """)
                                    
                                    if valor_verificacion == 'R4t4-2025':
                                        self.log("  ✓ Contraseña ingresada (Shadow DOM)")
                                        contraseña_escrita = True
                                    else:
                                        self.log(f"  ⚠️ Shadow DOM falló: valor='{valor_verificacion}'")
                                except Exception as e1:
                                    self.log(f"  ⚠️ Error Shadow DOM: {str(e1)[:30]}")
                                
                                # Método 2: Intentar con diferentes selectores de Shadow DOM
                                if not contraseña_escrita:
                                    try:
                                        driver.execute_script("""
                                            var text = 'R4t4-2025';
                                            
                                            // Probar diferentes selectores
                                            var selectors = [
                                                'dd-password[name="password"]',
                                                'dd-password[data-name="password"]',
                                                'dd-password',
                                                'input[name="password"]',
                                                'input[type="password"]'
                                            ];
                                            
                                            for (var i = 0; i < selectors.length; i++) {
                                                var element = document.querySelector(selectors[i]);
                                                if (element) {
                                                    // Si tiene shadowRoot, acceder al input interno
                                                    if (element.shadowRoot) {
                                                        var realInput = element.shadowRoot.querySelector('input');
                                                        if (realInput) {
                                                            realInput.focus();
                                                            realInput.value = '';
                                                            realInput.value = text;
                                                            realInput.dispatchEvent(new Event('input', { bubbles: true }));
                                                            realInput.dispatchEvent(new Event('change', { bubbles: true }));
                                                            return 'SUCCESS_' + selectors[i];
                                                        }
                                                    } else {
                                                        // Si no tiene shadowRoot, es un input normal
                                                        element.focus();
                                                        element.value = '';
                                                        element.value = text;
                                                        element.dispatchEvent(new Event('input', { bubbles: true }));
                                                        element.dispatchEvent(new Event('change', { bubbles: true }));
                                                        return 'SUCCESS_NORMAL_' + selectors[i];
                                                    }
                                                }
                                            }
                                            return 'NO_ELEMENT_FOUND';
                                        """)
                                        time.sleep(1)
                                        self.log("  ✓ Contraseña ingresada (Múltiples selectores)")
                                        contraseña_escrita = True
                                    except Exception as e2:
                                        self.log(f"  ⚠️ Error múltiples selectores: {str(e2)[:30]}")
                                
                                # Método 3: Selenium estándar como último recurso
                                if not contraseña_escrita:
                                    try:
                                        password_input.click()
                                        time.sleep(0.5)
                                        password_input.clear()
                                        time.sleep(0.5)
                                        password_input.send_keys("R4t4-2025")
                                        time.sleep(0.5)
                                        self.log("  ✓ Contraseña ingresada (Selenium)")
                                        contraseña_escrita = True
                                    except Exception as e3:
                                        self.log(f"  ⚠️ Error Selenium: {str(e3)[:30]}")
                            except Exception as e_general:
                                self.log(f"  ❌ Error general en contraseña: {str(e_general)[:30]}")
                        
                        # Verificar que la contraseña se haya escrito correctamente
                        if password_input:
                            try:
                                valor_actual = password_input.get_attribute('value')
                                if valor_actual == 'R4t4-2025':
                                    self.log("  ✅ Contraseña verificada correctamente")
                                else:
                                    self.log(f"  ⚠️ Contraseña no coincide: '{valor_actual}' (esperado: 'R4t4-2025')")
                            except:
                                self.log("  ⚠️ No se pudo verificar el valor de la contraseña")
                        else:
                            self.log("  ❌ No se encontró campo de contraseña")
                        
                        # Forzar validación del formulario después de escribir contraseña
                        try:
                            driver.execute_script("""
                                // Buscar el formulario y forzar validación
                                var forms = document.querySelectorAll('form');
                                for (var i = 0; i < forms.length; i++) {
                                    forms[i].dispatchEvent(new Event('input', { bubbles: true }));
                                    forms[i].dispatchEvent(new Event('change', { bubbles: true }));
                                }
                                
                                // También disparar eventos en los elementos dd-input
                                var ddInputs = document.querySelectorAll('dd-input, dd-password');
                                for (var i = 0; i < ddInputs.length; i++) {
                                    if (ddInputs[i].shadowRoot) {
                                        var inputs = ddInputs[i].shadowRoot.querySelectorAll('input');
                                        for (var j = 0; j < inputs.length; j++) {
                                            inputs[j].dispatchEvent(new Event('input', { bubbles: true }));
                                            inputs[j].dispatchEvent(new Event('change', { bubbles: true }));
                                        }
                                    }
                                }
                                
                                // Forzar validación HTML5
                                var inputs = document.querySelectorAll('input');
                                for (var i = 0; i < inputs.length; i++) {
                                    if (inputs[i].checkValidity) {
                                        inputs[i].checkValidity();
                                    }
                                }
                            """)
                            time.sleep(0.5)
                            self.log("  ✓ Validación de formulario forzada (contraseña)")
                        except Exception as e:
                            self.log(f"  ⚠️ Error forzando validación: {str(e)[:30]}")
                        
                        # Validación final antes del login
                        try:
                            driver.execute_script("""
                                // Verificar que los campos estén llenos y válidos
                                var ddUser = document.querySelector('dd-input[name="user"]');
                                var ddPassword = document.querySelector('dd-password[name="password"]');
                                
                                var userValid = false;
                                var passwordValid = false;
                                
                                if (ddUser && ddUser.shadowRoot) {
                                    var userInput = ddUser.shadowRoot.querySelector('input');
                                    if (userInput && userInput.value === 'THWORK') {
                                        userValid = true;
                                        // Forzar eventos finales
                                        userInput.dispatchEvent(new Event('blur', { bubbles: true }));
                                    }
                                }
                                
                                if (ddPassword && ddPassword.shadowRoot) {
                                    var passwordInput = ddPassword.shadowRoot.querySelector('input');
                                    if (passwordInput && passwordInput.value === 'R4t4-2025') {
                                        passwordValid = true;
                                        // Forzar eventos finales
                                        passwordInput.dispatchEvent(new Event('blur', { bubbles: true }));
                                    }
                                }
                                
                                // Forzar validación del formulario completo
                                var forms = document.querySelectorAll('form');
                                for (var i = 0; i < forms.length; i++) {
                                    forms[i].dispatchEvent(new Event('submit', { bubbles: true, cancelable: true }));
                                }
                                
                                return 'User: ' + userValid + ', Password: ' + passwordValid;
                            """)
                            time.sleep(1)
                            self.log("  ✓ Validación final completada")
                        except Exception as e:
                            self.log(f"  ⚠️ Error validación final: {str(e)[:30]}")
                        
                        # Click login - probar múltiples selectores más específicos
                        login_button = None
                        selectors_button = [
                            (By.XPATH, "//button[@type='submit']"),
                            (By.XPATH, "//input[@type='submit']"),
                            (By.XPATH, "//button[contains(text(), 'Login')]"),
                            (By.XPATH, "//button[contains(text(), 'Entrar')]"),
                            (By.XPATH, "//button[contains(text(), 'Iniciar')]"),
                            (By.CSS_SELECTOR, "button[type='submit']"),
                            (By.CSS_SELECTOR, "input[type='submit']"),
                            (By.CSS_SELECTOR, "button.btn-primary"),
                            (By.CSS_SELECTOR, "button.btn"),
                            (By.CLASS_NAME, "btn-primary"),
                            (By.CLASS_NAME, "btn")
                        ]
                        
                        for selector_type, selector_value in selectors_button:
                            try:
                                login_button = wait.until(EC.element_to_be_clickable((selector_type, selector_value)))
                                self.log(f"  ✓ Botón login encontrado con: {selector_type} = {selector_value}")
                                break
                            except:
                                continue
                        
                        if login_button:
                            # Asegurar que el botón esté visible y clickeable
                            try:
                                driver.execute_script("arguments[0].scrollIntoView(true);", login_button)
                                time.sleep(0.5)
                                
                                # Intentar click normal
                                login_button.click()
                                time.sleep(5)  # Más tiempo para el login
                                self.log("  ✓ Click en login realizado")
                                
                                # Verificar si el login fue exitoso
                                new_url = driver.current_url
                                if "login" not in new_url.lower() and "admin" in new_url.lower():
                                    self.log("  ✓ Login exitoso en Dondominio")
                                else:
                                    self.log("  ⚠️ Login puede no haber sido exitoso")
                                
                                # Verificar si aparece verificación 2FA
                                try:
                                    time.sleep(2)  # Esperar a que cargue la página
                                    twofa_element = driver.find_element(By.CLASS_NAME, "input-2fa-code")
                                    if twofa_element:
                                        self.log("  🔐 Verificación 2FA detectada")
                                        
                                        # Obtener código de verificación del email
                                        codigo = self.obtener_codigo_verificacion_email()
                                        if codigo and len(codigo) == 6:
                                            self.log(f"  📧 Código obtenido: {codigo}")
                                            
                                            # Introducir código dígito por dígito
                                            for i, digito in enumerate(codigo, 1):
                                                try:
                                                    input_field = driver.find_element(By.NAME, f"code-sub-{i}")
                                                    input_field.clear()
                                                    input_field.send_keys(digito)
                                                    time.sleep(0.2)
                                                    self.log(f"  ✓ Dígito {i}: {digito}")
                                                except Exception as e:
                                                    self.log(f"  ❌ Error introduciendo dígito {i}: {str(e)[:30]}")
                                            
                                            # Hacer click en el botón de verificación
                                            try:
                                                verify_button = driver.find_element(By.CSS_SELECTOR, "button.btn.btn-dark.btn-lg.btn-block[type='submit']")
                                                verify_button.click()
                                                time.sleep(3)
                                                self.log("  ✓ Código de verificación enviado")
                                                
                                                # Verificar si la verificación fue exitosa
                                                final_url = driver.current_url
                                                if "admin" in final_url.lower():
                                                    self.log("  ✅ Verificación 2FA exitosa")
                                                else:
                                                    self.log("  ⚠️ Verificación 2FA puede no haber sido exitosa")
                                                    
                                            except Exception as e:
                                                self.log(f"  ❌ Error haciendo click en verificar: {str(e)[:30]}")
                                        else:
                                            self.log("  ❌ No se pudo obtener código de verificación")
                                    else:
                                        self.log("  ✓ No se requiere verificación 2FA")
                                except:
                                    self.log("  ✓ No se requiere verificación 2FA")
                            except Exception as e:
                                self.log(f"  ⚠️ Error click normal: {str(e)[:30]}")
                                # Intentar click con JavaScript
                                try:
                                    driver.execute_script("arguments[0].click();", login_button)
                                    time.sleep(5)
                                    self.log("  ✓ Click en login realizado (JavaScript)")
                                except Exception as e2:
                                    self.log(f"  ❌ Error click JavaScript: {str(e2)[:30]}")
                                
                                # Verificar si aparece verificación 2FA (para el caso de JavaScript)
                                try:
                                    time.sleep(2)  # Esperar a que cargue la página
                                    twofa_element = driver.find_element(By.CLASS_NAME, "input-2fa-code")
                                    if twofa_element:
                                        self.log("  🔐 Verificación 2FA detectada (JavaScript)")
                                        
                                        # Obtener código de verificación del email
                                        codigo = self.obtener_codigo_verificacion_email()
                                        if codigo and len(codigo) == 6:
                                            self.log(f"  📧 Código obtenido: {codigo}")
                                            
                                            # Introducir código dígito por dígito
                                            for i, digito in enumerate(codigo, 1):
                                                try:
                                                    input_field = driver.find_element(By.NAME, f"code-sub-{i}")
                                                    input_field.clear()
                                                    input_field.send_keys(digito)
                                                    time.sleep(0.2)
                                                    self.log(f"  ✓ Dígito {i}: {digito}")
                                                except Exception as e:
                                                    self.log(f"  ❌ Error introduciendo dígito {i}: {str(e)[:30]}")
                                            
                                            # Hacer click en el botón de verificación
                                            try:
                                                verify_button = driver.find_element(By.CSS_SELECTOR, "button.btn.btn-dark.btn-lg.btn-block[type='submit']")
                                                verify_button.click()
                                                time.sleep(3)
                                                self.log("  ✓ Código de verificación enviado")
                                                
                                                # Verificar si la verificación fue exitosa
                                                final_url = driver.current_url
                                                if "admin" in final_url.lower():
                                                    self.log("  ✅ Verificación 2FA exitosa")
                                                else:
                                                    self.log("  ⚠️ Verificación 2FA puede no haber sido exitosa")
                                                    
                                            except Exception as e:
                                                self.log(f"  ❌ Error haciendo click en verificar: {str(e)[:30]}")
                                        else:
                                            self.log("  ❌ No se pudo obtener código de verificación")
                                    else:
                                        self.log("  ✓ No se requiere verificación 2FA")
                                except:
                                    self.log("  ✓ No se requiere verificación 2FA")
                        else:
                            self.log("  ❌ No se encontró botón de login")
                            
                    except Exception as login_error:
                        self.log(f"  ❌ Error en login Dondominio: {str(login_error)[:50]}")
                    
                    # Ir a la página de dominios después del login
                    self.log("  📍 Navegando a lista de dominios...")
                    driver.get("https://www.dondominio.com/admin/domains/list/")
                    time.sleep(3)
                    
                    # Buscar dominio
                    self.log(f"  🔍 Buscando dominio: {domain_name_without_tld}")
                    try:
                        # Probar múltiples selectores para el campo de búsqueda
                        filter_input = None
                        selectors_filter = [
                            (By.NAME, "filterName"),
                            (By.NAME, "search"),
                            (By.NAME, "filter"),
                            (By.NAME, "q"),
                            (By.NAME, "query"),
                            (By.ID, "filterName"),
                            (By.ID, "search"),
                            (By.ID, "filter"),
                            (By.ID, "q"),
                            (By.ID, "query"),
                            (By.CSS_SELECTOR, "input[name='filterName']"),
                            (By.CSS_SELECTOR, "input[name='search']"),
                            (By.CSS_SELECTOR, "input[name='filter']"),
                            (By.CSS_SELECTOR, "input[type='text']"),
                            (By.CSS_SELECTOR, "input[placeholder*='buscar']"),
                            (By.CSS_SELECTOR, "input[placeholder*='search']"),
                            (By.CSS_SELECTOR, "input[placeholder*='filter']"),
                            (By.CSS_SELECTOR, "input[placeholder*='dominio']"),
                            (By.CSS_SELECTOR, "input[placeholder*='domain']"),
                            (By.XPATH, "//input[@type='text']"),
                            (By.XPATH, "//input[contains(@placeholder, 'buscar')]"),
                            (By.XPATH, "//input[contains(@placeholder, 'search')]"),
                            (By.XPATH, "//input[contains(@placeholder, 'filter')]")
                        ]
                        
                        for selector_type, selector_value in selectors_filter:
                            try:
                                filter_input = wait.until(EC.element_to_be_clickable((selector_type, selector_value)))
                                self.log(f"  ✓ Campo búsqueda encontrado con: {selector_type} = {selector_value}")
                                break
                            except:
                                continue
                        
                        # Si no se encontró, mostrar todos los inputs disponibles para debug
                        if not filter_input:
                            self.log("  🔍 No se encontró campo de búsqueda, mostrando inputs disponibles:")
                            try:
                                all_inputs = driver.find_elements(By.TAG_NAME, "input")
                                for i, inp in enumerate(all_inputs[:10]):  # Solo los primeros 10
                                    try:
                                        name = inp.get_attribute('name') or 'sin_name'
                                        id_attr = inp.get_attribute('id') or 'sin_id'
                                        placeholder = inp.get_attribute('placeholder') or 'sin_placeholder'
                                        input_type = inp.get_attribute('type') or 'sin_type'
                                        self.log(f"    Input {i+1}: name='{name}', id='{id_attr}', type='{input_type}', placeholder='{placeholder}'")
                                    except:
                                        self.log(f"    Input {i+1}: No se pudo obtener información")
                            except Exception as debug_error:
                                self.log(f"  ❌ Error en debug: {str(debug_error)[:30]}")
                        
                        if filter_input:
                            # Limpiar campo de forma más robusta
                            try:
                                filter_input.click()
                                time.sleep(0.5)
                                filter_input.clear()
                                time.sleep(0.5)
                                filter_input.send_keys(domain_name_without_tld)
                                time.sleep(1)
                                
                                # Presionar Enter para buscar
                                from selenium.webdriver.common.keys import Keys
                                filter_input.send_keys(Keys.RETURN)
                                time.sleep(3)
                                self.log("  ✓ Dominio ingresado y búsqueda iniciada (Enter)")
                                
                            except Exception as e:
                                self.log(f"  ⚠️ Error ingresando dominio: {str(e)[:30]}")
                                # Intentar método alternativo solo con JavaScript
                                try:
                                    # Método 1: JavaScript directo con eventos
                                    driver.execute_script("""
                                        var element = arguments[0];
                                        var text = arguments[1];
                                        element.focus();
                                        element.value = '';
                                        element.value = text;
                                        
                                        // Disparar todos los eventos necesarios
                                        element.dispatchEvent(new Event('focus', { bubbles: true }));
                                        element.dispatchEvent(new Event('input', { bubbles: true }));
                                        element.dispatchEvent(new Event('change', { bubbles: true }));
                                        element.dispatchEvent(new Event('blur', { bubbles: true }));
                                    """, filter_input, domain_name_without_tld)
                                    time.sleep(0.5)
                                    
                                    # Presionar Enter con JavaScript
                                    driver.execute_script("""
                                        var element = arguments[0];
                                        var enterEvent = new KeyboardEvent('keydown', {
                                            key: 'Enter',
                                            code: 'Enter',
                                            keyCode: 13,
                                            which: 13,
                                            bubbles: true
                                        });
                                        element.dispatchEvent(enterEvent);
                                        
                                        var enterEvent2 = new KeyboardEvent('keyup', {
                                            key: 'Enter',
                                            code: 'Enter',
                                            keyCode: 13,
                                            which: 13,
                                            bubbles: true
                                        });
                                        element.dispatchEvent(enterEvent2);
                                    """, filter_input)
                                    time.sleep(3)
                                    
                                    self.log("  ✓ Dominio ingresado y búsqueda iniciada (JavaScript + Enter)")
                                    
                                except Exception as e2:
                                    self.log(f"  ❌ Error método alternativo: {str(e2)[:30]}")
                            
                            # Verificar que el dominio se haya escrito correctamente
                            try:
                                valor_actual = filter_input.get_attribute('value')
                                if valor_actual == domain_name_without_tld:
                                    self.log("  ✅ Dominio verificado correctamente")
                                else:
                                    self.log(f"  ⚠️ Dominio no coincide: '{valor_actual}' (esperado: '{domain_name_without_tld}')")
                            except:
                                self.log("  ⚠️ No se pudo verificar el valor del dominio")
                        else:
                            self.log("  ❌ No se encontró campo de búsqueda")
                            raise Exception("Campo de búsqueda no encontrado")
                        
                        # Buscar y hacer click en el dominio específico después de la búsqueda
                        self.log(f"  🔍 Buscando dominio específico: {domain_name_without_tld}")
                        domain_link = None
                        
                        # Esperar a que se carguen los resultados de la búsqueda
                        time.sleep(2)
                        
                        # Probar múltiples selectores para encontrar el dominio
                        selectors_domain = [
                            # Selectores por texto del dominio
                            (By.XPATH, f"//a[contains(text(), '{domain_name_without_tld}')]"),
                            (By.XPATH, f"//a[contains(text(), '{domain_name_without_tld}.')]"),
                            (By.XPATH, f"//td[contains(text(), '{domain_name_without_tld}')]//a"),
                            (By.XPATH, f"//tr[contains(., '{domain_name_without_tld}')]//a"),
                            
                            # Selectores por href
                            (By.CSS_SELECTOR, f"a[href*='{domain_name_without_tld}']"),
                            (By.CSS_SELECTOR, f"a[href*='{domain_name_without_tld}.']"),
                            
                            # Selectores por partial link text
                            (By.PARTIAL_LINK_TEXT, domain_name_without_tld),
                            
                            # Selectores más genéricos
                            (By.XPATH, "//a[contains(@href, 'domain')]"),
                            (By.XPATH, "//a[contains(@href, 'domains')]"),
                            (By.CSS_SELECTOR, "a[href*='domain']"),
                            (By.CSS_SELECTOR, "a[href*='domains']"),
                            
                            # Selectores por clase o atributos comunes
                            (By.CSS_SELECTOR, ".domain-link"),
                            (By.CSS_SELECTOR, ".domain-name"),
                            (By.CSS_SELECTOR, "[data-domain]"),
                        ]
                        
                        for selector_type, selector_value in selectors_domain:
                            try:
                                if selector_type == By.XPATH:
                                    domain_link = wait.until(EC.element_to_be_clickable((By.XPATH, selector_value)))
                                elif selector_type == By.CSS_SELECTOR:
                                    domain_link = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector_value)))
                                elif selector_type == By.PARTIAL_LINK_TEXT:
                                    domain_link = wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, selector_value)))
                                
                                # Verificar que el enlace contiene el dominio correcto
                                link_text = domain_link.text.lower()
                                link_href = domain_link.get_attribute('href') or ''
                                
                                if (domain_name_without_tld.lower() in link_text or 
                                    domain_name_without_tld.lower() in link_href.lower()):
                                    self.log(f"  ✓ Dominio encontrado con: {selector_type} = {selector_value}")
                                    break
                                else:
                                    domain_link = None
                                    continue
                                    
                            except:
                                continue
                        
                        if domain_link:
                            # Hacer scroll al elemento y hacer click
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", domain_link)
                            time.sleep(1)
                            
                            # Intentar click normal primero
                            try:
                                domain_link.click()
                                time.sleep(3)
                                self.log("  ✓ Click en dominio realizado")
                            except:
                                # Si falla, intentar con JavaScript
                                try:
                                    driver.execute_script("arguments[0].click();", domain_link)
                                    time.sleep(3)
                                    self.log("  ✓ Click en dominio realizado (JavaScript)")
                                except Exception as click_error:
                                    self.log(f"  ❌ Error haciendo click: {str(click_error)[:30]}")
                                    raise
                        else:
                            self.log("  ❌ No se encontró el dominio en la lista")
                            
                            # Debug: mostrar todos los enlaces disponibles
                            try:
                                self.log("  🔍 Enlaces disponibles en la página:")
                                all_links = driver.find_elements(By.TAG_NAME, "a")
                                for i, link in enumerate(all_links[:10]):  # Solo los primeros 10
                                    try:
                                        text = link.text.strip()[:50] or 'sin_texto'
                                        href = link.get_attribute('href') or 'sin_href'
                                        self.log(f"    Link {i+1}: '{text}' -> {href[:50]}")
                                    except:
                                        self.log(f"    Link {i+1}: No se pudo obtener información")
                            except:
                                pass
                            
                            raise Exception("Dominio no encontrado en la lista")
                            
                    except Exception as search_error:
                        self.log(f"  ❌ Error buscando dominio: {str(search_error)[:50]}")
                        raise
                    
                    # Zoom al 90%
                    driver.execute_script("document.body.style.zoom='0.9'")
                    time.sleep(1)
                
                    # Captura 1: Pantalla completa
                    screenshot_filename = f'{self.screenshots_dir}/dondominio_completa_{timestamp}.png'
                    pyautogui.screenshot(screenshot_filename)
                    screenshots.append(screenshot_filename)
                    self.log("  ✓ Dondominio completa")
                    
                    # Captura 2: Elemento específico
                    try:
                        elemento = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[1]/div[4]/div/div[3]/div/div/div/div[2]/div[2]")))
                        driver.execute_script("arguments[0].scrollIntoView({block: 'start'});", elemento)
                        driver.execute_script("window.scrollBy(0, -100);")
                        time.sleep(0.5)
                        
                        screenshot_filename = f'{self.screenshots_dir}/dondominio_elemento_{timestamp}.png'
                        elemento.screenshot(screenshot_filename)
                        screenshots.append(screenshot_filename)
                        self.log("  ✓ Dondominio elemento")
                    except Exception as e:
                        self.log(f"  ✗ Dondominio elemento: {str(e)[:30]}")
                    
                    # Click en el enlace específico
                    try:
                        enlace = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[1]/div[4]/div/div[3]/div/div/div/div[1]/div[2]/div[1]/div[3]/div[2]/div[1]/div[1]/a")))
                        enlace.click()
                        time.sleep(3)
                        
                        # Captura final: Pantalla completa de la nueva página
                        screenshot_filename = f'{self.screenshots_dir}/dondominio_final_{timestamp}.png'
                        pyautogui.screenshot(screenshot_filename)
                        screenshots.append(screenshot_filename)
                        self.log("  ✓ Dondominio final")
                    except Exception as e:
                        self.log(f"  ✗ Dondominio enlace: {str(e)[:30]}")
                    
                    return screenshots
                    
                except Exception as e:
                    self.log(f"  ❌ Error con Dondominio: {str(e)[:30]}")
                    return []
        
        except Exception as e:
            self.log(f"  💥 Error general: {str(e)[:30]}")
            return []

    def obtener_nombre_web_con_ia_gui(self, url):
        """Obtiene el nombre de la web usando IA local para la GUI"""
        try:
            self.log("  🤖 Obteniendo nombre con IA...")
            
            # Configurar Chrome en modo headless para extracción rápida
            chrome_options = Options()
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            driver_temp = webdriver.Chrome(options=chrome_options)
            
            try:
                driver_temp.get(url)
                time.sleep(3)
                
                # Extraer datos básicos
                title = ""
                h1_text = ""
                meta_description = ""
                domain = ""
                
                try:
                    title = driver_temp.find_element(By.TAG_NAME, 'title').text
                except:
                    pass
                    
                try:
                    h1_element = driver_temp.find_element(By.TAG_NAME, 'h1')
                    h1_text = h1_element.text
                except:
                    pass
                
                try:
                    meta_desc_element = driver_temp.find_element(By.CSS_SELECTOR, 'meta[name="description"]')
                    meta_description = meta_desc_element.get_attribute('content') or ""
                except:
                    pass
                
                # Extraer dominio
                try:
                    domain = urlparse(url).netloc.replace('www.', '')
                except:
                    pass
                
                # Crear prompt para la IA
                prompt = f"""Analiza estos datos de una página web y proporciona SOLO el nombre de la empresa o sitio web, sin explicaciones adicionales.

URL: {url}
Dominio: {domain}
Título: {title}
H1: {h1_text}
Meta descripción: {meta_description[:200]}

Responde únicamente con el nombre de la empresa/sitio web, máximo 3 palabras."""
                
                # Consultar IA local
                respuesta = consultar_ia_local(prompt)
                
                if respuesta and isinstance(respuesta, str):
                    nombre_web = respuesta.strip()
                    # Limpiar la respuesta
                    nombre_web = nombre_web.replace('"', '').replace("'", '').strip()
                    if len(nombre_web) > 50:  # Si es muy largo, usar el dominio
                        nombre_web = domain.split('.')[0].title()
                    self.log(f"  ✓ Nombre: {nombre_web}")
                    return nombre_web
                else:
                    self.log("  ❌ Error IA, usando dominio")
                    return domain.split('.')[0].title() if domain else "Sitio Web"
                    
            finally:
                driver_temp.quit()
                
        except Exception as e:
            self.log(f"  ❌ Error obteniendo nombre: {str(e)[:30]}")
            return None

    def analizar_competencia_con_ia(self, url, keywords, driver):
        """Analiza la competencia del cliente usando IA local"""
        try:
            self.log("  🤖 Analizando competencia con IA...")
            
            # Obtener información del sitio web del cliente
            driver.get(url)
            time.sleep(3)
            
            # Extraer datos del sitio web del cliente
            cliente_data = self.extraer_datos_sitio_web(driver)
            
            # Usar los competidores que ya se generaron automáticamente
            competidores_texto = self.competidores_entry.get().strip()
            if competidores_texto:
                competidores_encontrados = [comp.strip() for comp in competidores_texto.split(',') if comp.strip()]
            else:
                competidores_encontrados = []
            
            # Log de competidores encontrados para análisis
            self.log(f"    📊 Competidores encontrados para análisis: {len(competidores_encontrados)}")
            if competidores_encontrados:
                for i, comp in enumerate(competidores_encontrados, 1):
                    self.log(f"      {i}. {comp}")
            
            # Analizar competidores encontrados
            analisis_competidores = []
            for competidor_url in competidores_encontrados[:2]:  # Analizar máximo 2 competidores
                try:
                    self.log(f"    🔍 Analizando competidor: {competidor_url}")
                    driver.get(competidor_url)
                    time.sleep(2)
                    competidor_data = self.extraer_datos_sitio_web(driver)
                    analisis_competidores.append({
                        'url': competidor_url,
                        'data': competidor_data
                    })
                    self.log(f"    ✅ Competidor analizado exitosamente: {competidor_data.get('title', 'Sin título')[:50]}")
                except Exception as e:
                    self.log(f"    ⚠️ Error analizando competidor {competidor_url}: {str(e)[:30]}")
                    continue
            
            # Crear prompt para análisis de competencia
            prompt = f"""Analiza brevemente los competidores encontrados para este sitio web.

SITIO WEB CLIENTE:
URL: {url}
Título: {cliente_data.get('title', '')}
H1: {cliente_data.get('h1', '')}
Keywords principales: {', '.join(keywords)}

COMPETIDORES ENCONTRADOS:
{self.formatear_datos_competidores(analisis_competidores)}

Proporciona un análisis que incluya:

1. COMPETIDORES DIRECTOS:
   - Para cada competidor (máximo 2): 1-2 líneas describiendo a qué se dedica
   - Sin explicaciones exhaustivas ni recomendaciones
   - Solo descripción básica de servicios/negocio

2. ANÁLISIS DEL CLIENTE:
   - 1-2 líneas describiendo a qué se dedica la empresa del cliente
   - Basado en el título, H1 y keywords principales

3. ANÁLISIS COMPARATIVO:
   - 2-3 líneas comparando los 3 sitios web juntos
   - Similitudes y diferencias en servicios/negocio
   - Posicionamiento en el mercado

Formato esperado:
COMPETIDOR 1: [URL] - [1-2 líneas sobre a qué se dedica]
COMPETIDOR 2: [URL] - [1-2 líneas sobre a qué se dedica]

ANÁLISIS CLIENTE:
[1-2 líneas sobre a qué se dedica la empresa del cliente]

COMPARACIÓN GENERAL:
[2-3 líneas comparando los 3 sitios web]

No añadas asteriscos ni otros caracteres especiales.
Mantén el análisis conciso, directo y sin conclusiones extensas."""
            
            # Consultar IA local
            respuesta = consultar_ia_local(prompt)
            
            if respuesta and isinstance(respuesta, str):
                self.log("  ✓ Análisis de competencia completado")
                return {
                    'analisis': respuesta,
                    'competidores_encontrados': competidores_encontrados,
                    'datos_cliente': cliente_data
                }
            else:
                self.log("  ❌ Error en análisis de competencia")
                return None
                
        except Exception as e:
            self.log(f"  ❌ Error analizando competencia: {str(e)[:30]}")
            return None

    def extraer_datos_sitio_web(self, driver):
        """Extrae datos básicos de un sitio web"""
        try:
            data = {}
            
            # Título
            try:
                data['title'] = driver.find_element(By.TAG_NAME, 'title').text
            except:
                data['title'] = ""
            
            # H1
            try:
                h1_element = driver.find_element(By.TAG_NAME, 'h1')
                data['h1'] = h1_element.text
            except:
                data['h1'] = ""
            
            # Meta descripción
            try:
                meta_desc_element = driver.find_element(By.CSS_SELECTOR, 'meta[name="description"]')
                data['meta_description'] = meta_desc_element.get_attribute('content') or ""
            except:
                data['meta_description'] = ""
            
            # URL actual
            data['url'] = driver.current_url
            
            return data
            
        except Exception as e:
            return {'title': '', 'h1': '', 'meta_description': '', 'url': driver.current_url}

    def buscar_competidores_duckduckgo(self, keywords, driver):
        """Busca competidores en DuckDuckGo usando las keywords principales"""
        try:
            competidores = []
            
            # Usar las keywords para buscar competidores reales
            if keywords:
                # Combinar las primeras 2 keywords para una búsqueda más específica
                if len(keywords) >= 2:
                    search_query = f'"{keywords[0]}" "{keywords[1]}" empresas servicios'
                else:
                    search_query = f'"{keywords[0]}" empresas servicios competencia'
                
                self.log(f"    🔍 Buscando competidores para: {search_query}")
                self.log(f"    🔗 URL de búsqueda: {search_url}")
                
                # Buscar competidores reales del sector con términos más específicos
                # Usar solo la primera keyword para evitar búsquedas demasiado específicas
                search_terms = f'"{keywords[0]}" empresas servicios'
                search_url = f"https://www.bing.com/search?q={search_terms.replace(' ', '+')}"
                driver.get(search_url)
                time.sleep(3)
                
                # Extraer URLs de los resultados
                try:
                    # Intentar múltiples selectores para encontrar enlaces (DuckDuckGo)
                    selectores = [
                        'a[data-testid="result-title-a"]',
                        'a[href*="http"]',
                        'h2 a[href*="http"]',
                        'div[data-testid="result"] a[href*="http"]',
                        'a[href*="http"]:not([href*="duckduckgo.com"])',
                        'h2 a',
                        'a'
                    ]
                    
                    result_links = []
                    for selector in selectores:
                        try:
                            links = driver.find_elements(By.CSS_SELECTOR, selector)
                            if links:
                                result_links.extend(links)
                                self.log(f"    ✓ Encontrados {len(links)} enlaces con selector: {selector}")
                                break
                        except:
                            continue
                    
                    if not result_links:
                        # Fallback: buscar todos los enlaces
                        result_links = driver.find_elements(By.TAG_NAME, 'a')
                        self.log(f"    ⚠️ Usando fallback: {len(result_links)} enlaces totales")
                    
                    for link in result_links[:15]:  # Primeros 15 resultados
                        try:
                            href = link.get_attribute('href')
                            if href and 'duckduckgo.com' not in href and 'duckduckgo' not in href.lower() and 'youtube.com' not in href and 'facebook.com' not in href and 'play.google.com' not in href and 'apps.apple.com' not in href and 'itunes.apple.com' not in href and 'playstore' not in href.lower() and 'appstore' not in href.lower():
                                # Limpiar URL
                                if '/url?q=' in href:
                                    href = href.split('/url?q=')[1].split('&')[0]
                                if href.startswith('http') and href not in competidores:
                                    # Verificar que sea un sitio web real (no app stores, redes sociales, etc.)
                                    dominios_excluidos = [
                                        'duckduckgo.com', 'duckduckgo', 'google.com', 'youtube.com', 'facebook.com', 
                                        'instagram.com', 'twitter.com', 'linkedin.com', 'tiktok.com',
                                        'play.google.com', 'apps.apple.com', 'itunes.apple.com',
                                        'amazon.com', 'wikipedia.org', 'reddit.com', 'pinterest.com'
                                    ]
                                    
                                    es_dominio_valido = True
                                    for dominio in dominios_excluidos:
                                        if dominio in href.lower():
                                            es_dominio_valido = False
                                            break
                                    
                                    if es_dominio_valido:
                                        # Verificar que el enlace tenga un dominio válido (no sea muy genérico)
                                        try:
                                            from urllib.parse import urlparse
                                            parsed_url = urlparse(href)
                                            domain = parsed_url.netloc.lower()
                                            
                                            # Excluir dominios genéricos o de plataformas
                                            dominios_genericos = [
                                                'github.com', 'stackoverflow.com', 'medium.com', 
                                                'wordpress.com', 'blogspot.com', 'tumblr.com',
                                                'wix.com', 'squarespace.com', 'weebly.com'
                                            ]
                                            
                                            if not any(gen in domain for gen in dominios_genericos):
                                                competidores.append(href)
                                                self.log(f"    ✓ Competidor encontrado: {href}")
                                                if len(competidores) >= 5:  # Limitar a 5 competidores
                                                    break
                                        except:
                                            # Si hay error parseando la URL, saltar
                                            continue
                        except:
                            continue
                            
                except Exception as e:
                    self.log(f"    ⚠️ Error extrayendo resultados: {str(e)[:30]}")
            
            # Si no encontramos suficientes competidores, intentar búsqueda alternativa
            if len(competidores) < 2 and keywords:
                self.log(f"    ⚠️ Pocos competidores encontrados, intentando búsqueda alternativa...")
                
                # Búsqueda alternativa más específica
                if len(keywords) >= 2:
                    alt_query = f'"{keywords[0]}" "{keywords[1]}" empresas'
                else:
                    alt_query = f'"{keywords[0]}" empresas del sector'
                
                alt_search_url = f"https://www.bing.com/search?q={alt_query.replace(' ', '+')}"
                driver.get(alt_search_url)
                time.sleep(3)
                
                # Repetir el proceso de extracción
                try:
                    result_links = []
                    for selector in selectores:
                        try:
                            links = driver.find_elements(By.CSS_SELECTOR, selector)
                            if links:
                                result_links.extend(links)
                                break
                        except:
                            continue
                    
                    for link in result_links[:10]:
                        try:
                            href = link.get_attribute('href')
                            if href and 'duckduckgo.com' not in href and 'duckduckgo' not in href.lower() and 'youtube.com' not in href and 'facebook.com' not in href and 'play.google.com' not in href and 'apps.apple.com' not in href and 'itunes.apple.com' not in href and 'playstore' not in href.lower() and 'appstore' not in href.lower():
                                if href.startswith('http') and href not in competidores:
                                    # Verificar que sea un sitio web real
                                    dominios_excluidos = [
                                        'duckduckgo.com', 'duckduckgo', 'google.com', 'youtube.com', 'facebook.com', 
                                        'instagram.com', 'twitter.com', 'linkedin.com', 'tiktok.com',
                                        'play.google.com', 'apps.apple.com', 'itunes.apple.com',
                                        'amazon.com', 'wikipedia.org', 'reddit.com', 'pinterest.com'
                                    ]
                                    
                                    es_dominio_valido = True
                                    for dominio in dominios_excluidos:
                                        if dominio in href.lower():
                                            es_dominio_valido = False
                                            break
                                    
                                    if es_dominio_valido:
                                        try:
                                            from urllib.parse import urlparse
                                            parsed_url = urlparse(href)
                                            domain = parsed_url.netloc.lower()
                                            
                                            dominios_genericos = [
                                                'github.com', 'stackoverflow.com', 'medium.com', 
                                                'wordpress.com', 'blogspot.com', 'tumblr.com',
                                                'wix.com', 'squarespace.com', 'weebly.com'
                                            ]
                                            
                                            if not any(gen in domain for gen in dominios_genericos):
                                                competidores.append(href)
                                                self.log(f"    ✓ Competidor alternativo encontrado: {href}")
                                                if len(competidores) >= 3:
                                                    break
                                        except:
                                            continue
                        except:
                            continue
                except Exception as e:
                    self.log(f"    ⚠️ Error en búsqueda alternativa: {str(e)[:30]}")
            
            self.log(f"    ✓ Encontrados {len(competidores)} competidores potenciales")
            
            # Log detallado de competidores encontrados
            if competidores:
                self.log(f"    📋 Lista de competidores encontrados:")
                for i, competidor in enumerate(competidores, 1):
                    self.log(f"      {i}. {competidor}")
            else:
                self.log(f"    ⚠️ No se encontraron competidores válidos")
            
            return competidores
            
        except Exception as e:
            self.log(f"    ❌ Error buscando competidores: {str(e)[:30]}")
            return []

    def formatear_datos_competidores(self, analisis_competidores):
        """Formatea los datos de competidores para el prompt de IA"""
        if not analisis_competidores:
            return "No se encontraron competidores para analizar."
        
        texto = ""
        for i, competidor in enumerate(analisis_competidores, 1):
            data = competidor['data']
            texto += f"""
COMPETIDOR {i}:
URL: {competidor['url']}
Título: {data.get('title', '')}
H1: {data.get('h1', '')}
Meta descripción: {data.get('meta_description', '')}
"""
        
        return texto

    def buscar_competidores_simple(self, driver, keywords, url):
        """Busca competidores usando estrategias más directas y menos detectables"""
        try:
            competidores = []
            
            if keywords:
                keyword_principal = keywords[0]
                self.log(f"    🔍 Buscando competidores para: {keyword_principal}")
                
                # Usar SerpAPI para búsqueda confiable
                import requests
                
                # Configuración de SerpAPI
                api_key = "d6627f12f17390dd5a229ef805199b662065d1fc8e1276c686fb270381a242a2"
                search_terms = f'"{keyword_principal}" empresa servicios'
                
                self.log(f"    🔗 Buscando con SerpAPI: {search_terms}")
                
                # Parámetros para SerpAPI
                params = {
                    "q": search_terms,
                    "api_key": api_key,
                    "engine": "google",
                    "num": 10,
                    "gl": "es",  # España
                    "hl": "es"   # Español
                }
                
                try:
                    response = requests.get("https://serpapi.com/search", params=params, timeout=10)
                    
                    if response.status_code == 200:
                        data = response.json()
                        
                        # Extraer resultados orgánicos
                        organic_results = data.get("organic_results", [])
                        
                        self.log(f"    ✓ Encontrados {len(organic_results)} resultados orgánicos")
                        
                        # Dominios a excluir
                        dominios_excluidos = [
                            'google.com', 'youtube.com', 'facebook.com', 'instagram.com', 'twitter.com', 
                            'linkedin.com', 'wikipedia.org', 'amazon.com', 'reddit.com', 'github.com',
                            'stackoverflow.com', 'medium.com', 'wordpress.com', 'blogspot.com',
                            'play.google.com', 'apps.apple.com', 'duckduckgo.com', 'bing.com', 'microsoft.com',
                            'pinterest.com', 'tumblr.com', 'flickr.com', 'vimeo.com',
                            'startpage.com', 'startmail.com', 'yahoo.com', 'msn.com'
                        ]
                        
                        for result in organic_results:
                            try:
                                href = result.get("link", "")
                                if not href or not href.startswith('http'):
                                    continue
                                
                                self.log(f"    🔍 Resultado encontrado: {href}")
                                
                                # Verificar si es un dominio válido
                                es_valido = True
                                for dominio in dominios_excluidos:
                                    if dominio in href.lower():
                                        es_valido = False
                                        self.log(f"    ❌ Excluido por: {dominio}")
                                        break
                                
                                # Crear URL limpia para comparación
                                from urllib.parse import urlparse
                                parsed_url = urlparse(href)
                                domain = parsed_url.netloc.lower()
                                clean_url = f"https://{domain}"
                                
                                if es_valido and clean_url not in competidores:
                                    # Verificación rápida: solo verificar que tenga un dominio válido
                                    try:
                                        # Verificar que sea un dominio válido
                                        if ('.com' in domain or '.es' in domain or '.org' in domain or '.net' in domain or '.info' in domain) and len(domain.split('.')) >= 2:
                                            # Verificación básica: que no sea un dominio obviamente inválido
                                            if not any(x in domain for x in ['localhost', '127.0.0.1', '0.0.0.0', 'example.com', 'test.com', 'duck.ai', 'ai.', 'chat.', 'bot.']):
                                                # Verificación adicional: que no sea un sitio de IA, chat, o herramientas
                                                if not any(x in domain.lower() for x in ['ai', 'chat', 'bot', 'tool', 'app', 'api', 'dev', 'tech', 'software', 'platform', 'service']):
                                                    # Verificación final: que el dominio parezca ser de una empresa real
                                                    if len(domain.split('.')[0]) > 3 and not domain.split('.')[0].isdigit():
                                                        competidores.append(clean_url)
                                                        self.log(f"    ✅ Competidor válido (intento {len(competidores)}): {clean_url}")
                                                        if len(competidores) >= 6:
                                                            break
                                                    else:
                                                        self.log(f"    ❌ Dominio muy corto o numérico: {href}")
                                                else:
                                                    self.log(f"    ❌ Dominio de herramienta/IA: {href}")
                                            else:
                                                self.log(f"    ❌ Dominio inválido: {href}")
                                        else:
                                            self.log(f"    ❌ Dominio no válido: {href}")
                                            
                                    except Exception as e:
                                        self.log(f"    ❌ Error verificando: {href} - {str(e)[:30]}")
                                        continue
                                elif clean_url in competidores:
                                    self.log(f"    ⚠️ Ya existe: {clean_url}")
                                    
                            except Exception as e:
                                self.log(f"    ❌ Error procesando resultado: {str(e)[:30]}")
                                continue
                                
                    else:
                        self.log(f"    ❌ Error en SerpAPI: {response.status_code}")
                        
                except requests.exceptions.Timeout:
                    self.log("    ❌ Timeout en SerpAPI")
                except requests.exceptions.RequestException as e:
                    self.log(f"    ❌ Error de conexión con SerpAPI: {str(e)[:30]}")
                except Exception as e:
                    self.log(f"    ❌ Error inesperado con SerpAPI: {str(e)[:30]}")
                
                # Si no encontramos suficientes competidores, intentar con segunda keyword
                if len(competidores) < 2 and len(keywords) > 1:
                    self.log(f"    ⚠️ Intentando con segunda keyword...")
                    
                    keyword_secundaria = keywords[1]
                    search_terms_sec = f'"{keyword_secundaria}" empresa servicios'
                    
                    self.log(f"    🔗 Búsqueda con keyword secundaria: {search_terms_sec}")
                    
                    params_sec = {
                        "q": search_terms_sec,
                        "api_key": api_key,
                        "engine": "google",
                        "num": 10,
                        "gl": "es",
                        "hl": "es"
                    }
                    
                    try:
                        response_sec = requests.get("https://serpapi.com/search", params=params_sec, timeout=10)
                        
                        if response_sec.status_code == 200:
                            data_sec = response_sec.json()
                            organic_results_sec = data_sec.get("organic_results", [])
                            
                            self.log(f"    ✓ Encontrados {len(organic_results_sec)} resultados adicionales")
                            
                            for result in organic_results_sec:
                                try:
                                    href = result.get("link", "")
                                    if not href or not href.startswith('http'):
                                        continue
                                    
                                    # Verificar si es un dominio válido
                                    es_valido = True
                                    for dominio in dominios_excluidos:
                                        if dominio in href.lower():
                                            es_valido = False
                                            break
                                    
                                    # Crear URL limpia para comparación
                                    from urllib.parse import urlparse
                                    parsed_url = urlparse(href)
                                    domain = parsed_url.netloc.lower()
                                    clean_url = f"https://{domain}"
                                    
                                    if es_valido and clean_url not in competidores:
                                        if ('.com' in domain or '.es' in domain or '.org' in domain or '.net' in domain or '.info' in domain) and len(domain.split('.')) >= 2:
                                            if not any(x in domain for x in ['localhost', '127.0.0.1', '0.0.0.0', 'example.com', 'test.com', 'duck.ai', 'ai.', 'chat.', 'bot.']):
                                                if not any(x in domain.lower() for x in ['ai', 'chat', 'bot', 'tool', 'app', 'api', 'dev', 'tech', 'software', 'platform', 'service']):
                                                    if len(domain.split('.')[0]) > 3 and not domain.split('.')[0].isdigit():
                                                        competidores.append(clean_url)
                                                        self.log(f"    ✅ Competidor adicional: {clean_url}")
                                                        if len(competidores) >= 2:
                                                            break
                                except:
                                    continue
                                    
                    except Exception as e:
                        self.log(f"    ❌ Error en búsqueda secundaria: {str(e)[:30]}")
                
                
                # Si aún no tenemos suficientes, intentar con términos relacionados
                if len(competidores) < 6 and len(keywords) > 1:
                    self.log(f"    ⚠️ Intentando con segunda keyword...")
                    
                    # Usar la segunda keyword
                    keyword_secundaria = keywords[1] if len(keywords) > 1 else keywords[0]
                    search_terms_sec = f"{keyword_secundaria} empresa servicios"
                    search_url_sec = f"https://www.google.com/search?q={search_terms_sec.replace(' ', '+')}"
                    
                    self.log(f"    🔗 Búsqueda con keyword secundaria: {search_url_sec}")
                    driver.get(search_url_sec)
                    time.sleep(3)
                    
                    # Obtener enlaces con segunda keyword
                    sec_links = driver.find_elements(By.TAG_NAME, 'a')
                    
                    for link in sec_links[:30]:
                        try:
                            href = link.get_attribute('href')
                            if not href or not href.startswith('http'):
                                continue
                            
                            es_valido = True
                            for dominio in dominios_excluidos:
                                if dominio in href.lower():
                                    es_valido = False
                                    break
                            
                            if es_valido and href not in competidores:
                                try:
                                    from urllib.parse import urlparse
                                    parsed_url = urlparse(href)
                                    domain = parsed_url.netloc.lower()
                                    
                                    if ('.com' in domain or '.es' in domain or '.org' in domain or '.net' in domain) and len(domain.split('.')) >= 2:
                                        competidores.append(href)
                                        self.log(f"    ✓ Competidor con keyword secundaria: {href}")
                                        if len(competidores) >= 6:
                                            break
                                except:
                                    continue
                        except:
                            continue
                
                # Último recurso: intentar con Google si aún no tenemos suficientes
                if len(competidores) < 6:
                    self.log(f"    ⚠️ Último recurso: intentando con Google...")
                    
                    # Búsqueda en Google como último recurso
                    google_search_terms = f"{keyword_principal} empresa"
                    google_search_url = f"https://www.google.com/search?q={google_search_terms.replace(' ', '+')}"
                    
                    self.log(f"    🔗 Búsqueda en Google: {google_search_url}")
                    driver.get(google_search_url)
                    time.sleep(4)
                    
                    # Intentar múltiples selectores para Google
                    google_selectores = [
                        'div[data-ved] a[href*="http"]',
                        'h3 a[href*="http"]',
                        'div.g a[href*="http"]',
                        'a[href*="http"]'
                    ]
                    
                    google_links = []
                    for selector in google_selectores:
                        try:
                            links = driver.find_elements(By.CSS_SELECTOR, selector)
                            if links:
                                google_links.extend(links)
                                self.log(f"    ✓ Encontrados {len(links)} enlaces en Google con selector: {selector}")
                                break
                        except:
                            continue
                    
                    if not google_links:
                        google_links = driver.find_elements(By.TAG_NAME, 'a')
                    
                    for link in google_links[:25]:
                        try:
                            href = link.get_attribute('href')
                            if not href or not href.startswith('http'):
                                continue
                            
                            # Limpiar URL de Google
                            if '/url?q=' in href:
                                href = href.split('/url?q=')[1].split('&')[0]
                            
                            es_valido = True
                            for dominio in dominios_excluidos:
                                if dominio in href.lower():
                                    es_valido = False
                                    break
                            
                            if es_valido and href not in competidores:
                                try:
                                    from urllib.parse import urlparse
                                    parsed_url = urlparse(href)
                                    domain = parsed_url.netloc.lower()
                                    
                                    if ('.com' in domain or '.es' in domain or '.org' in domain or '.net' in domain) and len(domain.split('.')) >= 2:
                                        competidores.append(href)
                                        self.log(f"    ✓ Competidor de Google encontrado: {href}")
                                        if len(competidores) >= 6:
                                            break
                                except:
                                    continue
                        except:
                            continue
                        
            self.log(f"    ✓ Encontrados {len(competidores)} competidores")
            
            # OBLIGATORIO: Debe encontrar exactamente 2 competidores válidos
            if len(competidores) >= 2:
                # Si tenemos 2 o más, usar IA para seleccionar los mejores
                if len(competidores) > 2:
                    self.log(f"    🤖 Seleccionando los 2 mejores competidores con IA...")
                    competidores_seleccionados = self.seleccionar_mejores_competidores_ia(competidores, keywords, url)
                    return competidores_seleccionados
                else:
                    return competidores
            else:
                # Si no tenemos suficientes competidores, intentar búsquedas adicionales
                self.log(f"    ❌ Solo se encontraron {len(competidores)} competidores, se necesitan 2")
                return self.buscar_competidores_persistente(driver, keywords, url, competidores)
            
        except Exception as e:
            self.log(f"    ❌ Error buscando competidores: {str(e)[:30]}")
            return []

    def buscar_competidores_persistente(self, driver, keywords, url, competidores_actuales):
        """Busca competidores de forma persistente hasta encontrar exactamente 2"""
        try:
            competidores = competidores_actuales.copy()
            intentos = 0
            max_intentos = 10
            
            while len(competidores) < 2 and intentos < max_intentos:
                intentos += 1
                self.log(f"    🔄 Intento {intentos}/{max_intentos} - Buscando más competidores...")
                
                # Estrategias adicionales de búsqueda
                if intentos == 1:
                    # Estrategia: Búsqueda con términos más generales
                    keyword_principal = keywords[0]
                    search_terms = f"{keyword_principal} servicios"
                    search_url = f"https://www.bing.com/search?q={search_terms.replace(' ', '+')}"
                    self.log(f"    🔗 Búsqueda general: {search_url}")
                    
                elif intentos == 2:
                    # Estrategia: Búsqueda con "empresa" solamente
                    keyword_principal = keywords[0]
                    search_terms = f"{keyword_principal} empresa"
                    search_url = f"https://www.bing.com/search?q={search_terms.replace(' ', '+')}"
                    self.log(f"    🔗 Búsqueda empresa: {search_url}")
                    
                elif intentos == 3:
                    # Estrategia: Búsqueda con segunda keyword
                    if len(keywords) > 1:
                        keyword_secundaria = keywords[1]
                        search_terms = f"{keyword_secundaria} empresa"
                        search_url = f"https://www.bing.com/search?q={search_terms.replace(' ', '+')}"
                        self.log(f"    🔗 Búsqueda keyword 2: {search_url}")
                    else:
                        continue
                        
                elif intentos == 4:
                    # Estrategia: Búsqueda con tercera keyword
                    if len(keywords) > 2:
                        keyword_terciaria = keywords[2]
                        search_terms = f"{keyword_terciaria} empresa"
                        search_url = f"https://www.bing.com/search?q={search_terms.replace(' ', '+')}"
                        self.log(f"    🔗 Búsqueda keyword 3: {search_url}")
                    else:
                        continue
                        
                elif intentos == 5:
                    # Estrategia: Búsqueda en Google
                    keyword_principal = keywords[0]
                    search_terms = f"{keyword_principal} empresa"
                    search_url = f"https://www.bing.com/search?q={search_terms.replace(' ', '+')}"
                    self.log(f"    🔗 Búsqueda Google: {search_url}")
                    
                elif intentos == 6:
                    # Estrategia: Búsqueda con términos relacionados
                    keyword_principal = keywords[0]
                    search_terms = f"{keyword_principal} negocio"
                    search_url = f"https://www.bing.com/search?q={search_terms.replace(' ', '+')}"
                    self.log(f"    🔗 Búsqueda negocio: {search_url}")
                    
                elif intentos == 7:
                    # Estrategia: Búsqueda con "compañía"
                    keyword_principal = keywords[0]
                    search_terms = f"{keyword_principal} compañía"
                    search_url = f"https://www.bing.com/search?q={search_terms.replace(' ', '+')}"
                    self.log(f"    🔗 Búsqueda compañía: {search_url}")
                    
                elif intentos == 8:
                    # Estrategia: Búsqueda con "corporación"
                    keyword_principal = keywords[0]
                    search_terms = f"{keyword_principal} corporación"
                    search_url = f"https://www.bing.com/search?q={search_terms.replace(' ', '+')}"
                    self.log(f"    🔗 Búsqueda corporación: {search_url}")
                    
                elif intentos == 9:
                    # Estrategia: Búsqueda con "organización"
                    keyword_principal = keywords[0]
                    search_terms = f"{keyword_principal} organización"
                    search_url = f"https://www.bing.com/search?q={search_terms.replace(' ', '+')}"
                    self.log(f"    🔗 Búsqueda organización: {search_url}")
                    
                else:
                    # Estrategia: Búsqueda aleatoria con términos genéricos
                    terminos_genericos = ["servicios", "soluciones", "consultoría", "asesoría", "gestión"]
                    termino = terminos_genericos[intentos % len(terminos_genericos)]
                    keyword_principal = keywords[0]
                    search_terms = f"{keyword_principal} {termino}"
                    search_url = f"https://www.bing.com/search?q={search_terms.replace(' ', '+')}"
                    self.log(f"    🔗 Búsqueda {termino}: {search_url}")
                
                # Ejecutar búsqueda
                driver.get(search_url)
                time.sleep(3)
                
                # Obtener enlaces
                if "google.com" in search_url:
                    # Selectores para Google
                    selectores = [
                        'div[data-ved] a[href*="http"]',
                        'h3 a[href*="http"]',
                        'div.g a[href*="http"]',
                        'a[href*="http"]'
                    ]
                else:
                    # Selectores para DuckDuckGo
                    selectores = [
                        'a[data-testid="result-title-a"]',
                        'h2 a[href*="http"]',
                        'div.result__body a[href*="http"]',
                        'a[href*="http"]'
                    ]
                
                result_links = []
                for selector in selectores:
                    try:
                        links = driver.find_elements(By.CSS_SELECTOR, selector)
                        if links:
                            result_links.extend(links)
                            break
                    except:
                        continue
                
                if not result_links:
                    result_links = driver.find_elements(By.TAG_NAME, 'a')
                
                # Dominios a excluir
                dominios_excluidos = [
                    'google.com', 'youtube.com', 'facebook.com', 'instagram.com', 'twitter.com', 
                    'linkedin.com', 'wikipedia.org', 'amazon.com', 'reddit.com', 'github.com',
                    'stackoverflow.com', 'medium.com', 'wordpress.com', 'blogspot.com',
                    'play.google.com', 'apps.apple.com', 'duckduckgo.com', 'bing.com', 'microsoft.com',
                    'pinterest.com', 'tumblr.com', 'flickr.com', 'vimeo.com',
                    'startpage.com', 'startmail.com', 'yahoo.com', 'msn.com'
                ]
                
                # Procesar enlaces encontrados
                for link in result_links[:30]:
                    try:
                        href = link.get_attribute('href')
                        if not href or not href.startswith('http'):
                            continue
                        
                        # Limpiar URL de Google
                        if '/url?q=' in href:
                            href = href.split('/url?q=')[1].split('&')[0]
                        
                        # Debug: mostrar todos los enlaces encontrados
                        self.log(f"    🔍 Enlace encontrado: {href}")
                        
                        # Verificar si es un dominio válido
                        es_valido = True
                        for dominio in dominios_excluidos:
                            if dominio in href.lower():
                                es_valido = False
                                self.log(f"    ❌ Excluido por dominio: {dominio}")
                                break
                        
                        if es_valido and href not in competidores:
                            # Verificar que la página responda y no esté en venta
                            try:
                                import requests
                                
                                # Configurar headers para parecer un navegador real
                                headers = {
                                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                                    'Accept-Language': 'es-ES,es;q=0.8,en-US;q=0.5,en;q=0.3',
                                    'Accept-Encoding': 'gzip, deflate',
                                    'Connection': 'keep-alive',
                                    'Upgrade-Insecure-Requests': '1'
                                }
                                
                                # Hacer petición con timeout corto
                                response = requests.get(href, headers=headers, timeout=5, allow_redirects=True)
                                
                                if response.status_code == 200:
                                    # Verificar que no sea una página de error o dominio en venta
                                    contenido = response.text.lower()
                                    
                                    # Palabras que indican páginas no válidas
                                    palabras_prohibidas = [
                                        'domain for sale', 'dominio en venta', 'this domain is for sale',
                                        'domain parking', 'parked domain', 'dominio aparcado',
                                        'page not found', 'página no encontrada', '404 error',
                                        'under construction', 'en construcción', 'coming soon',
                                        'temporarily unavailable', 'temporalmente no disponible',
                                        'server error', 'error del servidor', '500 error',
                                        'access denied', 'acceso denegado', 'forbidden',
                                        'expired domain', 'dominio expirado', 'domain expired'
                                    ]
                                    
                                    es_pagina_valida = True
                                    for palabra in palabras_prohibidas:
                                        if palabra in contenido:
                                            es_pagina_valida = False
                                            self.log(f"    ❌ Página no válida (contiene '{palabra}'): {href}")
                                            break
                                    
                                    if es_pagina_valida and len(contenido) > 200:  # Contenido mínimo
                                        competidores.append(href)
                                        self.log(f"    ✅ Competidor válido (intento {intentos}): {href}")
                                        if len(competidores) >= 2:
                                            break
                                    else:
                                        self.log(f"    ❌ Página con poco contenido: {href}")
                                else:
                                    self.log(f"    ❌ Página no accesible (código {response.status_code}): {href}")
                                    
                            except requests.exceptions.Timeout:
                                self.log(f"    ❌ Timeout en: {href}")
                            except requests.exceptions.ConnectionError:
                                self.log(f"    ❌ Error de conexión en: {href}")
                            except Exception as e:
                                self.log(f"    ❌ Error verificando: {href} - {str(e)[:30]}")
                                continue
                        elif href in competidores:
                            self.log(f"    ⚠️ Ya existe: {href}")
                    except Exception as e:
                        self.log(f"    ❌ Error procesando enlace: {str(e)[:30]}")
                        continue
                
                if len(competidores) >= 2:
                    break
                    
                time.sleep(2)  # Pausa entre intentos
            
            if len(competidores) >= 2:
                self.log(f"    ✅ Se encontraron {len(competidores)} competidores después de {intentos} intentos")
                return competidores[:2]
            else:
                self.log(f"    ❌ No se pudieron encontrar 2 competidores después de {max_intentos} intentos")
                return competidores  # Devolver los que se encontraron
                
        except Exception as e:
            self.log(f"    ❌ Error en búsqueda persistente: {str(e)[:30]}")
            return competidores_actuales

    def seleccionar_mejores_competidores_ia(self, competidores, keywords, url_cliente):
        """Usa IA para seleccionar los 2 mejores competidores de la lista"""
        try:
            if len(competidores) <= 2:
                return competidores
            
            # Crear prompt para la IA
            keywords_texto = ", ".join(keywords) if keywords else "servicios empresariales"
            
            prompt = f"""
Analiza estos {len(competidores)} competidores potenciales y selecciona los 2 MÁS RELEVANTES para una empresa que ofrece: {keywords_texto}

URL del cliente: {url_cliente}

Competidores encontrados:
{chr(10).join([f"{i+1}. {comp}" for i, comp in enumerate(competidores)])}

Criterios de selección:
1. Relevancia del sector (más similar al cliente)
2. Tamaño y profesionalidad del sitio web
3. Servicios similares ofrecidos
4. Posicionamiento en el mercado

Responde SOLO con los números de los 2 mejores competidores separados por comas.
Ejemplo: 1, 3
"""
            
            self.log(f"    🤖 Enviando {len(competidores)} competidores a IA para selección...")
            
            # Llamar a la IA
            respuesta_ia = consultar_ia_local(prompt)
            
            if respuesta_ia:
                # Extraer números de la respuesta
                numeros = []
                for palabra in respuesta_ia.split():
                    if palabra.replace(',', '').isdigit():
                        numeros.append(int(palabra.replace(',', '')))
                
                # Seleccionar los competidores
                competidores_seleccionados = []
                for num in numeros[:2]:  # Máximo 2
                    if 1 <= num <= len(competidores):
                        competidores_seleccionados.append(competidores[num-1])
                        self.log(f"    ✓ Competidor seleccionado por IA: {competidores[num-1]}")
                
                if len(competidores_seleccionados) == 2:
                    # Validar que los competidores sean páginas reales y operativas
                    competidores_validados = self.validar_competidores_operativos(competidores_seleccionados)
                    self.log(f"    ✅ IA seleccionó {len(competidores_validados)} competidores válidos")
                    return competidores_validados
                else:
                    self.log(f"    ⚠️ IA no pudo seleccionar correctamente, usando los primeros 2")
                    return competidores[:2]
            else:
                self.log(f"    ⚠️ Error en IA, usando los primeros 2 competidores")
                return competidores[:2]
                
        except Exception as e:
            self.log(f"    ❌ Error en selección IA: {str(e)[:30]}")
            return competidores[:2]

    def validar_competidores_operativos(self, competidores):
        """Valida que los competidores sean páginas reales y operativas"""
        try:
            competidores_validos = []
            
            for competidor in competidores:
                self.log(f"    🔍 Validando competidor: {competidor}")
                
                try:
                    import requests
                    from urllib.parse import urlparse
                    
                    # Configurar headers para parecer un navegador real
                    headers = {
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                        'Accept-Language': 'es-ES,es;q=0.8,en-US;q=0.5,en;q=0.3',
                        'Accept-Encoding': 'gzip, deflate',
                        'Connection': 'keep-alive',
                        'Upgrade-Insecure-Requests': '1'
                    }
                    
                    # Hacer petición con timeout
                    response = requests.get(competidor, headers=headers, timeout=10, allow_redirects=True)
                    
                    # Verificar código de respuesta
                    if response.status_code == 200:
                        # Verificar que no sea una página de error o dominio en venta
                        contenido = response.text.lower()
                        
                        # Palabras que indican páginas no válidas
                        palabras_prohibidas = [
                            'domain for sale', 'dominio en venta', 'this domain is for sale',
                            'domain parking', 'parked domain', 'dominio aparcado',
                            'page not found', 'página no encontrada', '404 error',
                            'under construction', 'en construcción', 'coming soon',
                            'temporarily unavailable', 'temporalmente no disponible',
                            'server error', 'error del servidor', '500 error',
                            'access denied', 'acceso denegado', 'forbidden',
                            'expired domain', 'dominio expirado', 'domain expired'
                        ]
                        
                        es_valida = True
                        for palabra in palabras_prohibidas:
                            if palabra in contenido:
                                self.log(f"    ❌ Competidor inválido (contiene '{palabra}'): {competidor}")
                                es_valida = False
                                break
                        
                        if es_valida:
                            # Verificar que tenga contenido mínimo (más de 500 caracteres)
                            if len(contenido) > 500:
                                # Verificar que no sea solo una página de "bienvenido" o similar
                                if any(word in contenido for word in ['welcome', 'bienvenido', 'hello', 'hola']):
                                    if len(contenido) < 1000:  # Páginas muy básicas
                                        self.log(f"    ❌ Competidor muy básico: {competidor}")
                                        continue
                                
                                competidores_validos.append(competidor)
                                self.log(f"    ✅ Competidor válido: {competidor}")
                            else:
                                self.log(f"    ❌ Competidor con poco contenido: {competidor}")
                        else:
                            continue
                    else:
                        self.log(f"    ❌ Competidor no accesible (código {response.status_code}): {competidor}")
                        
                except requests.exceptions.Timeout:
                    self.log(f"    ❌ Competidor con timeout: {competidor}")
                except requests.exceptions.ConnectionError:
                    self.log(f"    ❌ Competidor no accesible (error de conexión): {competidor}")
                except requests.exceptions.RequestException as e:
                    self.log(f"    ❌ Error validando competidor: {competidor} - {str(e)[:30]}")
                except Exception as e:
                    self.log(f"    ❌ Error inesperado validando: {competidor} - {str(e)[:30]}")
            
            # Si no tenemos suficientes competidores válidos, devolver los que tenemos
            if len(competidores_validos) < 2:
                self.log(f"    ⚠️ Solo {len(competidores_validos)} competidores válidos encontrados")
                if len(competidores_validos) == 0:
                    self.log(f"    ❌ No se encontraron competidores válidos")
                    return []
                else:
                    self.log(f"    ✓ Usando {len(competidores_validos)} competidores válidos disponibles")
                    return competidores_validos
            
            return competidores_validos[:2]  # Máximo 2
            
        except Exception as e:
            self.log(f"    ❌ Error en validación: {str(e)[:30]}")
            return competidores[:2]

    def generar_competidores_automaticos(self, driver, keywords, url):
        """Genera competidores automáticamente usando las keywords"""
        try:
            competidores = []
            
            # Usar las keywords para buscar competidores reales
            if keywords:
                # Estrategia simplificada: buscar directamente empresas del sector
                keyword_principal = keywords[0]
                self.log(f"    🔍 Buscando competidores para: {keyword_principal}")
                
                # Usar una búsqueda más simple y directa
                search_terms = f"{keyword_principal} empresa servicios"
                search_url = f"https://www.bing.com/search?q={search_terms.replace(' ', '+')}"
                
                self.log(f"    🔗 URL de búsqueda: {search_url}")
                driver.get(search_url)
                time.sleep(4)
                
                # Extraer URLs de los resultados
                try:
                    # Intentar múltiples selectores para encontrar enlaces (DuckDuckGo)
                    selectores = [
                        'a[data-testid="result-title-a"]',
                        'a[href*="http"]',
                        'h2 a[href*="http"]',
                        'div[data-testid="result"] a[href*="http"]',
                        'a[href*="http"]:not([href*="duckduckgo.com"])',
                        'h2 a',
                        'a'
                    ]
                    
                    result_links = []
                    for selector in selectores:
                        try:
                            links = driver.find_elements(By.CSS_SELECTOR, selector)
                            if links:
                                result_links.extend(links)
                                self.log(f"    ✓ Encontrados {len(links)} enlaces con selector: {selector}")
                                break
                        except:
                            continue
                    
                    if not result_links:
                        # Fallback: buscar todos los enlaces
                        result_links = driver.find_elements(By.TAG_NAME, 'a')
                        self.log(f"    ⚠️ Usando fallback: {len(result_links)} enlaces totales")
                    
                    for link in result_links[:15]:  # Primeros 15 resultados
                        try:
                            href = link.get_attribute('href')
                            if href and 'duckduckgo.com' not in href and 'duckduckgo' not in href.lower() and 'youtube.com' not in href and 'facebook.com' not in href and 'play.google.com' not in href and 'apps.apple.com' not in href and 'itunes.apple.com' not in href and 'playstore' not in href.lower() and 'appstore' not in href.lower():
                                # Limpiar URL
                                if '/url?q=' in href:
                                    href = href.split('/url?q=')[1].split('&')[0]
                                if href.startswith('http') and href not in competidores:
                                    # Excluir la URL del cliente
                                    if url not in href and href not in url:
                                        competidores.append(href)
                                        self.log(f"    ✓ Competidor encontrado: {href}")
                                        if len(competidores) >= 2:  # Solo necesitamos 2
                                            break
                        except:
                            continue
                            
                except Exception as e:
                    self.log(f"    ⚠️ Error extrayendo resultados: {str(e)[:30]}")
            
            self.log(f"    ✓ Encontrados {len(competidores)} competidores")
            
            # Log detallado de competidores generados automáticamente
            if competidores:
                self.log(f"    📋 Competidores generados automáticamente:")
                for i, competidor in enumerate(competidores, 1):
                    self.log(f"      {i}. {competidor}")
            else:
                self.log(f"    ⚠️ No se pudieron generar competidores automáticamente")
            
            return competidores
            
        except Exception as e:
            self.log(f"    ❌ Error generando competidores: {str(e)[:30]}")
            return []


    def generar_pdf_titularidad_dominio(self, screenshots, url_web, nombre_web, timestamp):
        """Genera PDF de titularidad del dominio con estilo igual al de justificación"""
        try:
            if not screenshots:
                self.log("  ❌ No hay screenshots")
                return None
            
            pdf_filename = f'Titularidad_Dominio_{timestamp}.pdf'
            pdf_path = os.path.join(self.pdfs_dir, pdf_filename)
            
            # Crear PDF
            c = canvas.Canvas(pdf_path, pagesize=A4)
            page_width, page_height = A4
            
            # Página 1: Portada con estilo igual al de justificación
            # Fondo blanco
            c.setFillColorRGB(1, 1, 1)
            c.rect(0, 0, page_width, page_height, fill=1, stroke=0)
            
            # Línea de acento superior
            c.setFillColorRGB(0.2, 0.45, 0.75)
            c.rect(0, page_height - 3, page_width, 3, fill=1, stroke=0)
            
            # Título principal
            c.setFillColorRGB(0.1, 0.1, 0.1)
            c.setFont("Helvetica-Bold", 20)
            c.drawString(50, page_height - 50, "SITIO WEB Y PRESENCIA BÁSICA EN INTERNET")
            
            # Subtítulo
            c.setFillColorRGB(0.2, 0.2, 0.2)
            c.setFont("Helvetica-Bold", 14)
            c.drawString(50, page_height - 75, "1. Mantenimiento del Dominio")
            
            # Texto descriptivo
            c.setFillColorRGB(0.3, 0.3, 0.3)
            c.setFont("Helvetica", 12)
            c.drawString(50, page_height - 105, "Verificación de titularidad del dominio mediante capturas de pantalla")
            c.drawString(50, page_height - 125, "de los paneles de control de los registradores correspondientes.")
            
            # Información del sitio
            c.setFillColorRGB(0.1, 0.1, 0.1)
            c.setFont("Helvetica-Bold", 12)
            c.drawString(50, page_height - 160, f"Sitio Web: {nombre_web if nombre_web else 'N/A'}")
            c.drawString(50, page_height - 180, f"URL: {url_web}")
            c.drawString(50, page_height - 200, f"Fecha de análisis: {datetime.now().strftime('%d/%m/%Y a las %H:%M')}")
            
            # Pie de página
            c.setFillColorRGB(0.5, 0.5, 0.5)
            c.setFont("Helvetica", 7)
            c.drawCentredString(page_width/2, 30, f"Justificado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            
            c.showPage()
            
            # Páginas de screenshots con formato igual al de justificación
            for i, screenshot_path in enumerate(screenshots):
                if os.path.exists(screenshot_path):
                    try:
                        # Fondo blanco
                        c.setFillColorRGB(1, 1, 1)
                        c.rect(0, 0, page_width, page_height, fill=1, stroke=0)
                        
                        # Línea de acento superior
                        c.setFillColorRGB(0.2, 0.45, 0.75)
                        c.rect(0, page_height - 3, page_width, 3, fill=1, stroke=0)
                        
                        # Título de la sección
                        c.setFillColorRGB(0.1, 0.1, 0.1)
                        c.setFont("Helvetica-Bold", 16)
                        titulo_captura = f"EVIDENCIA {i+1}: {self.obtener_descripcion_captura(screenshot_path)}"
                        c.drawString(50, page_height - 50, titulo_captura)
                        
                        # Línea decorativa
                        c.setStrokeColorRGB(0.2, 0.4, 0.8)
                        c.setLineWidth(1)
                        c.line(50, page_height - 70, page_width - 50, page_height - 70)
                        
                        # Cargar imagen
                        img = Image.open(screenshot_path)
                        img_width, img_height = img.size
                        
                        # Calcular dimensiones para ajustar a la página
                        max_width = page_width - 100
                        max_height = page_height - 200
                        
                        scale_x = max_width / img_width
                        scale_y = max_height / img_height
                        scale = min(scale_x, scale_y)
                        
                        new_width = img_width * scale
                        new_height = img_height * scale
                        
                        # Centrar imagen
                        x = (page_width - new_width) / 2
                        y = (page_height - new_height) / 2 - 50
                        
                        # Marco decorativo
                        c.setStrokeColorRGB(0.8, 0.8, 0.8)
                        c.setLineWidth(2)
                        c.rect(x - 10, y - 10, new_width + 20, new_height + 20)
                        
                        # Agregar imagen
                        c.drawImage(screenshot_path, x, y, width=new_width, height=new_height)
                        
                        # Descripción de la captura
                        c.setFillColorRGB(0.1, 0.1, 0.1)
                        c.setFont("Helvetica", 10)
                        descripcion = f"Archivo: {os.path.basename(screenshot_path)}"
                        c.drawString(50, y - 30, descripcion)
                        
                        # Pie de página
                        c.setFillColorRGB(0.5, 0.5, 0.5)
                        c.setFont("Helvetica", 7)
                        c.drawCentredString(page_width/2, 30, f"Justificado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                        
                        c.showPage()
                        
                    except Exception as e:
                        self.log(f"  ✗ Error screenshot {i+1}: {str(e)[:30]}")
            
            c.save()
            self.log(f"  ✓ PDF de titularidad generado: {pdf_filename}")
            return pdf_filename
            
        except Exception as e:
            self.log(f"  ❌ Error generando PDF de titularidad: {str(e)[:30]}")
            return None
    
    def obtener_descripcion_captura(self, screenshot_path):
        """Obtiene una descripción legible del nombre del archivo de captura"""
        filename = os.path.basename(screenshot_path)
        
        if "ionos" in filename.lower():
            if "completa" in filename.lower():
                return "Panel de control completo - Ionos"
            elif "especifica" in filename.lower():
                return "Sección específica - Ionos"
            else:
                return "Panel de control - Ionos"
        elif "dondominio" in filename.lower():
            if "completa" in filename.lower():
                return "Panel de control completo - Dondominio"
            elif "final" in filename.lower():
                return "Página final - Dondominio"
            else:
                return "Panel de control - Dondominio"
        else:
            return "Captura de pantalla del panel de control"

    def generar_pdf_publicidad(self, fullpage_screenshot, url_web, timestamp):
        """Genera PDF de publicidad que solo contiene la captura de la web del cliente (scroll al máximo abajo) sin título ni texto"""
        try:
            if not fullpage_screenshot or not os.path.exists(fullpage_screenshot):
                self.log("  ❌ No hay captura de página completa para generar PDF de publicidad")
                return None
            
            pdf_filename = f'Publicidad_{timestamp}.pdf'
            pdf_path = os.path.join(self.pdfs_dir, pdf_filename)
            
            # Crear PDF
            c = canvas.Canvas(pdf_path, pagesize=A4)
            page_width, page_height = A4
            
            # Solo agregar la imagen sin título ni texto
            try:
                # Cargar imagen
                img = Image.open(fullpage_screenshot)
                img_width, img_height = img.size
                
                # Calcular dimensiones para ajustar a la página
                max_width = page_width - 20  # Márgenes mínimos
                max_height = page_height - 20
                
                scale_x = max_width / img_width
                scale_y = max_height / img_height
                scale = min(scale_x, scale_y)
                
                new_width = img_width * scale
                new_height = img_height * scale
                
                # Centrar imagen
                x = (page_width - new_width) / 2
                y = (page_height - new_height) / 2
                
                # Agregar imagen
                c.drawImage(fullpage_screenshot, x, y, width=new_width, height=new_height)
                
            except Exception as e:
                self.log(f"  ❌ Error cargando imagen: {str(e)}")
                return None
            
            c.save()
            self.log(f"  ✓ PDF de publicidad generado: {pdf_filename}")
            return pdf_filename
            
        except Exception as e:
            self.log(f"  ❌ Error generando PDF de publicidad: {str(e)[:30]}")
            return None


if __name__ == "__main__":
    root = tk.Tk()
    app = SEOAnalyzerGUI(root)
    root.mainloop()